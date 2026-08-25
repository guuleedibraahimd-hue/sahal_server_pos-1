from flask import (
    Flask,
    render_template,
    request,
    redirect,
    jsonify,
    session,
    url_for,
    flash,
    send_from_directory,
    Response,
    make_response
)

from flask_socketio import (
    SocketIO,
    emit,
    join_room
)

from werkzeug.utils import secure_filename

from werkzeug.security import (
    generate_password_hash,
    check_password_hash
)

import sqlite3
import os
import qrcode
import socket
import random
import json
import re
import time
from urllib.parse import quote

from zoneinfo import ZoneInfo
from datetime import datetime, timedelta, timezone

import firebase_admin
from firebase_admin import credentials, firestore, storage

from io import BytesIO
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.enums import TA_CENTER
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import requests

app = Flask(__name__)

# ==========================================
# 💳 SYSTEM RENEWAL PAYMENT NUMBER
# The number account owners are told to send subscription
# renewal payments to. Shown on every suspended dashboard.
# ==========================================
ADMIN_PAYMENT_NUMBER = "618276993"

@app.route('/manifest.json')
def serve_manifest():
    return send_from_directory('static', 'manifest.json')

@app.route('/sw.js')
def serve_sw():
    return send_from_directory('static', 'sw.js')

# =========================
# 🚀 FLASK APP
# =========================

app = Flask(
    __name__,
    static_url_path="/static",
    static_folder="static",
    template_folder="templates"
)

# =========================
# 🔐 SECRET KEY
# =========================

app.secret_key = "supersecretkey123"

# =========================
# 🔌 SOCKET IO
# polling keliya — Railway WebSocket proxy dhibaato buu leeyahay
# =========================

socketio = SocketIO(
    app,
    cors_allowed_origins="*",
    async_mode="threading",
    allow_upgrades=False,
    transports=["polling"]
)

# =========================
# 🔥 SAHAL SERVER FIREBASE
# =========================

firebase_key_str = os.environ.get("FIREBASE_KEY")
firebase_key = json.loads(firebase_key_str)
cred1 = credentials.Certificate(firebase_key)

# Storage bucket for uploaded files (menu photos, ad images/audio, etc).
# Local disk on most hosts (Railway included) is EPHEMERAL — uploaded files
# vanish on every restart/redeploy even though Firestore still remembers
# their filename. Firebase Storage keeps them permanently instead.
# Override with the FIREBASE_STORAGE_BUCKET env var if your bucket name
# doesn't follow the default "<project-id>.appspot.com" pattern (newer
# Firebase projects sometimes use "<project-id>.firebasestorage.app").
SAHAL_STORAGE_BUCKET = os.environ.get(
    "FIREBASE_STORAGE_BUCKET",
    f"{firebase_key.get('project_id', '')}.appspot.com"
)

sahal_app = firebase_admin.initialize_app(
    cred1,
    {"storageBucket": SAHAL_STORAGE_BUCKET},
    name="sahal_app"
)

db = firestore.client(sahal_app)  # ✅ hal mar oo kaliya

# =========================
# 💎 DHIBIC DAHAB FIREBASE
# =========================

dhibic_key_str = os.environ.get("DHIBIC_FIREBASE_KEY")
dhibic_key = json.loads(dhibic_key_str)
cred2 = credentials.Certificate(dhibic_key)

dhibic_app = firebase_admin.initialize_app(
    cred2,
    name="dhibic_app"
)

dhibic_db = firestore.client(dhibic_app)  # ✅ hal mar oo kaliya
# =========================
# 📁 FOLDERS
# =========================

UPLOAD_FOLDER = "static/uploads"
QR_FOLDER = "static/qr"

os.makedirs(
    UPLOAD_FOLDER,
    exist_ok=True
)

os.makedirs(
    QR_FOLDER,
    exist_ok=True
)

app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER

# =========================
# ☁️ FIREBASE STORAGE UPLOAD HELPER
# Waxay bedeshaa kaydinta faylasha (disk-ka ephemeral-ka ah) una wareejisaa
# Firebase Storage — sidaas faylashu marnaba kuma lumi doonaan deploy/restart.
# =========================
def upload_to_firebase_storage(file_obj, folder="uploads"):
    """Upload a Flask FileStorage object to Firebase Storage and return its
    public URL. Returns "" if file_obj is empty/missing."""
    if not file_obj or not file_obj.filename:
        return ""
    safe_name = secure_filename(file_obj.filename)
    blob_path = f"{folder}/{int(time.time() * 1000)}_{safe_name}"
    bucket = storage.bucket(app=sahal_app)
    blob = bucket.blob(blob_path)
    blob.upload_from_file(file_obj.stream, content_type=file_obj.content_type)
    blob.make_public()
    return blob.public_url


def delete_from_firebase_storage(public_url):
    """Best-effort delete of a blob given its public URL. Safe to call with
    a plain local filename too — it just won't match anything and no-ops."""
    try:
        if not public_url or not public_url.startswith("http"):
            return
        bucket = storage.bucket(app=sahal_app)
        # public_url looks like https://storage.googleapis.com/<bucket>/<blob_path>
        marker = f"{bucket.name}/"
        idx = public_url.find(marker)
        if idx == -1:
            return
        blob_path = public_url[idx + len(marker):]
        bucket.blob(blob_path).delete()
    except Exception:
        pass
# =========================
# DATABASE PATH
# =========================
DB_PATH = os.environ.get("DB_PATH", "database.db")

# =========================
# INIT DATABASE
# =========================
def init_db():
    print("INIT DB RUNNING...")

    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()

    c.execute("""
    CREATE TABLE IF NOT EXISTS orders (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        restaurant_id TEXT,
        table_no TEXT,
        food TEXT,
        price REAL,
        qty INTEGER DEFAULT 1,
        total REAL,
        time TEXT DEFAULT CURRENT_TIMESTAMP,
        status TEXT DEFAULT 'pending'
    )
    """)

    c.execute("""
    CREATE INDEX IF NOT EXISTS idx_orders
    ON orders(restaurant_id, table_no)
    """)

    c.execute("""
    CREATE TABLE IF NOT EXISTS restaurants(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT,
        phone TEXT,
        username TEXT,
        password TEXT,
        price INTEGER,
        expiry TEXT,
        active INTEGER,
        payment_number TEXT,
        kitchen_password TEXT
    )
    """)

    c.execute("""
    CREATE TABLE IF NOT EXISTS settings(
        id INTEGER PRIMARY KEY,
        admin_password TEXT,
        register_password TEXT
    )
    """)

    c.execute("SELECT id FROM settings WHERE id=1")
    if not c.fetchone():
        c.execute("""
            INSERT INTO settings
            (id, admin_password, register_password)
            VALUES (1, '8880', '8880')
        """)

    c.execute("""
    CREATE TABLE IF NOT EXISTS supermarkets(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT,
        username TEXT,
        password TEXT,
        price INTEGER,
        expiry TEXT,
        active INTEGER DEFAULT 1
    )
    """)

    c.execute("""
    CREATE TABLE IF NOT EXISTS supermarket_products(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        barcode TEXT UNIQUE,
        product_name TEXT,
        price REAL
    )
    """)

    c.execute("""
    CREATE TABLE IF NOT EXISTS supermarket_orders(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        receipt_no TEXT,
        total REAL,
        created_at TEXT
    )
    """)

    conn.commit()
    conn.close()

    print("DATABASE READY ✅")

init_db()

# =========================
# 🇸🇴 SOMALIA TIME
# =========================
def somalia_time():
    return datetime.now(timezone(timedelta(hours=3)))

def get_somali_time():
    return datetime.now(timezone.utc) + timedelta(hours=3)

# =========================
# 🌐 GET SERVER IP
# =========================
def get_ip():
    s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
    try:
        s.connect(("8.8.8.8", 80))
        ip = s.getsockname()[0]
    finally:
        s.close()
    return ip

SERVER_IP = get_ip()

# =========================
# 🎤 WEBRTC SIGNALING EVENTS
# =========================
@socketio.on("voice_call")
def voice_call(data):
    emit("incoming_call", data, broadcast=True)

@socketio.on("offer")
def handle_offer(data):
    emit("offer", data, broadcast=True)

@socketio.on("answer")
def handle_answer(data):
    emit("answer", data, broadcast=True)

@socketio.on("ice_candidate")
def handle_ice(data):
    emit("ice_candidate", data, broadcast=True)

# =========================
# 🏠 SOCKET ROOM JOIN
# =========================
@socketio.on("join_customer_room")
def join_customer(data):
    room = f"{data['rid']}_{data['table']}"
    join_room(room)
    emit("joined_room", {"room": room})

@socketio.on("join_kitchen_room")
def join_kitchen(data):
    room = f"kitchen_{data['rid']}"
    join_room(room)
    emit("joined_kitchen", {"room": room})

# =========================
# 🔐 SYSTEM PASSWORDS FROM FIREBASE
# =========================
def get_system_passwords():
    try:
        doc_ref = db.collection("evote").document("system")
        doc = doc_ref.get()

        if doc.exists:
            return doc.to_dict()

        return {
            "admin_password": "6993",
            "register_password": "6993"
        }

    except Exception as e:
        print("Firebase password error:", e)
        return {
            "admin_password": "6993",
            "register_password": "6993"
        }

# =========================
# ⏰ AUTO CHECK EXPIRY
# =========================
def auto_check_expiry(rid):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    try:
        c.execute("SELECT expiry, active FROM restaurants WHERE id=?", (rid,))
        row = c.fetchone()
        if row and row[0]:
            expiry = datetime.strptime(row[0], "%Y-%m-%d")
            if datetime.now() >= expiry:
                c.execute("UPDATE restaurants SET active=0 WHERE id=?", (rid,))
                conn.commit()
    except Exception as e:
        print("Auto Expiry Error:", e)
    conn.close()

# =========================
# 🔥 FIRESTORE FUNCTIONS
# =========================
def get_restaurants_firestore():
    restaurants = []
    try:
        docs = db.collection("restaurants").stream()
        for doc in docs:
            item = doc.to_dict()
            item["id"] = doc.id
            item["active"] = item.get("active", False)
            item["name"] = item.get("name", "N/A")
            item["phone"] = item.get("phone", "N/A")
            item["username"] = item.get("username", "N/A")
            item["kitchen_password"] = item.get("kitchen_password", "N/A")
            item["password"] = item.get("password", "N/A")
            item["expiry"] = item.get("expiry", "N/A")
            restaurants.append(item)
    except Exception as e:
        print("Restaurant Load Error:", e)
    return restaurants


def get_supermarkets_firestore():
    supermarkets = []
    try:
        docs = db.collection("supermarkets").stream()
        for doc in docs:
            item = doc.to_dict()
            item["id"] = doc.id
            item["active"] = item.get("active", False)
            item["name"] = item.get("name", "N/A")
            item["username"] = item.get("username", "N/A")
            item["expiry"] = item.get("expiry", "N/A")
            supermarkets.append(item)
    except Exception as e:
        print("Supermarket Load Error:", e)
    return supermarkets


def get_pharmacies_list():
    pharmacies = []
    try:
        today = datetime.now().date()
        docs = db.collection("pharmacies").stream()
        for doc in docs:
            item = doc.to_dict()
            item["id"] = doc.id
            expiry_date = item.get("expiry_date")
            if expiry_date:
                try:
                    item["days_left"] = (datetime.strptime(expiry_date, "%Y-%m-%d").date() - today).days
                except Exception:
                    pass
            pharmacies.append(item)
    except Exception as e:
        print("Pharmacy Load Error:", e)
    return pharmacies


def get_renewal_requests():
    """Pending renewal payment claims submitted from suspended dashboards
    (restaurant / supermarket / pharmacy) for the admin to review."""
    reqs = []
    try:
        docs = db.collection("renewal_requests").where("status", "==", "pending").stream()
        for doc in docs:
            item = doc.to_dict()
            item["id"] = doc.id
            reqs.append(item)
        reqs.sort(key=lambda x: x.get("submitted_at", ""), reverse=True)
    except Exception as e:
        print("Renewal Requests Load Error:", e)
    return reqs


# ==========================================
# 💳 SUBMIT RENEWAL PAYMENT CLAIM
# Called from the suspended dashboard (restaurant / supermarket /
# pharmacy) once the owner has sent payment to ADMIN_PAYMENT_NUMBER.
# Public route — no admin login required, only a valid entity_id.
# ==========================================
RENEWAL_COLLECTION_MAP  = {"restaurant": "restaurants", "supermarket": "supermarkets", "pharmacy": "pharmacies"}
RENEWAL_NAME_FIELD_MAP  = {"restaurant": "name", "supermarket": "name", "pharmacy": "pharmacy_name"}

@app.route("/submit_renewal/<entity_type>/<entity_id>", methods=["POST"])
def submit_renewal(entity_type, entity_id):
    try:
        if entity_type not in RENEWAL_COLLECTION_MAP:
            return jsonify({"success": False, "error": "Invalid account type"}), 400

        data              = request.get_json() or {}
        sender_name       = data.get("sender_name", "").strip()
        paid_from_number  = data.get("paid_from_number", "").strip()

        if not sender_name or not paid_from_number:
            return jsonify({"success": False, "error": "Fill all fields"})

        entity_doc = db.collection(RENEWAL_COLLECTION_MAP[entity_type]).document(entity_id).get()
        if not entity_doc.exists:
            return jsonify({"success": False, "error": "Account not found"})

        business_name = entity_doc.to_dict().get(RENEWAL_NAME_FIELD_MAP[entity_type], entity_id)

        db.collection("renewal_requests").add({
            "entity_type":      entity_type,
            "entity_id":        entity_id,
            "business_name":    business_name,
            "sender_name":      sender_name,
            "paid_from_number": paid_from_number,
            "submitted_at":     datetime.now().isoformat(),
            "status":           "pending"
        })
        return jsonify({"success": True, "message": "Renewal request sent"})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


# ==========================================
# 💳 PUSH-PAYMENT RENEWAL (Hormuud EVC Plus via Waafi)
# The owner enters their phone number and picks how many months; this
# calls Waafi's MWALLET_ACCOUNT push flow with the account's real
# merchant credentials. Waafi/Hormuud then send the PIN prompt
# straight to the OWNER'S OWN PHONE via their own system — this app
# never sees, asks for, or handles the PIN at any point. The
# subscription is only extended and reactivated once Waafi confirms
# the payment actually succeeded.
# ==========================================
WAAFI_MERCHANT_UID = "M0914174"
WAAFI_API_USER_ID  = "1008694"
WAAFI_API_KEY       = "API-QMfqbsf1V6qFSxyQgQ2Nbq3DjHoF"
WAAFI_URL           = "https://api.waafipay.net/asm"


def _monthly_fee_for(entity_type, entity_data):
    if entity_type == "pharmacy":
        return float(entity_data.get("monthly_fee", 0) or 0)
    try:
        return float(entity_data.get("monthly_fee") or entity_data.get("price") or entity_data.get("fee") or 0)
    except (TypeError, ValueError):
        return 0.0


# ==========================================
# ✏️ ADMIN — EDIT RESTAURANT / SUPERMARKET
# Full edit: name, username, all passwords, phone/admin info, and the
# monthly fee (used everywhere renew/push-payment computes an amount —
# changing it here is the one place that "free" account can be given
# a real price, or an existing price corrected).
# ==========================================
EDITABLE_ENTITY_COLLECTIONS = {"restaurant": "restaurants", "supermarket": "supermarkets"}

RESTAURANT_EDIT_FIELDS = [
    "name", "phone", "username", "password",
    "admin_name", "admin_email",
    "restaurant_admin_password", "kitchen_password",
    "monthly_fee", "payment"
]
SUPERMARKET_EDIT_FIELDS = [
    "name", "username", "password", "monthly_fee", "payment"
]


@app.route("/admin/edit_entity/<entity_type>/<entity_id>", methods=["GET", "POST"])
def admin_edit_entity(entity_type, entity_id):
    if not session.get("admin_ok"):
        return jsonify({"success": False, "error": "Unauthorized"}), 401
    if entity_type not in EDITABLE_ENTITY_COLLECTIONS:
        return jsonify({"success": False, "error": "Invalid account type"}), 400

    collection = EDITABLE_ENTITY_COLLECTIONS[entity_type]
    entity_ref = db.collection(collection).document(entity_id)
    entity_doc = entity_ref.get()
    if not entity_doc.exists:
        return jsonify({"success": False, "error": "Account not found"}), 404

    if request.method == "GET":
        data = entity_doc.to_dict()
        data["id"] = entity_id
        data["monthly_fee"] = _monthly_fee_for(entity_type, data)
        return jsonify({"success": True, "data": data})

    try:
        body = request.get_json() or {}
        fields = RESTAURANT_EDIT_FIELDS if entity_type == "restaurant" else SUPERMARKET_EDIT_FIELDS
        update_fields = {}

        for field in fields:
            if field not in body:
                continue
            value = body[field]
            if isinstance(value, str):
                value = value.strip()
            if field == "monthly_fee":
                try:
                    update_fields["monthly_fee"] = float(value)
                except (TypeError, ValueError):
                    return jsonify({"success": False, "error": "Monthly fee must be a number"})
            elif value != "":
                update_fields[field] = value

        if not update_fields:
            return jsonify({"success": False, "error": "Nothing to update"})

        # Legacy typo'd field some restaurants were created with — keep
        # it in sync so old templates reading it still show the update.
        if entity_type == "restaurant" and "restaurant_admin_password" in update_fields:
            update_fields["resturen_admin password"] = update_fields["restaurant_admin_password"]

        entity_ref.update(update_fields)
        updated = entity_ref.get().to_dict()
        updated["id"] = entity_id
        updated["monthly_fee"] = _monthly_fee_for(entity_type, updated)
        return jsonify({"success": True, "data": updated})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


# ==========================================
# 📢 ADMIN BROADCAST NOTIFICATIONS
# System admin can message: everyone, every account of one type, or
# one specific account — shown as a dismissable banner directly on
# that account's own dashboard (restaurant/restaurant-admin/supermarket/
# pharmacy). Dismissal is per-recipient, stored on the entity's own doc.
# ==========================================
BROADCAST_ENTITY_COLLECTIONS = {"restaurant": "restaurants", "supermarket": "supermarkets", "pharmacy": "pharmacies"}


@app.route("/admin/broadcasts")
def admin_list_broadcasts():
    if not session.get("admin_ok"):
        return jsonify({"success": False, "error": "Unauthorized"}), 401
    try:
        broadcasts = []
        for doc in db.collection("admin_broadcasts").order_by(
                "created_at", direction=firestore.Query.DESCENDING).limit(100).stream():
            b = doc.to_dict()
            b["id"] = doc.id
            broadcasts.append(b)
        return jsonify({"success": True, "broadcasts": broadcasts})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


@app.route("/admin/send_broadcast", methods=["POST"])
def admin_send_broadcast():
    if not session.get("admin_ok"):
        return jsonify({"success": False, "error": "Unauthorized"}), 401
    try:
        data        = request.get_json() or {}
        message     = data.get("message", "").strip()
        target_type = data.get("target_type", "").strip()
        target_id   = data.get("target_id", "").strip()

        if not message:
            return jsonify({"success": False, "error": "Write a message first"})
        if target_type not in ("everyone", "restaurant", "pharmacy", "supermarket"):
            return jsonify({"success": False, "error": "Invalid target"})

        target_name = ""
        if target_type != "everyone" and target_id:
            collection = BROADCAST_ENTITY_COLLECTIONS[target_type]
            name_field = "pharmacy_name" if target_type == "pharmacy" else "name"
            target_doc = db.collection(collection).document(target_id).get()
            if not target_doc.exists:
                return jsonify({"success": False, "error": "Selected account not found"})
            target_name = target_doc.to_dict().get(name_field, target_id)

        db.collection("admin_broadcasts").add({
            "message": message,
            "target_type": target_type,
            "target_id": target_id or "",
            "target_name": target_name,
            "created_at": datetime.now().isoformat()
        })
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


@app.route("/admin/delete_broadcast/<broadcast_id>", methods=["DELETE"])
def admin_delete_broadcast(broadcast_id):
    if not session.get("admin_ok"):
        return jsonify({"success": False, "error": "Unauthorized"}), 401
    try:
        db.collection("admin_broadcasts").document(broadcast_id).delete()
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


def _get_active_broadcasts(entity_type, entity_id):
    """Every broadcast aimed at this entity (targeted by name, targeted
    at 'all of this type', or sent to everyone) that this entity hasn't
    dismissed yet. Fetches the whole (small) broadcasts collection and
    filters in Python — avoids the composite-index requirement that
    mixing multiple equality/OR conditions in one Firestore query would
    trigger."""
    try:
        collection = BROADCAST_ENTITY_COLLECTIONS.get(entity_type)
        if not collection:
            return []
        entity_doc = db.collection(collection).document(entity_id).get()
        dismissed = set((entity_doc.to_dict() or {}).get("dismissed_broadcasts", [])) if entity_doc.exists else set()

        active = []
        for doc in db.collection("admin_broadcasts").stream():
            b = doc.to_dict()
            bid = doc.id
            if bid in dismissed:
                continue
            if b.get("target_type") == "everyone":
                active.append({**b, "id": bid})
            elif b.get("target_type") == entity_type and (not b.get("target_id") or b.get("target_id") == entity_id):
                active.append({**b, "id": bid})
        active.sort(key=lambda x: x.get("created_at", ""), reverse=True)
        return active
    except Exception:
        return []


@app.route("/dismiss_broadcast/<entity_type>/<entity_id>/<broadcast_id>", methods=["POST"])
def dismiss_broadcast(entity_type, entity_id, broadcast_id):
    collection = BROADCAST_ENTITY_COLLECTIONS.get(entity_type)
    if not collection:
        return jsonify({"success": False, "error": "Invalid account type"}), 400
    try:
        db.collection(collection).document(entity_id).update({
            "dismissed_broadcasts": firestore.ArrayUnion([broadcast_id])
        })
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


# ==========================================
# 📢 ADMIN BROADCAST MESSAGES
# Admin composes a message and targets it at: everyone, every account
# of one type, or one specific account — it then shows as a dismissible
# notification directly on that account's own dashboard(s). Restaurants
# have two dashboards (the menu/ads one and the analytics/staff one) —
# a message sent to a restaurant shows on both.
# ==========================================
BROADCAST_TARGET_TYPES = {"all", "restaurant", "supermarket", "pharmacy"}
BROADCAST_NAME_FIELD = {"restaurant": "name", "supermarket": "name", "pharmacy": "pharmacy_name"}


@app.route("/admin/send_message", methods=["POST"])
def admin_send_message():
    if not session.get("admin_ok"):
        return jsonify({"success": False, "error": "Unauthorized"}), 401
    try:
        data = request.get_json() or {}
        message     = data.get("message", "").strip()
        target_type = data.get("target_type", "all")
        target_id   = data.get("target_id", "").strip()

        if not message:
            return jsonify({"success": False, "error": "Message text is required"})
        if target_type not in BROADCAST_TARGET_TYPES:
            return jsonify({"success": False, "error": "Invalid target type"})

        target_label = "Everyone"
        if target_type != "all":
            if target_id:
                collection = RENEWAL_COLLECTION_MAP.get(target_type)
                doc = db.collection(collection).document(target_id).get()
                if not doc.exists:
                    return jsonify({"success": False, "error": "That account was not found"})
                target_label = doc.to_dict().get(BROADCAST_NAME_FIELD[target_type], target_id)
            else:
                target_label = f"All {target_type.capitalize()}s"

        db.collection("admin_messages").add({
            "message": message,
            "target_type": target_type,
            "target_id": target_id,
            "target_label": target_label,
            "created_at": datetime.now().isoformat()
        })
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


@app.route("/admin/broadcast_messages")
def admin_list_messages():
    if not session.get("admin_ok"):
        return jsonify({"success": False, "error": "Unauthorized"}), 401
    try:
        messages = []
        for doc in db.collection("admin_messages").stream():
            m = doc.to_dict()
            m["id"] = doc.id
            messages.append(m)
        messages.sort(key=lambda x: x.get("created_at", ""), reverse=True)
        return jsonify({"success": True, "messages": messages})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


@app.route("/admin/delete_message/<msg_id>", methods=["DELETE"])
def admin_delete_message(msg_id):
    if not session.get("admin_ok"):
        return jsonify({"success": False, "error": "Unauthorized"}), 401
    try:
        db.collection("admin_messages").document(msg_id).delete()
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


def _messages_for_entity(entity_type, entity_id):
    """Every broadcast that applies to this specific account — sent to
    everyone, sent to all accounts of this type, or sent to this one
    account by name. Filtered in Python (not a Firestore query) so no
    composite index is ever needed here."""
    relevant = []
    for doc in db.collection("admin_messages").stream():
        m = doc.to_dict()
        if m.get("target_type") == "all":
            relevant.append(m | {"id": doc.id})
        elif m.get("target_type") == entity_type:
            t_id = m.get("target_id", "")
            if not t_id or t_id == entity_id:
                relevant.append(m | {"id": doc.id})
    relevant.sort(key=lambda x: x.get("created_at", ""), reverse=True)
    return relevant[:20]


@app.route("/dashboard_messages/<entity_type>/<entity_id>")
def dashboard_messages(entity_type, entity_id):
    if entity_type not in BROADCAST_TARGET_TYPES or entity_type == "all":
        return jsonify({"success": False, "error": "Invalid type"}), 400
    try:
        return jsonify({"success": True, "messages": _messages_for_entity(entity_type, entity_id)})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


@app.route("/renew_push_payment/<entity_type>/<entity_id>", methods=["POST"])
def renew_push_payment(entity_type, entity_id):
    try:
        if entity_type not in RENEWAL_COLLECTION_MAP:
            return jsonify({"success": False, "error": "Invalid account type"}), 400

        data   = request.get_json() or {}
        phone  = data.get("phone", "").strip()
        months = int(data.get("months", 0) or 0)

        if not phone or months < 1:
            return jsonify({"success": False, "error": "Enter a phone number and choose at least 1 month"})

        clean_phone = re.sub(r'\D', '', phone)
        if clean_phone.startswith("252"):
            clean_phone = clean_phone[3:]
        elif clean_phone.startswith("0"):
            clean_phone = clean_phone[1:]
        if len(clean_phone) != 9:
            return jsonify({"success": False, "error": "Phone number must be 9 digits (e.g. 61XXXXXXX)"})

        collection = RENEWAL_COLLECTION_MAP[entity_type]
        entity_ref = db.collection(collection).document(entity_id)
        entity_doc = entity_ref.get()
        if not entity_doc.exists:
            return jsonify({"success": False, "error": "Account not found"})

        entity_data   = entity_doc.to_dict()
        business_name = entity_data.get(RENEWAL_NAME_FIELD_MAP[entity_type], entity_id)
        monthly_fee   = _monthly_fee_for(entity_type, entity_data)
        if monthly_fee <= 0:
            return jsonify({"success": False, "error": "This account has no subscription price set — contact support"})

        total_amount = round(monthly_fee * months, 2)
        reference_id = f"renew_{entity_id}_{int(datetime.now().timestamp())}"

        payload = {
            "schemaVersion": "1.0",
            "requestId": str(int(datetime.now().timestamp() * 1000)),
            "timestamp": datetime.now().isoformat(),
            "channelName": "WEB",
            "serviceName": "API_PURCHASE",
            "serviceParams": {
                "merchantUid": WAAFI_MERCHANT_UID,
                "apiUserId": WAAFI_API_USER_ID,
                "apiKey": WAAFI_API_KEY,
                "paymentMethod": "MWALLET_ACCOUNT",
                "payerInfo": {"accountNo": clean_phone},
                "transactionInfo": {
                    "referenceId": reference_id,
                    "invoiceId": reference_id,
                    "amount": f"{total_amount:.2f}",
                    "currency": "USD",
                    "description": f"Sahal Server subscription — {business_name} ({months} mo)"
                }
            }
        }

        try:
            waafi_res = requests.post(WAAFI_URL, json=payload, timeout=45)
            waafi_data = waafi_res.json()
        except Exception as e:
            return jsonify({"success": False, "error": f"Payment gateway error: {str(e)}"})

        response_code = str(waafi_data.get("responseCode", ""))
        succeeded = response_code in ("0", "2001")

        # Log every attempt (success or fail) for a real audit trail —
        # replaces the old "trust me, I paid" renewal_requests flow.
        db.collection("subscription_payments").add({
            "entity_type": entity_type, "entity_id": entity_id, "business_name": business_name,
            "phone": clean_phone, "months": months, "amount": total_amount,
            "reference_id": reference_id, "response_code": response_code,
            "waafi_response": waafi_data,
            "status": "SUCCESS" if succeeded else "FAILED",
            "created_at": datetime.now().isoformat()
        })

        if not succeeded:
            return jsonify({"success": False, "error": waafi_data.get("responseMsg", "Payment was declined or cancelled")})

        # Extend from whichever is later — today, or the current expiry
        # (if it's still in the future) — so renewing early never loses
        # already-paid time.
        current_expiry_str = entity_data.get("expiry_date") or entity_data.get("expiry")
        base_date = datetime.now()
        if current_expiry_str:
            try:
                current_expiry = datetime.strptime(current_expiry_str, "%Y-%m-%d")
                if current_expiry > base_date:
                    base_date = current_expiry
            except Exception:
                pass
        new_expiry = (base_date + timedelta(days=months * 30)).strftime("%Y-%m-%d")

        entity_ref.update({
            "active": True,
            "status": "active",
            "expiry_date": new_expiry,
            "expiry": new_expiry
        })

        return jsonify({"success": True, "expiry_date": new_expiry, "amount": total_amount})

    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


@app.route("/admin/dismiss_renewal/<req_id>", methods=["DELETE"])
def admin_dismiss_renewal(req_id):
    if not session.get("admin_ok"):
        return jsonify({"success": False, "error": "Unauthorized"}), 401
    try:
        db.collection("renewal_requests").document(req_id).delete()
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


def get_orders_firestore():
    orders = []
    try:
        docs = db.collection("orders").stream()
        for doc in docs:
            item = doc.to_dict()
            item["id"] = doc.id
            item["restaurant_name"] = item.get("restaurant_name", "N/A")
            item["food"] = item.get("food", "N/A")
            item["table"] = item.get("table", "N/A")
            item["time"] = item.get("time", "N/A")
            item["status"] = item.get("status", "Pending")
            orders.append(item)
    except Exception as e:
        print("Orders Load Error:", e)
    return orders


def save_restaurant_firestore(data):
    db.collection("restaurants").add(data)

def save_supermarket_firestore(data):
    db.collection("supermarkets").add(data)

def save_order_firestore(data):
    db.collection("orders").add(data)

@app.route("/")
def home():
    return render_template("home.html")


@app.route("/index")
def index():
    return render_template("index.html")



@app.route("/submit_order", methods=["POST"])
def submit_order():
    try:
        rid = request.form.get("rid")
        table = request.form.get("table")
        item_name = request.form.get("item_name")
        price = request.form.get("price")

        db.collection("orders").add({
            "restaurant_id": rid,
            "table_no": table,
            "item_name": item_name,
            "price": price,
            "status": "pending"
        })

        return redirect(f"/menu/{rid}/{table}")

    except Exception as e:
        return f"Order Error ❌ {str(e)}"


# =========================
# 🔐 ADMIN ROUTE
# =========================
@app.route("/admin", methods=["GET", "POST"])
def admin():

    if request.method == "POST" and not session.get("admin_ok"):
        try:
            passwords = get_system_passwords()
            real_pass = passwords.get("admin_password")
            entered   = request.form.get("password", "").strip()

            if entered != real_pass:
                return render_template(
                    "admin_login.html",
                    error="Wrong password ❌"
                )

            session["admin_ok"] = True
            return redirect("/admin")

        except Exception as e:
            print("ADMIN LOGIN ERROR:", e)
            return render_template(
                "admin_login.html",
                error=f"System Error ❌ {str(e)}"
            )

    if not session.get("admin_ok"):
        return render_template("admin_login.html")

    try:
        restaurants  = get_restaurants_firestore()
        supermarkets = get_supermarkets_firestore()
        pharmacies   = get_pharmacies_list()
        renewal_requests = get_renewal_requests()
        orders       = get_orders_firestore()
        total        = len(orders)

        info_docs = db.collection("system_info").stream()
        all_info  = []

        for doc in info_docs:
            data = doc.to_dict()
            all_info.append({
                "id":       doc.id,
                "title":    data.get("title", ""),
                "content":  data.get("content", ""),
                "image":    data.get("image", ""),
                "video":    data.get("video", ""),
                "date":     str(data.get("date", "")),
                "position": data.get("position", 0)
            })

        all_info.sort(key=lambda x: x.get("position", 0))

        review_docs      = db.collection("reviews").stream()
        review_count_map = {}

        for doc in review_docs:
            item = doc.to_dict()
            rid  = item.get("restaurant_id")
            if rid:
                review_count_map[rid] = review_count_map.get(rid, 0) + 1

        for r in restaurants:
            rid              = r.get("id")
            r["review_count"]= review_count_map.get(rid, 0)

        top_reviews = sorted(
            restaurants,
            key=lambda x: x.get("review_count", 0),
            reverse=True
        )[:3]

        sent_broadcasts = []
        for doc in db.collection("admin_broadcasts").order_by(
                "created_at", direction=firestore.Query.DESCENDING).limit(50).stream():
            b = doc.to_dict()
            b["id"] = doc.id
            sent_broadcasts.append(b)

        return render_template(
            "admin.html",
            restaurants=restaurants,
            supermarkets=supermarkets,
            pharmacies=pharmacies,
            renewal_requests=renewal_requests,
            now_date=datetime.now().strftime("%Y-%m-%d"),
            orders=orders,
            total=total,
            top_reviews=top_reviews,
            all_info=all_info,
            sent_broadcasts=sent_broadcasts
        )

    except Exception as e:
        print("ADMIN LOAD ERROR:", e)
        return render_template(
            "admin_login.html",
            error=f"Admin Error ❌ {str(e)}"
        )


# =========================
# 🔓 LOGOUT ADMIN
# =========================
@app.route("/logout_admin")
def logout_admin():
    session.pop("admin_ok",      None)
    session.pop("register_ok",   None)
    return redirect("/admin")


# =========================
# 🔓 LOGOUT REGISTER
# =========================
@app.route("/logout_register")
def logout_register():
    session.pop("register_ok", None)
    return redirect("/register")


@app.route("/logout")
def logout():
    session.clear()
    return redirect("/")


# =========================
# 🔐 CHANGE SYSTEM PASSWORDS
# =========================
@app.route("/change_system_passwords", methods=["POST"])
def change_system_passwords():
    try:
        if not session.get("admin_ok"):
            return jsonify({"success": False, "message": "Unauthorized"})

        data          = request.get_json()
        admin_pass    = data.get("admin_password")
        register_pass = data.get("register_password")

        if not admin_pass or not register_pass:
            return jsonify({"success": False, "message": "Fill all fields"})

        db.collection("system_passwords").document("main").set({
            "admin_password":    admin_pass,
            "register_password": register_pass
        })

        return jsonify({"success": True, "message": "Passwords updated successfully ✅"})

    except Exception as e:
        return jsonify({"success": False, "message": str(e)})


@app.route("/change_passwords", methods=["POST"])
def change_passwords():
    new_admin    = request.form.get("admin_pass")
    new_register = request.form.get("register_pass")

    conn = sqlite3.connect("database.db")
    c    = conn.cursor()

    c.execute("""
        UPDATE settings
        SET admin_password=?,
            register_password=?
        WHERE id=1
    """, (new_admin, new_register))

    conn.commit()
    conn.close()

    return redirect("/admin")


# =========================
# ✅ ACTIVATE RESTAURANT
# =========================
@app.route("/activate/<string:rid>")
@app.route("/activate/restaurant/<string:rid>")
def activate_restaurant(rid):
    try:
        if not session.get("admin_ok"):
            return redirect("/admin")

        restaurant_ref = db.collection("restaurants").document(rid)
        restaurant_doc = restaurant_ref.get()

        if not restaurant_doc.exists:
            return f"Restaurant not found ❌ ID: {rid}"

        restaurant_ref.update({
            "active":       True,
            "status":       "active",
            "activated_at": datetime.now()
        })

        return redirect("/admin")

    except Exception as e:
        return f"Activate restaurant error ❌ {e}"


# =========================
# ❌ DISABLE RESTAURANT
# =========================
@app.route("/disable/<string:rid>")
@app.route("/disable/restaurant/<string:rid>")
def disable_restaurant(rid):
    try:
        if not session.get("admin_ok"):
            return redirect("/admin")

        restaurant_ref = db.collection("restaurants").document(rid)
        restaurant_doc = restaurant_ref.get()

        if not restaurant_doc.exists:
            return f"Restaurant not found ❌ ID: {rid}"

        # Disabling also marks the subscription as expired as of today —
        # so re-activating later can't accidentally inherit leftover
        # "paid" days from before the disable.
        today_str = datetime.now().strftime("%Y-%m-%d")
        restaurant_ref.update({
            "active":      False,
            "status":      "disabled",
            "expiry_date": today_str,
            "expiry":      today_str,
            "disabled_at": datetime.now()
        })

        return redirect("/admin")

    except Exception as e:
        return f"Disable restaurant error ❌ {e}"

# =========================
# 🗑 DELETE RESTAURANT
# =========================
@app.route("/delete_restaurant/<string:rid>")
@app.route("/delete/restaurant/<string:rid>")
def delete_restaurant(rid):
    try:
        if not session.get("admin_ok"):
            return redirect("/admin")

        db.collection("restaurants").document(rid).delete()
        return redirect("/admin")

    except Exception as e:
        return f"Delete restaurant error ❌ {e}"


# =========================
# 🛒 SUPERMARKET FUNCTIONS
# =========================

# =========================
# ✅ ACTIVATE SUPERMARKET
# =========================
@app.route("/activate_market/<string:mid>")
@app.route("/activate/supermarket/<string:mid>")
def activate_market(mid):
    try:
        if not session.get("admin_ok"):
            return redirect("/admin")
        db.collection("supermarkets").document(mid).update({"active": True})
        return redirect("/admin")
    except Exception as e:
        return f"Activate market error ❌ {e}"


# =========================
# ❌ DISABLE SUPERMARKET
# =========================
@app.route("/disable_market/<string:mid>")
@app.route("/disable/supermarket/<string:mid>")
def disable_market(mid):
    try:
        if not session.get("admin_ok"):
            return redirect("/admin")
        today_str = datetime.now().strftime("%Y-%m-%d")
        db.collection("supermarkets").document(mid).update({
            "active": False,
            "status": "disabled",
            "expiry_date": today_str,
            "expiry": today_str
        })
        return redirect("/admin")
    except Exception as e:
        return f"Disable market error ❌ {e}"


# =========================
# 🗑 DELETE SUPERMARKET
# =========================
@app.route("/delete_market/<string:mid>")
@app.route("/delete/supermarket/<string:mid>")
def delete_market(mid):
    try:
        if not session.get("admin_ok"):
            return redirect("/admin")
        db.collection("supermarkets").document(mid).delete()
        return redirect("/admin")
    except Exception as e:
        return f"Delete market error ❌ {e}"


# =========================
# 🔄 RENEW SUPERMARKET
# =========================
@app.route("/renew/supermarket/<string:mid>", methods=["POST"])
def renew_market(mid):
    if not session.get("admin_ok"):
        return jsonify({"success": False, "error": "Unauthorized"}), 401
    try:
        market_ref = db.collection("supermarkets").document(mid)
        market_doc = market_ref.get()
        if not market_doc.exists:
            return jsonify({"success": False, "error": f"Supermarket not found — ID: {mid}"})

        market_data = market_doc.to_dict()
        months = int((request.get_json() or {}).get("months", 3) or 3)
        if months < 1:
            return jsonify({"success": False, "error": "Months must be at least 1"})

        monthly_fee  = _monthly_fee_for("supermarket", market_data)
        total_amount = round(monthly_fee * months, 2)

        # Always resets from TODAY — any leftover time from before this
        # renewal is not carried over, matching a fresh subscription
        # purchase rather than an extension.
        new_expiry = (datetime.now() + timedelta(days=months * 30)).strftime("%Y-%m-%d")
        market_ref.update({
            "active": True,
            "status": "active",
            "expiry_date": new_expiry,
            "expiry": new_expiry,
            "renewed_at": datetime.now().isoformat()
        })

        db.collection("subscription_payments").add({
            "entity_type": "supermarket", "entity_id": mid,
            "business_name": market_data.get("name", mid),
            "months": months, "amount": total_amount,
            "payment_method": "admin_manual", "status": "SUCCESS",
            "created_at": datetime.now().isoformat()
        })

        return jsonify({"success": True, "expiry_date": new_expiry, "amount": total_amount, "months": months})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


# =========================
# 🗑 DELETE MENU
# =========================
@app.route("/delete_menu/<mid>/<rid>")
def delete_menu(mid, rid):
    # Menu management is Cashier-only (matches /add_menu, /edit_menu_item,
    # /delete_menu_item) — this older route was previously reachable with
    # no auth check at all; gating it the same way closes that gap rather
    # than leaving two ways to delete a menu item with different rules.
    if not session.get("staff_ok") or session.get("staff_role") != "cashier" or session.get("staff_rid") != rid:
        return "Unauthorized — only the Cashier can delete menu items ❌", 401
    try:
        restaurant_ref = db.collection("restaurants").document(rid)
        menu_ref = restaurant_ref.collection("menu").document(mid)
        menu_doc = menu_ref.get()

        if not menu_doc.exists:
            return "Menu item not found ❌"

        menu_data = menu_doc.to_dict()
        image_name = menu_data.get("image")

        if image_name:
            if image_name.startswith("http"):
                delete_from_firebase_storage(image_name)
            else:
                image_path = os.path.join("static", "uploads", image_name)
                if os.path.exists(image_path):
                    os.remove(image_path)

        menu_ref.delete()
        return redirect(f"/restaurant_admin/{rid}")

    except Exception as e:
        return f"Delete menu error ❌ {str(e)}"


# =========================
# 🔄 ADMIN RENEW RESTAURANT
# =========================
@app.route("/renew/restaurant/<string:rid>", methods=["POST"])
def renew_restaurant(rid):
    if not session.get("admin_ok"):
        return jsonify({"success": False, "error": "Unauthorized"}), 401
    try:
        restaurant_ref = db.collection("restaurants").document(rid)
        restaurant_doc = restaurant_ref.get()
        if not restaurant_doc.exists:
            return jsonify({"success": False, "error": f"Restaurant not found — ID: {rid}"})

        restaurant_data = restaurant_doc.to_dict()
        months = int((request.get_json() or {}).get("months", 3) or 3)
        if months < 1:
            return jsonify({"success": False, "error": "Months must be at least 1"})

        monthly_fee  = _monthly_fee_for("restaurant", restaurant_data)
        total_amount = round(monthly_fee * months, 2)

        # Always resets from TODAY — see renew_market for why.
        new_expiry = (datetime.now() + timedelta(days=months * 30)).strftime("%Y-%m-%d")
        restaurant_ref.update({
            "active": True,
            "status": "active",
            "expiry_date": new_expiry,
            "expiry": new_expiry,
            "renewed_at": datetime.now().isoformat()
        })

        db.collection("subscription_payments").add({
            "entity_type": "restaurant", "entity_id": rid,
            "business_name": restaurant_data.get("name", rid),
            "months": months, "amount": total_amount,
            "payment_method": "admin_manual", "status": "SUCCESS",
            "created_at": datetime.now().isoformat()
        })

        return jsonify({"success": True, "expiry_date": new_expiry, "amount": total_amount, "months": months})

    except Exception as e:
        print("RENEW RESTAURANT ERROR:", e)
        return jsonify({"success": False, "error": str(e)})


# =========================
# 📝 REGISTER RESTAURANT
# =========================
@app.route("/register", methods=["GET", "POST"])
def register():
    try:
        if not session.get("register_ok"):
            if request.method == "POST":
                passwords = get_system_passwords()
                real_pass = passwords.get("register_password", "6993")

                if request.form.get("access_password") == real_pass:
                    session["register_ok"] = True
                    return redirect("/register")

                return render_template(
                    "access_register.html",
                    error="Wrong password ❌"
                )
            return render_template("access_register.html")

        if request.method == "POST":
            months = int(request.form["months"])
            expiry_date = (
                datetime.now() + timedelta(days=months * 30)
            ).strftime("%Y-%m-%d")

            try:
                payment_methods = json.loads(request.form.get("payment_methods_json", "[]"))
            except Exception:
                payment_methods = []
            payment_methods = [
                {"type": m.get("type", "").strip(), "code": m.get("code", "").strip()}
                for m in payment_methods if m.get("code", "").strip()
            ]
            # `payment` stays as a readable summary string for any
            # older template/report that only ever displayed one line.
            payment_summary = " | ".join(f"{m['type']}: {m['code']}" for m in payment_methods)

            data = {
                "name": request.form["name"].strip(),
                "phone": request.form.get("phone", "").strip(),
                "username": request.form["username"].strip(),
                "password": request.form["password"].strip(),
                "kitchen_password": request.form["kitchen_password"].strip(),
                "restaurant_admin_password": request.form["restaurant_admin_password"].strip(),
                "admin_name": request.form.get("admin_name", "").strip(),
                "admin_email": request.form.get("admin_email", "").strip(),
                "price": request.form["price"].strip(),
                "payment": payment_summary,
                "payment_methods": payment_methods,
                "expiry": expiry_date,
                "active": True,
                "review_count": 0,
                "average_rating": 0,
                "created_at": datetime.now()
            }

            doc_ref = db.collection("restaurants").add(data)
            rid = doc_ref[1].id

            restaurant_ref = db.collection("restaurants").document(rid)
            restaurant_ref.collection("menu").document("init").set({"created_at": datetime.now()})
            restaurant_ref.collection("orders").document("init").set({"created_at": datetime.now()})

            return redirect("/admin")

        return render_template("register.html")

    except Exception as e:
        print("Register Error:", e)
        return f"Register Error ❌ {str(e)}"


# =========================
# 🔐 LOGIN RESTAURANT
# =========================
@app.route("/login", methods=["GET", "POST"])
def login():
    try:
        if request.method == "POST":
            username = request.form.get("username", "").strip()
            password = request.form.get("password", "").strip()

            docs = db.collection("restaurants").stream()

            for doc in docs:
                data = doc.to_dict()

                if (
                    data.get("username") == username and
                    data.get("password") == password
                ):
                    if not data.get("active", True):
                        return render_template(
                            "login.html",
                            error="Account disabled ❌"
                        )

                    session["restaurant_login"] = True
                    session["restaurant_id"] = doc.id
                    session["restaurant_name"] = data.get("name")

                    return redirect(f"/dashboard/{doc.id}")

            return render_template(
                "login.html",
                error="Wrong username or password ❌"
            )

        return render_template("login.html")

    except Exception as e:
        print("LOGIN ERROR:", e)
        return render_template(
            "login.html",
            error=f"System Error ❌ {str(e)}"
        )


# ==========================================
# 🛒 SUPERMARKET — ROUTES
# ==========================================

import string

# ==========================================
# 🗄️ INIT SUPERMARKET SALES TABLE (SQLite)
# ==========================================
def init_supermarket_sales(conn, c):
    c.execute("""
        CREATE TABLE IF NOT EXISTS supermarket_sales (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            receipt_no TEXT,
            market_id  TEXT,
            total      REAL DEFAULT 0,
            profit     REAL DEFAULT 0,
            items_json TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
    """)
    conn.commit()


# ==========================================
# 🔐 SUPERMARKET LOGIN
# ==========================================
@app.route("/supermarket_login", methods=["GET", "POST"])
def supermarket_login():
    try:
        if request.method == "POST":
            username = request.form.get("username", "").strip()
            password = request.form.get("password", "").strip()

            for doc in db.collection("supermarkets").stream():
                data = doc.to_dict()

                if (data.get("username") == username and
                        data.get("password") == password):

                    # Hubi active — Firestore wuxuu isticmaalaa "active" ama "status"
                    is_active = data.get("active", data.get("status", True))
                    if is_active is False or is_active == "disabled":
                        return render_template(
                            "supermarket_login.html",
                            error="Account disabled ❌"
                        )

                    session["market_id"]    = doc.id
                    session["market_name"]  = data.get("name", "Supermarket")
                    session["market_login"] = True
                    return redirect("/supermarket_dashboard")

            return render_template(
                "supermarket_login.html",
                error="Wrong username or password ❌"
            )

        return render_template("supermarket_login.html")

    except Exception as e:
        print("SUPERMARKET LOGIN ERROR:", e)
        return render_template(
            "supermarket_login.html",
            error=f"System Error ❌ {str(e)}"
        )


# ==========================================
# 🏠 SUPERMARKET DASHBOARD
# ==========================================
@app.route("/supermarket_dashboard")
def supermarket_dashboard():
    try:
        mid = session.get("market_id")
        if not mid:
            return redirect("/supermarket_login")

        market_doc = db.collection("supermarkets").document(mid).get()
        if market_doc.exists:
            market = market_doc.to_dict()
            is_active = market.get("active", market.get("status", True))
            if is_active is False or is_active == "disabled":
                return render_template(
                    "renew.html",
                    entity_type    = "supermarket",
                    entity_label   = "Supermarket",
                    entity_id      = mid,
                    business_name  = market.get("name", session.get("market_name", "Supermarket")),
                    monthly_fee    = _monthly_fee_for("supermarket", market),
                    payment_number = ADMIN_PAYMENT_NUMBER,
                    logout_url     = "/logout"
                )

        # PRODUCTS — KA SOO QAAD FIRESTORE
        products_raw = []
        try:
            docs = db.collection("supermarkets").document(mid)\
                     .collection("products").stream()
            for doc in docs:
                p          = doc.to_dict()
                p["id"]    = doc.id
                p["name"]  = p.get("name", "")
                p["barcode"]= p.get("barcode", "")
                p["price"] = p.get("price", 0)
                p["cost_price"] = p.get("cost_price", 0)
                p["stock"] = p.get("stock", 0)
                p["category"]   = p.get("category", "General")
                p["unit"]       = p.get("unit", "pcs")
                p["box_qty"]    = p.get("box_qty", 0)
                p["box_price"]  = p.get("box_price", 0)
                products_raw.append(p)
        except Exception as e:
            print("Products load error:", e)

        # ORDERS — KA SOO QAAD SQLITE
        orders = []
        try:
            conn = sqlite3.connect(DB_PATH)
            c    = conn.cursor()
            init_supermarket_sales(conn, c)
            c.execute("""
                SELECT id, receipt_no, total, created_at, items_json
                FROM supermarket_sales
                WHERE market_id=?
                ORDER BY created_at DESC LIMIT 50
            """, (mid,))
            orders = c.fetchall()
            conn.close()
        except Exception as e:
            print("Orders load error:", e)

        return render_template(
            "supermarket_dashboard.html",
            products          = products_raw,
            supermarket_orders= orders,
            market_name       = session.get("market_name", "Supermarket"),
            market_id         = mid,
            today             = datetime.now().strftime("%Y-%m-%d"),
            broadcasts        = _get_active_broadcasts("supermarket", mid)
        )

    except Exception as e:
        return f"Dashboard Error ❌ {str(e)}"


# ==========================================
# 📦 GET PRODUCTS JSON (for JS barcode scan)
# ==========================================
@app.route("/supermarket/products_json")
def supermarket_products_json():
    try:
        mid = session.get("market_id")
        if not mid:
            return jsonify([])

        docs    = db.collection("supermarkets").document(mid)\
                    .collection("products").stream()
        results = []
        for doc in docs:
            p = doc.to_dict()
            results.append({
                "id":        doc.id,
                "barcode":   p.get("barcode", ""),
                "name":      p.get("name", ""),
                "price":     p.get("price", 0),
                "cost":      p.get("cost_price", 0),
                "stock":     p.get("stock", 0),
                "category":  p.get("category", "General"),
                "unit":      p.get("unit", "pcs"),
                "box_qty":   p.get("box_qty", 0),
                "box_price": p.get("box_price", 0)
            })
        return jsonify(results)

    except Exception as e:
        return jsonify({"error": str(e)})


# ==========================================
# ➕ ADD PRODUCT (JSON) → FIRESTORE
# ==========================================
@app.route("/supermarket/add_product_json", methods=["POST"])
def supermarket_add_product_json():
    try:
        mid = session.get("market_id")
        if not mid:
            return jsonify({"success": False, "error": "Not logged in"}), 401

        data = request.get_json()
        name = data.get("name", "").strip()
        if not name:
            return jsonify({"success": False, "error": "Name required ❌"})

        db.collection("supermarkets").document(mid)\
          .collection("products").add({
            "barcode":    data.get("barcode", ""),
            "name":       name,
            "price":      float(data.get("price", 0)),
            "cost_price": float(data.get("cost_price", 0)),
            "stock":      int(data.get("stock", 0)),
            "category":   data.get("category", "General"),
            "unit":       data.get("unit", "pcs"),
            "box_qty":    int(data.get("box_qty", 0)),
            "box_price":  float(data.get("box_price", 0)),
            "created_at": datetime.now().isoformat()
        })

        return jsonify({"success": True})

    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


# ==========================================
# ✏️ EDIT PRODUCT → FIRESTORE
# ==========================================
@app.route("/supermarket/edit_product/<pid>", methods=["PUT"])
def supermarket_edit_product(pid):
    try:
        mid = session.get("market_id")
        if not mid:
            return jsonify({"success": False, "error": "Not logged in"}), 401

        data = request.get_json()
        db.collection("supermarkets").document(mid)\
          .collection("products").document(pid).update({
            "name":       data.get("name"),
            "barcode":    data.get("barcode", ""),
            "price":      float(data.get("price", 0)),
            "cost_price": float(data.get("cost_price", 0)),
            "stock":      int(data.get("stock", 0)),
            "category":   data.get("category", "General"),
            "unit":       data.get("unit", "pcs"),
            "box_qty":    int(data.get("box_qty", 0)),
            "box_price":  float(data.get("box_price", 0))
        })

        return jsonify({"success": True})

    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


# ==========================================
# 🗑 DELETE PRODUCT → FIRESTORE
# ==========================================
@app.route("/supermarket/delete_product/<pid>", methods=["DELETE"])
def supermarket_delete_product(pid):
    try:
        mid = session.get("market_id")
        if not mid:
            return jsonify({"success": False, "error": "Not logged in"}), 401

        db.collection("supermarkets").document(mid)\
          .collection("products").document(pid).delete()

        return jsonify({"success": True})

    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


# ==========================================
# 🛒 COMPLETE SALE
# ==========================================
@app.route("/supermarket/complete_sale", methods=["POST"])
def supermarket_complete_sale():
    try:
        mid = session.get("market_id")
        if not mid:
            return jsonify({"success": False, "error": "Not logged in"}), 401

        data = request.get_json()
        cart = data.get("cart", [])
        if not cart:
            return jsonify({"success": False, "error": "Cart is empty ❌"})

        total_revenue = 0
        total_profit  = 0
        receipt_items = []

        for item in cart:
            pid   = item.get("id")
            qty   = int(item.get("qty", 1))
            price = float(item.get("price", 0))
            label = item.get("label", "pcs")

            # GET PRODUCT FROM FIRESTORE
            prod_doc = db.collection("supermarkets").document(mid)\
                         .collection("products").document(pid).get()

            if not prod_doc.exists:
                return jsonify({"success": False, "error": f"Product not found ❌"})

            p         = prod_doc.to_dict()
            name      = p.get("name", "")
            cost      = float(p.get("cost_price", 0))
            stock     = int(p.get("stock", 0))
            box_qty   = int(p.get("box_qty", 0))
            stock_dec = box_qty * qty if "Box" in label and box_qty else qty

            if stock_dec > stock:
                return jsonify({"success": False, "error": f"Not enough stock for {name} ❌"})

            revenue = price * qty
            profit  = (price - cost) * qty
            total_revenue += revenue
            total_profit  += profit

            # UPDATE STOCK IN FIRESTORE
            db.collection("supermarkets").document(mid)\
              .collection("products").document(pid).update({
                "stock": stock - stock_dec
            })

            receipt_items.append({
                "name":  name,
                "qty":   qty,
                "price": price,
                "label": label,
                "total": revenue
            })

        receipt_no = "RCP-" + "".join(random.choices(string.digits, k=6))

        # SAVE SALE TO SQLITE
        conn = sqlite3.connect(DB_PATH)
        c    = conn.cursor()
        init_supermarket_sales(conn, c)
        c.execute("""
            INSERT INTO supermarket_sales
            (receipt_no, market_id, total, profit, items_json, created_at)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (receipt_no, mid, round(total_revenue, 2),
              round(total_profit, 2), json.dumps(receipt_items),
              datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
        conn.commit()
        conn.close()

        return jsonify({
            "success":    True,
            "receipt_no": receipt_no,
            "total":      round(total_revenue, 2),
            "profit":     round(total_profit, 2),
            "items":      receipt_items
        })

    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


# ==========================================
# 🖨️ REPRINT RECEIPT
# ==========================================
@app.route("/supermarket/receipt/<receipt_no>")
def supermarket_get_receipt(receipt_no):
    try:
        mid = session.get("market_id")
        if not mid:
            return jsonify({"success": False})

        conn = sqlite3.connect(DB_PATH)
        c    = conn.cursor()
        init_supermarket_sales(conn, c)
        c.execute("""
            SELECT items_json, total FROM supermarket_sales
            WHERE receipt_no=? AND market_id=?
        """, (receipt_no, mid))
        row = c.fetchone()
        conn.close()

        if not row:
            return jsonify({"success": False})

        return jsonify({
            "success": True,
            "items":   json.loads(row[0]),
            "total":   row[1]
        })

    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


# ==========================================
# 📤 BULK CSV IMPORT → FIRESTORE
# ==========================================
@app.route("/supermarket/bulk_import", methods=["POST"])
def supermarket_bulk_import():
    try:
        mid = session.get("market_id")
        if not mid:
            return jsonify({"success": False, "error": "Not logged in"}), 401

        data     = request.get_json()
        products = data.get("products", [])
        if not products:
            return jsonify({"success": False, "error": "No products ❌"})

        imported   = 0
        prod_ref   = db.collection("supermarkets").document(mid).collection("products")

        for p in products:
            name = p.get("name", "").strip()
            if not name:
                continue
            try:
                prod_ref.add({
                    "barcode":    p.get("barcode", ""),
                    "name":       name,
                    "price":      float(p.get("price", 0)),
                    "cost_price": float(p.get("cost_price", 0)),
                    "stock":      int(p.get("stock", 0)),
                    "category":   p.get("category", "General"),
                    "unit":       p.get("unit", "pcs"),
                    "box_qty":    int(p.get("box_qty", 0)),
                    "box_price":  float(p.get("box_price", 0)),
                    "created_at": datetime.now().isoformat()
                })
                imported += 1
            except:
                pass

        return jsonify({"success": True, "imported": imported})

    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


# ==========================================
# 📊 STATS
# ==========================================
@app.route("/supermarket/stats")
def supermarket_stats():
    try:
        mid = session.get("market_id")
        if not mid:
            return jsonify({})

        # Products count from Firestore
        prod_docs      = db.collection("supermarkets").document(mid)\
                           .collection("products").stream()
        total_products = 0
        low_stock      = 0
        for doc in prod_docs:
            p = doc.to_dict()
            total_products += 1
            if int(p.get("stock", 0)) <= 10:
                low_stock += 1

        # Sales from SQLite
        today = datetime.now().strftime("%Y-%m-%d")
        conn  = sqlite3.connect(DB_PATH)
        c     = conn.cursor()
        init_supermarket_sales(conn, c)
        c.execute("""
            SELECT SUM(total), COUNT(*)
            FROM supermarket_sales
            WHERE market_id=? AND date(created_at)=?
        """, (mid, today))
        row           = c.fetchone()
        today_revenue = round(row[0] or 0, 2)
        today_orders  = row[1] or 0
        conn.close()

        return jsonify({
            "total_products": total_products,
            "today_revenue":  today_revenue,
            "today_orders":   today_orders,
            "low_stock":      low_stock
        })

    except Exception as e:
        return jsonify({"error": str(e)})


# ==========================================
# 📊 ANALYTICS
# ==========================================
@app.route("/supermarket/analytics")
def supermarket_analytics():
    try:
        mid       = session.get("market_id")
        if not mid:
            return jsonify({})

        date_from = request.args.get("from", datetime.now().strftime("%Y-%m-%d"))
        date_to   = request.args.get("to",   datetime.now().strftime("%Y-%m-%d"))

        conn = sqlite3.connect(DB_PATH)
        c    = conn.cursor()
        init_supermarket_sales(conn, c)

        c.execute("""
            SELECT SUM(total), SUM(profit), COUNT(*)
            FROM supermarket_sales
            WHERE market_id=? AND date(created_at) BETWEEN ? AND ?
        """, (mid, date_from, date_to))
        row = c.fetchone()

        c.execute("""
            SELECT date(created_at), COUNT(*), SUM(total), SUM(profit)
            FROM supermarket_sales
            WHERE market_id=? AND date(created_at) BETWEEN ? AND ?
            GROUP BY date(created_at)
            ORDER BY date(created_at) DESC
        """, (mid, date_from, date_to))
        daily = [{"date":r[0],"orders":r[1],
                  "revenue":round(r[2],2),"profit":round(r[3],2)}
                 for r in c.fetchall()]

        # Top products from items_json
        all_sales   = c.execute("""
            SELECT items_json FROM supermarket_sales
            WHERE market_id=? AND date(created_at) BETWEEN ? AND ?
        """, (mid, date_from, date_to)).fetchall()

        conn.close()

        product_map = {}
        items_total = 0
        for sale in all_sales:
            try:
                items = json.loads(sale[0])
                for item in items:
                    k = item.get("name", "")
                    product_map[k] = product_map.get(k, 0) + item.get("qty", 0)
                    items_total   += item.get("qty", 0)
            except:
                pass

        top_products = sorted(
            [{"name": k, "qty": v} for k, v in product_map.items()],
            key=lambda x: x["qty"], reverse=True
        )[:10]

        return jsonify({
            "revenue":      round(row[0] or 0, 2),
            "profit":       round(row[1] or 0, 2),
            "orders":       row[2] or 0,
            "items_sold":   items_total,
            "daily":        daily,
            "top_products": top_products
        })

    except Exception as e:
        return jsonify({"error": str(e)})


# ==========================================
# 📋 REGISTER SUPERMARKET (ADMIN)
# ==========================================
@app.route("/supermarket_register", methods=["GET", "POST"])
def supermarket_register():
    try:
        if request.method == "POST":
            months = int(request.form.get("months", 1))
            expiry = (datetime.now() + timedelta(days=months*30)).strftime("%Y-%m-%d")
            db.collection("supermarkets").add({
                "name":       request.form.get("name", ""),
                "username":   request.form.get("username", ""),
                "password":   request.form.get("password", ""),
                "price":      request.form.get("price", ""),
                "payment":    request.form.get("payment", ""),
                "expiry_date":expiry,
                "active":     True,
                "created_at": datetime.now()
            })
            return redirect("/supermarket_login")
        return render_template("supermarket_register.html")
    except Exception as e:
        return f"Register Error ❌ {str(e)}"

# =====================================
# 📊 RESTAURANT DASHBOARD
# =====================================
@app.route("/dashboard/<rid>")
def dashboard(rid):
    try:
        if not session.get("restaurant_login"):
            return redirect("/login")

        restaurant_ref = db.collection("restaurants").document(rid)
        restaurant_doc = restaurant_ref.get()

        if not restaurant_doc.exists:
            return "Restaurant not found ❌"

        restaurant = restaurant_doc.to_dict()

        def suspended_page():
            return render_template(
                "renew.html",
                rid            = rid,
                entity_type    = "restaurant",
                entity_label   = "Restaurant",
                entity_id      = rid,
                business_name  = restaurant.get("name", "Restaurant"),
                monthly_fee    = _monthly_fee_for("restaurant", restaurant),
                payment_number = ADMIN_PAYMENT_NUMBER,
                logout_url     = "/logout"
            )

        if not restaurant.get("active", True):
            return suspended_page()

        expiry = restaurant.get("expiry", "")
        if expiry:
            try:
                expiry_date = datetime.strptime(expiry, "%Y-%m-%d")
                if datetime.now() >= expiry_date:
                    restaurant_ref.update({"active": False})
                    return suspended_page()
            except Exception as expiry_error:
                print("Expiry Error:", expiry_error)

        menu = []
        menu_docs = restaurant_ref.collection("menu").stream()
        for doc in menu_docs:
            item = doc.to_dict()
            item["id"] = doc.id
            item["name"] = item.get("name", "No Name")
            item["price"] = item.get("price", 0)
            item["image"] = item.get("image", "")
            menu.append(item)

        ads = []
        ad_docs = restaurant_ref.collection("ads").stream()
        for doc in ad_docs:
            ad = doc.to_dict()
            ad["id"] = doc.id
            ad["title"] = ad.get("title", "")
            ad["image"] = ad.get("image", "")
            ad["audio"] = ad.get("audio", "")
            ad["created_at"] = ad.get("created_at", None)
            ads.append(ad)

        ads = list(reversed(ads))

        # Real, computed stats — no placeholders. Views has no tracking
        # system behind it yet, so it's left out rather than faked.
        today = datetime.now().strftime("%Y-%m-%d")
        total_orders_today = 0
        total_sales_today = 0.0
        for doc in restaurant_ref.collection("orders").stream():
            o = doc.to_dict()
            created = o.get("created_at")
            created_date = created.strftime("%Y-%m-%d") if hasattr(created, "strftime") else str(created)[:10]
            if created_date == today:
                total_orders_today += 1
                if str(o.get("status", "")).lower() == "paid":
                    total_sales_today += float(o.get("price", 0))

        return render_template(
            "dashboard.html",
            rid=rid,
            restaurant=restaurant.get("name", "Restaurant"),
            restaurant_phone=restaurant.get("phone", ""),
            restaurant_expiry=restaurant.get("expiry", ""),
            menu=menu,
            ads=ads,
            total_orders_today=total_orders_today,
            total_sales_today=round(total_sales_today, 2),
            broadcasts=_get_active_broadcasts("restaurant", rid)
        )

    except Exception as e:
        print("Dashboard Error:", e)
        return f"Dashboard Error ❌ {str(e)}"


@app.route("/verify_admin_login/<rid>", methods=["POST"])
def verify_admin_login(rid):
    """Quick inline re-verification widget on the dashboard topbar —
    the page itself is already gated by session['restaurant_login'];
    this just re-checks the same username/password against THIS
    restaurant so an admin whose session lapsed mid-work can confirm
    they're still who they say they are without leaving the page."""
    if not session.get("restaurant_login") or session.get("restaurant_id") != rid:
        return jsonify({"success": False, "error": "Session expired — please log in again"}), 401
    try:
        data = request.get_json() or {}
        username = data.get("username", "").strip()
        password = data.get("password", "").strip()

        restaurant_doc = db.collection("restaurants").document(rid).get()
        if not restaurant_doc.exists:
            return jsonify({"success": False, "error": "Restaurant not found"})

        r = restaurant_doc.to_dict()
        if r.get("username") == username and r.get("password") == password:
            session["restaurant_login"] = True
            session["restaurant_id"] = rid
            session["restaurant_name"] = r.get("name")
            return jsonify({"success": True})

        return jsonify({"success": False, "error": "Wrong username or password"})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


# =====================================
# 📱 CUSTOMER MOBILE MENU
# =====================================
@app.route("/menu/<rid>/<table_no>")
def mobile_menu(rid, table_no):
    try:
        restaurant_ref = db.collection("restaurants").document(rid)
        restaurant_doc = restaurant_ref.get()

        if not restaurant_doc.exists:
            return "Restaurant not found ❌"

        restaurant = restaurant_doc.to_dict()

        payment = restaurant.get("payment", "")
        payment_name = restaurant.get("payment_name", "")
        payment_number = restaurant.get("payment_number", payment)

        menu = []
        menu_docs = restaurant_ref.collection("menu").stream()
        for doc in menu_docs:
            if doc.id == "init":
                continue
            item = doc.to_dict()
            item["id"] = doc.id
            item["image"] = item.get("image", "")
            item["name"] = item.get("name", "No Name")
            item["price"] = item.get("price", 0)
            menu.append(item)

        ads = []
        ads_docs = restaurant_ref.collection("ads").stream()
        for doc in ads_docs:
            if doc.id == "init":
                continue
            ad = doc.to_dict()
            ad["id"] = doc.id
            ad["image"] = ad.get("image", "")
            ad["audio"] = ad.get("audio", "")
            ad["title"] = ad.get("title", "")
            ads.append(ad)

        ads = list(reversed(ads))

        return render_template(
            "customer_menu.html",
            menu=menu,
            table=table_no,
            rid=rid,
            ads=ads,
            restaurant=restaurant.get("name", "Restaurant"),
            payment=payment,
            payment_name=payment_name,
            payment_number=payment_number,
            order_status=None
        )

    except Exception as e:
        print("Menu Error:", e)
        return f"Menu Error ❌ {str(e)}"


# =====================================
# 🛒 CUSTOMER ORDER
# =====================================
@app.route("/customer_order/<rid>", methods=["POST"])
def customer_order(rid):
    try:
        items = request.form.get("items", "")
        price = request.form.get("price", "0")
        table = request.form.get("table", "")
        drink_option = request.form.get("drink_option", "")
        food_option = request.form.get("food_option", "")
        tea_option = request.form.get("tea_option", "")

        if not items:
            return "No items selected ❌"
        if not table:
            return "Table number missing ❌"

        db.collection("restaurants").document(rid).collection("orders").add({
            "items": items,
            "price": float(price),
            "table": str(table),
            "drink_option": drink_option,
            "food_option": food_option,
            "tea_option": tea_option,
            "status": "pending",
            "created_at": datetime.utcnow()
        })

        return redirect(f"/menu/{rid}/{table}")

    except Exception as e:
        print("Order Error:", e)
        return f"Order failed ❌ {str(e)}"


# =====================================
# 📊 SALES DATA
# =====================================
@app.route("/sales_data/<rid>")
def sales_data(rid):
    try:
        from_date = request.args.get("from")
        to_date = request.args.get("to")

        order_docs = db.collection("orders") \
            .where("restaurant_id", "==", rid) \
            .stream()

        data = []
        total_sales = 0

        for doc in order_docs:
            item = doc.to_dict()
            created_at = str(item.get("created_at", ""))

            if from_date and to_date:
                if created_at[:10] < from_date or created_at[:10] > to_date:
                    continue

            total_sales += float(item.get("total", 0))
            data.append({
                "table": item.get("table_no"),
                "food": item.get("food"),
                "total": item.get("total"),
                "date": created_at
            })

        return jsonify({
            "orders": data,
            "total_orders": len(data),
            "total_sales": total_sales
        })

    except Exception as e:
        return jsonify({"error": str(e)})


# =====================================
# 🍽 RESTAURANT ADMIN PANEL
# =====================================
from collections import defaultdict


def _parse_created_at(value):
    """Waxay isku dayaysaa in ay convert-garayso created_at (Firestore timestamp,
    datetime, ama string) una soo celiso datetime object oo timezone-aware.
    Haddii ay fashilanto waxay soo celisaa None."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value if value.tzinfo else value.replace(tzinfo=timezone.utc)
    try:
        # Firestore DatetimeWithNanoseconds ama string ISO
        return datetime.fromisoformat(str(value).replace("Z", "+00:00"))
    except Exception:
        return None


def _range_start(range_key, now):
    """Waxay soo celisaa taariikhda bilowga range-ka la doortay (day/week/month/year)."""
    if range_key == "day":
        return now - timedelta(days=1)
    if range_key == "week":
        return now - timedelta(days=7)
    if range_key == "month":
        return now - timedelta(days=30)
    if range_key == "year":
        return now - timedelta(days=365)
    return now - timedelta(days=7)  # default = week


@app.route("/restaurant_admin/<rid>", methods=["GET", "POST"])
def restaurant_admin(rid):
    try:
        if not session.get("admin_" + str(rid)):
            return redirect(f"/restaurant_admin_login/{rid}")

        restaurant_ref = db.collection("restaurants").document(rid)
        restaurant_doc = restaurant_ref.get()

        if not restaurant_doc.exists:
            return "Restaurant not found ❌"

        restaurant = restaurant_doc.to_dict()
        restaurant["id"] = rid

        if request.method == "POST":
            update_data = {
                "name": request.form.get("name", "").strip(),
                "username": request.form.get("username", "").strip(),
                "password": request.form.get("password", "").strip(),
                "kitchen_password": request.form.get("kitchen_password", "").strip(),
                "restaurant_admin_password": request.form.get("restaurant_admin_password", "").strip(),
                "updated_at": datetime.now(timezone.utc)
            }
            restaurant_ref.update(update_data)
            return redirect(f"/restaurant_admin/{rid}")

        # ---------- Range filter (?range=day/week/month/year) ----------
        range_key = request.args.get("range", "week")
        now = datetime.now(timezone.utc)
        range_start = _range_start(range_key, now)

        # ---------- Menu ----------
        menu = []
        menu_by_id = {}
        menu_docs = restaurant_ref.collection("menu").stream()
        for doc in menu_docs:
            item = doc.to_dict()
            item["id"] = doc.id
            menu.append(item)
            menu_by_id[doc.id] = item

        # ---------- Orders ----------
        orders = []
        total = 0.0
        order_docs = restaurant_ref.collection("orders").stream()

        # Xogta chart-yada ee waxaan ka soo xisaabin doono orders-ka dhabta ah
        daily_revenue = defaultdict(float)   # {date_str: revenue}
        daily_orders = defaultdict(int)      # {date_str: order count}
        status_counts = defaultdict(int)     # {status: count}
        item_sales = defaultdict(lambda: {"qty": 0, "revenue": 0.0, "image": "", "name": ""})
        category_revenue = defaultdict(float)  # {category: revenue}

        # Real money report figures — ONLY orders that were actually
        # paid count as revenue collected (unlike `total` below, which
        # sums every order regardless of status).
        paid_revenue_all_time = 0.0
        paid_orders_all_time = 0

        for doc in order_docs:
            order = doc.to_dict()
            order["id"] = doc.id

            try:
                price = float(order.get("price", 0))
            except Exception:
                price = 0.0
            total += price
            orders.append(order)

            created_at = _parse_created_at(order.get("created_at"))
            status = str(order.get("status") or "pending").lower()
            status_counts[status] += 1

            if status == "paid":
                paid_revenue_all_time += price
                paid_orders_all_time += 1

            # Only bucket into the range-filtered charts if we have a valid date
            if created_at and created_at >= range_start:
                if range_key == "day":
                    bucket = created_at.strftime("%H:00")
                elif range_key == "year":
                    bucket = created_at.strftime("%b")
                else:
                    bucket = created_at.strftime("%b %d")

                daily_revenue[bucket] += price
                daily_orders[bucket] += 1

            # Line items for top-selling & category breakdown, if the order stores them
            items = order.get("items")
            if isinstance(items, list):
                for it in items:
                    name = it.get("name") or "Unknown"
                    qty = int(it.get("qty", it.get("quantity", 1)) or 1)
                    line_price = float(it.get("price", 0) or 0)
                    item_sales[name]["qty"] += qty
                    item_sales[name]["revenue"] += line_price * qty
                    item_sales[name]["name"] = name

                    menu_match = next((m for m in menu if m.get("name") == name), None)
                    item_sales[name]["image"] = menu_match.get("image", "") if menu_match else it.get("image", "")

                    category = (menu_match.get("category") if menu_match else it.get("category")) or "Other"
                    category_revenue[category] += line_price * qty

        # ---------- Order status breakdown (for template list + donut chart) ----------
        status_breakdown = [
            {"label": s.title(), "count": c} for s, c in sorted(status_counts.items())
        ]
        status_labels = [s["label"] for s in status_breakdown] or ["No Data"]
        status_counts_list = [s["count"] for s in status_breakdown] or [1]

        # ---------- Overview chart (revenue + orders over time) ----------
        chart_labels = sorted(daily_revenue.keys(), key=lambda k: k)
        chart_revenue = [round(daily_revenue[k], 2) for k in chart_labels]
        chart_orders = [daily_orders[k] for k in chart_labels]

        # ---------- Top selling items ----------
        top_items = sorted(item_sales.values(), key=lambda x: x["qty"], reverse=True)[:5]
        for t in top_items:
            t["revenue"] = round(t["revenue"], 2)

        # ---------- Revenue by category ----------
        category_labels = list(category_revenue.keys()) or ["No Data"]
        category_values = [round(v, 2) for v in category_revenue.values()] or [0]

        # ---------- Staff (waiters + cashiers) ----------
        staff = []
        for doc in restaurant_ref.collection("staff_accounts").stream():
            s = doc.to_dict()
            s["id"] = doc.id
            staff.append(s)
        staff.sort(key=lambda x: (x.get("role", ""), x.get("employee_id", "")))

        # ---------- Staff work summary (ALL-TIME, not just today) — for
        # the Staff page's "work they've done" overview ----------
        staff_work_summary = []
        for s in staff:
            eid = s.get("employee_id")
            role = s.get("role")
            if not eid or role not in ("waiter", "cashier"):
                continue
            txs, tx_total = _staff_all_time_transactions(rid, eid, role)
            staff_work_summary.append({
                "name": s.get("name", ""),
                "employee_id": eid,
                "role": role,
                "transaction_count": len(txs),
                "total_amount": tx_total
            })

        # Percentage share within each role group (waiters compared to
        # waiters, cashiers to cashiers — comparing across roles isn't
        # meaningful since they earn differently).
        role_totals = defaultdict(float)
        for w in staff_work_summary:
            role_totals[w["role"]] += w["total_amount"]
        for w in staff_work_summary:
            group_total = role_totals[w["role"]]
            w["percentage"] = round((w["total_amount"] / group_total) * 100, 1) if group_total > 0 else 0.0
        staff_work_summary.sort(key=lambda x: (x["role"], -x["total_amount"]))

        # ---------- Waiter performance (today's paid orders, % of today's sales) ----------
        today_str = datetime.now().strftime("%Y-%m-%d")
        waiter_agg = {}
        for s in staff:
            if s.get("role") == "waiter":
                waiter_agg[s.get("employee_id")] = {
                    "employee_id": s.get("employee_id"), "name": s.get("name", ""),
                    "orders": 0, "sales": 0.0, "percentage": 0.0
                }
        today_waiter_sales_total = 0.0
        for o in orders:
            if str(o.get("status", "")).lower() != "paid":
                continue
            created_at = _parse_created_at(o.get("created_at"))
            if not created_at or created_at.strftime("%Y-%m-%d") != today_str:
                continue
            wid = o.get("employee_id")
            if not wid or wid not in waiter_agg:
                continue
            amount = float(o.get("price", 0))
            waiter_agg[wid]["orders"] += 1
            waiter_agg[wid]["sales"] += amount
            today_waiter_sales_total += amount
        for w in waiter_agg.values():
            w["sales"] = round(w["sales"], 2)
            w["percentage"] = round((w["sales"] / today_waiter_sales_total) * 100, 1) if today_waiter_sales_total > 0 else 0.0
        waiter_performance = sorted(waiter_agg.values(), key=lambda x: x["sales"], reverse=True)

        # ---------- Cashier performance (most recent shift: hours worked +
        # orders they personally confirmed/paid, as a % of today's total
        # confirmed orders across all cashiers) ----------
        cashier_agg = {}
        for s in staff:
            if s.get("role") == "cashier":
                cashier_agg[s.get("employee_id")] = {
                    "employee_id": s.get("employee_id"), "name": s.get("name", ""),
                    "orders_confirmed": 0, "shift_hours": "—", "shift_status": "No shift today", "percentage": 0.0
                }

        today_payments_count = {}
        total_confirmed_today = 0
        for doc in restaurant_ref.collection("payments").where("date", "==", today_str).stream():
            p = doc.to_dict()
            cid = p.get("cashier_id")
            if cid:
                today_payments_count[cid] = today_payments_count.get(cid, 0) + 1
                total_confirmed_today += 1

        for cid, agg in cashier_agg.items():
            agg["orders_confirmed"] = today_payments_count.get(cid, 0)
            agg["percentage"] = round((agg["orders_confirmed"] / total_confirmed_today) * 100, 1) if total_confirmed_today > 0 else 0.0

            # Most recent shift today for this cashier — sorted in
            # Python rather than via Firestore .order_by(), since
            # combining an equality filter with order_by on a
            # different field needs a composite index that doesn't
            # exist by default (this was the exact 400 error).
            shift_docs = list(restaurant_ref.collection("cashier_shifts")
                               .where("cashier_employee_id", "==", cid)
                               .stream())
            if shift_docs:
                shift = max(
                    (d.to_dict() for d in shift_docs),
                    key=lambda s: s.get("opened_at", "")
                )
                opened_at = shift.get("opened_at", "")
                try:
                    opened_dt = datetime.strptime(opened_at[:19], "%Y-%m-%dT%H:%M:%S") if "T" in opened_at else datetime.strptime(opened_at, "%Y-%m-%d %H:%M:%S")
                    if shift.get("status") == "closed" and shift.get("closed_at"):
                        closed_at = shift.get("closed_at")
                        closed_dt = datetime.strptime(closed_at[:19], "%Y-%m-%dT%H:%M:%S") if "T" in closed_at else datetime.strptime(closed_at, "%Y-%m-%d %H:%M:%S")
                        span = closed_dt - opened_dt
                        agg["shift_status"] = "Closed"
                    else:
                        span = datetime.now() - opened_dt
                        agg["shift_status"] = "Active now"
                    total_minutes = int(span.total_seconds() // 60)
                    h, m = divmod(max(0, total_minutes), 60)
                    agg["shift_hours"] = f"{h}h {m}m"
                except Exception:
                    pass

        cashier_performance = sorted(cashier_agg.values(), key=lambda x: x["orders_confirmed"], reverse=True)

        return render_template(
            "restaurant_admin.html",
            r=restaurant,
            menu=menu,
            orders=orders,
            staff=staff,
            waiter_performance=waiter_performance,
            cashier_performance=cashier_performance,
            staff_work_summary=staff_work_summary,
            paid_revenue_all_time=round(paid_revenue_all_time, 2),
            paid_orders_all_time=paid_orders_all_time,
            broadcasts=_get_active_broadcasts("restaurant", rid),
            total=round(total, 2),
            profit=round(total, 2),
            loss=0,
            compare_text="System working",
            rid=rid,
            range=range_key,
            status_breakdown=status_breakdown,
            status_labels=status_labels,
            status_counts=status_counts_list,
            chart_labels=chart_labels,
            chart_revenue=chart_revenue,
            chart_orders=chart_orders,
            top_items=top_items,
            category_labels=category_labels,
            category_values=category_values,
        )

    except Exception as e:
        return f"Error ❌ {str(e)}"


# =====================================
# 🧹 CLEAR KITCHEN ORDERS
# =====================================
@app.route("/clear_kitchen_orders/<rid>")
def clear_kitchen_orders(rid):
    try:
        if not session.get("admin_" + str(rid)):
            return redirect(f"/restaurant_admin_login/{rid}")

        orders_ref = db.collection("restaurants").document(rid).collection("orders")
        docs = orders_ref.stream()

        for doc in docs:
            doc.reference.update({
                "kitchen_cleared": True,
                "cleared_at": datetime.now(timezone.utc)
            })

        return redirect(f"/restaurant_admin/{rid}")

    except Exception as e:
        return f"Kitchen clear error ❌ {str(e)}"


# =====================================
# 🗑 CLEAR ALL ADS
# =====================================
@app.route("/clear_ads/<rid>")
def clear_ads(rid):
    try:
        if not session.get("admin_" + str(rid)):
            return redirect(f"/restaurant_admin_login/{rid}")

        ads_ref = db.collection("restaurants").document(rid).collection("ads")
        docs = ads_ref.stream()

        for doc in docs:
            doc.reference.delete()

        return redirect(f"/restaurant_admin/{rid}")

    except Exception as e:
        return f"Ads clear error ❌ {str(e)}"


# =====================================
# 🔐 RESTAURANT ADMIN LOGIN
# =====================================
@app.route("/restaurant_admin_login/<rid>", methods=["GET", "POST"])
def restaurant_admin_login(rid):
    try:
        restaurant_ref = db.collection("restaurants").document(rid)
        restaurant_doc = restaurant_ref.get()

        if not restaurant_doc.exists:
            return "Restaurant not found ❌"

        restaurant = restaurant_doc.to_dict()

        if request.method == "POST":
            entered_password = request.form.get("password", "").strip()

            real_password = str(
                restaurant.get("restaurant_admin_password")
                or restaurant.get("resturen_admin password")
                or ""
            ).strip()

            if entered_password == real_password:
                session["admin_" + str(rid)] = True
                return redirect(f"/restaurant_admin/{rid}")

            return f'''
            <div style="max-width:400px;margin:50px auto;font-family:Arial;">
                <h3 style="color:red;">Wrong password ❌</h3>
                <a href="/restaurant_admin_login/{rid}">Try again</a>
            </div>
            '''

        return f'''
        <form method="post"
              style="max-width:400px;margin:50px auto;font-family:Arial;
                     background:white;padding:25px;border-radius:12px;
                     box-shadow:0 0 10px rgba(0,0,0,0.1);">
            <h2 style="text-align:center;">Admin Login 🔐</h2>
            <input type="password" name="password"
                   placeholder="Enter admin password" required
                   style="width:100%;padding:12px;margin:15px 0;
                          border:1px solid #ddd;border-radius:8px;
                          box-sizing:border-box;">
            <button type="submit"
                    style="width:100%;padding:12px;background:#0a7cff;
                           color:white;border:none;border-radius:8px;
                           font-weight:bold;cursor:pointer;">
                Login
            </button>
        </form>
        '''

    except Exception as e:
        print("Login Error:", e)
        return f"Login error ❌ {str(e)}"


# =====================================
# 👥 ADD STAFF (legacy form — kept working as-is)
# =====================================
@app.route("/add_staff/<rid>", methods=["POST"])
def add_staff(rid):
    try:
        staff_data = {
            "restaurant_id": rid,
            "name": request.form["name"],
            "email": request.form["email"],
            "password": request.form["password"],
            "role": "staff",
            "created_at": datetime.now()
        }
        db.collection("restaurants").document(rid).collection("staff").add(staff_data)
        return redirect("/dashboard/" + rid)

    except Exception as e:
        return f"Add staff error ❌ {str(e)}"


# =====================================
# 🆔 SEQUENTIAL EMPLOYEE ID (per restaurant, per role)
# =====================================
def get_next_employee_id(rid, role):
    prefix = "WTR" if role == "waiter" else "CSH"
    counter_ref = db.collection("restaurants").document(rid) \
                    .collection("meta").document("staff_counter")
    field = f"{role}_count"
    counter_ref.set({field: firestore.Increment(1)}, merge=True)
    snap = counter_ref.get()
    count = snap.to_dict().get(field, 1)
    return f"{prefix}-{count:03d}"


def restaurant_is_active(rid):
    doc = db.collection("restaurants").document(rid).get()
    if not doc.exists:
        return False, None
    data = doc.to_dict()
    return data.get("active", True), data


# =====================================
# 💰 CASHIER SHIFTS + PAYMENTS
# =====================================
def get_next_payment_id(rid):
    date_str = datetime.now().strftime("%Y%m%d")
    counter_ref = db.collection("restaurants").document(rid) \
                    .collection("meta").document("payment_counter")
    counter_ref.set({"count": firestore.Increment(1)}, merge=True)
    snap = counter_ref.get()
    count = snap.to_dict().get("count", 1)
    return f"PAY-{date_str}-{count:05d}"


def get_active_cashier_shift(rid, employee_id):
    """Returns (shift_id, shift_dict) for the open shift, or (None, None)."""
    docs = db.collection("restaurants").document(rid).collection("cashier_shifts") \
        .where("cashier_employee_id", "==", employee_id) \
        .where("status", "==", "open").limit(1).stream()
    for doc in docs:
        s = doc.to_dict()
        s["id"] = doc.id
        return doc.id, s
    return None, None


# =====================================
# 🧑‍💼 STAFF MANAGEMENT (Admin — create/list/toggle/delete Waiter & Cashier accounts)
# Reachable from either restaurant admin surface: the menu/ads dashboard
# (session["restaurant_login"]) or the analytics panel at
# /restaurant_admin (session["admin_<rid>"]) — both are legitimate
# "I am this restaurant's owner" credentials, just two different
# passwords, so either one authorizes staff management.
# =====================================
def _restaurant_admin_authorized(rid):
    return (
        (session.get("restaurant_login") and session.get("restaurant_id") == rid)
        or session.get("admin_" + str(rid))
    )


@app.route("/staff_manage/<rid>")
def staff_manage(rid):
    if not _restaurant_admin_authorized(rid):
        return redirect("/login")
    try:
        restaurant_doc = db.collection("restaurants").document(rid).get()
        restaurant_name = restaurant_doc.to_dict().get("name", "Restaurant") if restaurant_doc.exists else "Restaurant"

        staff = []
        docs = db.collection("restaurants").document(rid).collection("staff_accounts").stream()
        for doc in docs:
            item = doc.to_dict()
            item["id"] = doc.id
            staff.append(item)
        staff.sort(key=lambda x: x.get("employee_id", ""))

        return render_template("staff_panel.html", rid=rid, restaurant_name=restaurant_name, staff=staff)
    except Exception as e:
        return f"Staff manage error ❌ {str(e)}"


@app.route("/staff_manage/<rid>/create", methods=["POST"])
def staff_create(rid):
    if not _restaurant_admin_authorized(rid):
        return jsonify({"success": False, "error": "Unauthorized"}), 401
    try:
        data = request.get_json() or {}
        name = data.get("name", "").strip()
        role = data.get("role", "").strip().lower()
        pin  = data.get("pin", "").strip()
        phone = data.get("phone", "").strip()
        custom_employee_id = data.get("employee_id", "").strip().upper()

        if not name or role not in ("waiter", "cashier") or not pin:
            return jsonify({"success": False, "error": "Fill all fields (role must be waiter or cashier)"})
        if not pin.isdigit() or not (4 <= len(pin) <= 6):
            return jsonify({"success": False, "error": "PIN must be 4-6 digits"})

        staff_ref = db.collection("restaurants").document(rid).collection("staff_accounts")

        if custom_employee_id:
            if not re.match(r'^[A-Z0-9\-_]{2,20}$', custom_employee_id):
                return jsonify({"success": False, "error": "Employee ID may only use letters, numbers, - and _ (2-20 characters)"})
            existing = list(staff_ref.where("employee_id", "==", custom_employee_id).limit(1).stream())
            if existing:
                return jsonify({"success": False, "error": f"Employee ID '{custom_employee_id}' is already in use"})
            employee_id = custom_employee_id
        else:
            employee_id = get_next_employee_id(rid, role)

        staff_ref.add({
            "name": name,
            "role": role,
            "employee_id": employee_id,
            "pin": pin,
            "phone": phone,
            "active": True,
            "created_at": datetime.now().isoformat()
        })
        return jsonify({"success": True, "employee_id": employee_id})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


@app.route("/staff_manage/<rid>/edit/<staff_id>", methods=["POST"])
def staff_edit(rid, staff_id):
    if not _restaurant_admin_authorized(rid):
        return jsonify({"success": False, "error": "Unauthorized"}), 401
    try:
        data = request.get_json() or {}
        name = data.get("name", "").strip()
        pin  = data.get("pin", "").strip()
        phone = data.get("phone", None)
        new_employee_id = data.get("employee_id", "").strip().upper()

        staff_ref = db.collection("restaurants").document(rid).collection("staff_accounts")
        item_ref = staff_ref.document(staff_id)
        item_doc = item_ref.get()
        if not item_doc.exists:
            return jsonify({"success": False, "error": "Staff account not found"})

        update_fields = {}
        if name:
            update_fields["name"] = name
        if pin:
            if not pin.isdigit() or not (4 <= len(pin) <= 6):
                return jsonify({"success": False, "error": "PIN must be 4-6 digits"})
            update_fields["pin"] = pin
        if phone is not None:
            update_fields["phone"] = phone.strip()
        if new_employee_id and new_employee_id != item_doc.to_dict().get("employee_id"):
            if not re.match(r'^[A-Z0-9\-_]{2,20}$', new_employee_id):
                return jsonify({"success": False, "error": "Employee ID may only use letters, numbers, - and _ (2-20 characters)"})
            existing = list(staff_ref.where("employee_id", "==", new_employee_id).limit(1).stream())
            if existing and existing[0].id != staff_id:
                return jsonify({"success": False, "error": f"Employee ID '{new_employee_id}' is already in use"})
            update_fields["employee_id"] = new_employee_id

        if not update_fields:
            return jsonify({"success": False, "error": "Nothing to update"})

        item_ref.update(update_fields)
        updated = item_ref.get().to_dict()
        updated["id"] = staff_id
        return jsonify({"success": True, "staff": updated})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


@app.route("/staff_manage/<rid>/toggle/<staff_id>", methods=["POST"])
def staff_toggle(rid, staff_id):
    if not _restaurant_admin_authorized(rid):
        return jsonify({"success": False, "error": "Unauthorized"}), 401
    try:
        ref = db.collection("restaurants").document(rid).collection("staff_accounts").document(staff_id)
        doc = ref.get()
        if not doc.exists:
            return jsonify({"success": False, "error": "Not found"})
        current = doc.to_dict().get("active", True)
        ref.update({"active": not current})
        return jsonify({"success": True, "active": not current})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


@app.route("/staff_manage/<rid>/delete/<staff_id>", methods=["DELETE"])
def staff_delete(rid, staff_id):
    if not _restaurant_admin_authorized(rid):
        return jsonify({"success": False, "error": "Unauthorized"}), 401
    try:
        db.collection("restaurants").document(rid).collection("staff_accounts").document(staff_id).delete()
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


# =====================================
# 📊 STAFF TRANSACTION REPORT — all-time history (not scoped to "today"),
# for a single waiter or cashier, plus a landscape PDF export.
# =====================================
def _staff_all_time_transactions(rid, employee_id, role):
    """Returns (transactions, total_amount) — every paid order (waiter)
    or every payment processed (cashier) this employee has ever been
    tied to, no date filter. Each transaction: {time, table, waiter, amount}."""
    restaurant_ref = db.collection("restaurants").document(rid)
    transactions = []
    total_amount = 0.0

    if role == "waiter":
        for doc in restaurant_ref.collection("orders") \
                .where("employee_id", "==", employee_id) \
                .where("status", "==", "paid").stream():
            o = doc.to_dict()
            created = o.get("created_at")
            if hasattr(created, "strftime"):
                time_str = created.strftime("%Y-%m-%d %H:%M")
            else:
                time_str = str(created)[:16] if created else "—"
            amount = float(o.get("price", 0))
            transactions.append({
                "time": time_str,
                "table": o.get("table", "—"),
                "waiter": o.get("employee_name", ""),
                "amount": round(amount, 2)
            })
            total_amount += amount
    elif role == "cashier":
        for doc in restaurant_ref.collection("payments") \
                .where("cashier_id", "==", employee_id).stream():
            p = doc.to_dict()
            date_str = p.get("date", "")
            time_str = p.get("time", "")
            amount = float(p.get("amount", 0))
            transactions.append({
                "time": f"{date_str} {time_str}".strip() or "—",
                "table": p.get("table", "—"),
                "waiter": p.get("waiter_name", "—"),
                "amount": round(amount, 2)
            })
            total_amount += amount

    transactions.sort(key=lambda t: t["time"], reverse=True)
    return transactions, round(total_amount, 2)


@app.route("/staff_daily_breakdown/<rid>/<employee_id>")
def staff_daily_breakdown(rid, employee_id):
    """For a WAITER: every day they've ever worked, with that day's
    order count and total — searchable by date on the frontend."""
    if not _restaurant_admin_authorized(rid):
        return jsonify({"success": False, "error": "Unauthorized"}), 401
    try:
        transactions, total = _staff_all_time_transactions(rid, employee_id, "waiter")
        daily = defaultdict(lambda: {"orders": 0, "total": 0.0})
        for t in transactions:
            date_key = t["time"][:10]  # "YYYY-MM-DD HH:MM" -> "YYYY-MM-DD"
            daily[date_key]["orders"] += 1
            daily[date_key]["total"] += t["amount"]

        days = [
            {"date": d, "orders": v["orders"], "total": round(v["total"], 2)}
            for d, v in daily.items()
        ]
        days.sort(key=lambda x: x["date"], reverse=True)

        return jsonify({"success": True, "days": days, "grand_total": total})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


@app.route("/staff_shift_history/<rid>/<employee_id>")
def staff_shift_history(rid, employee_id):
    """For a CASHIER: every shift they've ever worked (clock-in/out
    time) with the exact orders they processed inside each one — order
    ref, table, time, amount — so the shift totals can be reconciled."""
    if not _restaurant_admin_authorized(rid):
        return jsonify({"success": False, "error": "Unauthorized"}), 401
    try:
        restaurant_ref = db.collection("restaurants").document(rid)

        shifts = []
        for doc in restaurant_ref.collection("cashier_shifts") \
                .where("cashier_employee_id", "==", employee_id).stream():
            s = doc.to_dict()
            s["id"] = doc.id
            shifts.append(s)

        payments_by_shift = defaultdict(list)
        for doc in restaurant_ref.collection("payments") \
                .where("cashier_id", "==", employee_id).stream():
            p = doc.to_dict()
            payments_by_shift[p.get("shift_id", "")].append(p)

        result = []
        for s in shifts:
            shift_payments = payments_by_shift.get(s["id"], [])
            orders = [{
                "ref": p.get("payment_id", ""),
                "table": p.get("table", ""),
                "time": f"{p.get('date','')} {p.get('time','')}".strip(),
                "amount": round(float(p.get("amount", 0)), 2)
            } for p in shift_payments]
            orders.sort(key=lambda x: x["time"])
            shift_total = round(sum(o["amount"] for o in orders), 2)

            result.append({
                "shift_code": s.get("shift_code", s["id"][:8]),
                "opened_at": s.get("opened_at", ""),
                "closed_at": s.get("closed_at", ""),
                "status": s.get("status", ""),
                "opening_cash": s.get("opening_cash", 0),
                "expected_cash": s.get("expected_cash", ""),
                "actual_cash": s.get("actual_cash", ""),
                "orders": orders,
                "shift_total": shift_total
            })

        result.sort(key=lambda x: x["opened_at"], reverse=True)
        return jsonify({"success": True, "shifts": result})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


@app.route("/staff_report_data/<rid>/<employee_id>")
def staff_report_data(rid, employee_id):
    if not _restaurant_admin_authorized(rid):
        return jsonify({"success": False, "error": "Unauthorized"}), 401
    try:
        role = request.args.get("role", "waiter")
        staff_docs = list(db.collection("restaurants").document(rid).collection("staff_accounts")
                           .where("employee_id", "==", employee_id).limit(1).stream())
        if not staff_docs:
            return jsonify({"success": False, "error": "Staff account not found"})
        staff = staff_docs[0].to_dict()

        transactions, total_amount = _staff_all_time_transactions(rid, employee_id, role)

        return jsonify({
            "success": True,
            "name": staff.get("name", ""),
            "employee_id": employee_id,
            "role": role,
            "transaction_count": len(transactions),
            "total_amount": total_amount,
            "transactions": transactions[:500]
        })
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


@app.route("/staff_report_pdf/<rid>/<employee_id>")
def staff_report_pdf(rid, employee_id):
    if not _restaurant_admin_authorized(rid):
        return "Unauthorized ❌", 401
    try:
        role = request.args.get("role", "waiter")
        restaurant_doc = db.collection("restaurants").document(rid).get()
        restaurant_name = restaurant_doc.to_dict().get("name", "Restaurant") if restaurant_doc.exists else "Restaurant"

        staff_docs = list(db.collection("restaurants").document(rid).collection("staff_accounts")
                           .where("employee_id", "==", employee_id).limit(1).stream())
        if not staff_docs:
            return "Staff account not found ❌", 404
        staff = staff_docs[0].to_dict()

        transactions, total_amount = _staff_all_time_transactions(rid, employee_id, role)

        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer, pagesize=landscape(letter),
            topMargin=0.5 * inch, bottomMargin=0.5 * inch,
            leftMargin=0.6 * inch, rightMargin=0.6 * inch
        )
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle("RestName", parent=styles["Title"], alignment=TA_CENTER, fontSize=20)
        sub_style = ParagraphStyle("Sub", parent=styles["Normal"], alignment=TA_CENTER, fontSize=11, textColor=colors.HexColor("#555555"))

        story = [
            Paragraph(restaurant_name, title_style),
            Spacer(1, 6),
            Paragraph(
                f"Staff Transaction Report — {staff.get('name','')} ({employee_id}) — {role.capitalize()}",
                sub_style
            ),
            Spacer(1, 4),
            Paragraph(f"Generated {datetime.now().strftime('%Y-%m-%d %H:%M')} — {len(transactions)} transactions", sub_style),
            Spacer(1, 18),
        ]

        table_data = [["Time", "Table", "Waiter", "Amount"]]
        for t in transactions:
            table_data.append([t["time"], str(t["table"]), t["waiter"] or "—", f"${t['amount']:.2f}"])
        table_data.append(["", "", "TOTAL", f"${total_amount:.2f}"])

        tbl = Table(table_data, colWidths=[2.2*inch, 1.3*inch, 2.5*inch, 1.5*inch], repeatRows=1)
        tbl.setStyle(TableStyle([
            # Header row — dark fill + white text so it still reads
            # clearly if printed in black & white (not relying on a
            # light color alone to separate it from the body rows).
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a1a2e")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 10),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
            ("TOPPADDING", (0, 0), (-1, 0), 8),
            # Body — plain borders (visible in grayscale) instead of
            # alternating background colors as the only distinguisher.
            ("FONTSIZE", (0, 1), (-1, -1), 9),
            ("GRID", (0, 0), (-1, -2), 0.5, colors.HexColor("#cccccc")),
            ("ALIGN", (3, 0), (3, -1), "RIGHT"),
            ("ALIGN", (1, 0), (1, -1), "CENTER"),
            # Total row
            ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
            ("FONTSIZE", (0, -1), (-1, -1), 11),
            ("LINEABOVE", (0, -1), (-1, -1), 1, colors.black),
            ("TOPPADDING", (0, -1), (-1, -1), 10),
        ]))
        story.append(tbl)

        if not transactions:
            story.append(Spacer(1, 20))
            story.append(Paragraph("No transactions found for this staff member.", styles["Normal"]))

        doc.build(story)
        pdf_bytes = buffer.getvalue()
        buffer.close()

        filename = f"{restaurant_name}_{staff.get('name','staff')}_report.pdf".replace(" ", "_")
        return Response(
            pdf_bytes,
            mimetype="application/pdf",
            headers={"Content-Disposition": f"inline; filename={filename}"}
        )
    except Exception as e:
        return f"Report PDF Error ❌ {str(e)}", 500


def _date_range_paid_rows(rid, from_date, to_date):
    """Shared by the JSON preview and the Excel export — every PAID
    order in [from_date, to_date] with time/table/items/amount and the
    serving waiter's name/employee_id/phone. Returns (rows, total)."""
    day_start = datetime.strptime(from_date, "%Y-%m-%d")
    day_end = datetime.strptime(to_date, "%Y-%m-%d") + timedelta(days=1)
    restaurant_ref = db.collection("restaurants").document(rid)

    waiter_lookup = {}
    for doc in restaurant_ref.collection("staff_accounts").where("role", "==", "waiter").stream():
        s = doc.to_dict()
        waiter_lookup[s.get("employee_id", "")] = {
            "name": s.get("name", ""),
            "phone": s.get("phone", "") or "—"
        }

    rows = []
    total_amount = 0.0
    # Single-field range filter on created_at only (no equality filter
    # mixed in) — avoids the composite-index requirement; status=='paid'
    # is filtered in Python instead.
    for doc in restaurant_ref.collection("orders") \
            .where("created_at", ">=", day_start) \
            .where("created_at", "<", day_end).stream():
        o = doc.to_dict()
        if str(o.get("status", "")).lower() != "paid":
            continue

        created = o.get("created_at")
        time_str = created.strftime("%Y-%m-%d %H:%M") if hasattr(created, "strftime") else str(created)[:16]
        amount = float(o.get("price", 0))
        eid = o.get("employee_id", "")
        waiter_info = waiter_lookup.get(eid, {"name": o.get("employee_name", "") or "—", "phone": "—"})

        rows.append({
            "time": time_str,
            "table": o.get("table", "—"),
            "items": o.get("items", ""),
            "amount": round(amount, 2),
            "waiter_name": waiter_info["name"] or "—",
            "waiter_id": eid or "—",
            "waiter_phone": waiter_info["phone"]
        })
        total_amount += amount

    rows.sort(key=lambda r: r["time"], reverse=True)
    return rows, round(total_amount, 2)


@app.route("/orders_report_data/<rid>")
def orders_report_data(rid):
    """JSON preview for the date-range report, shown in the browser
    before the admin decides to download the Excel file."""
    if not _restaurant_admin_authorized(rid):
        return jsonify({"success": False, "error": "Unauthorized"}), 401
    try:
        from_date = request.args.get("from", "")
        to_date = request.args.get("to", "")
        if not from_date or not to_date:
            return jsonify({"success": False, "error": "Both dates are required"})

        rows, total_amount = _date_range_paid_rows(rid, from_date, to_date)
        return jsonify({
            "success": True,
            "rows": rows[:500],
            "row_count": len(rows),
            "total_amount": total_amount
        })
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


@app.route("/orders_report_excel/<rid>")
def orders_report_excel(rid):
    """Date-range Excel export of every PAID order: time, table, items,
    amount, and the waiter who served it (name, employee ID, phone)."""
    if not _restaurant_admin_authorized(rid):
        return "Unauthorized ❌", 401
    try:
        from_date = request.args.get("from", "")
        to_date = request.args.get("to", "")
        if not from_date or not to_date:
            return "Both 'from' and 'to' dates are required ❌", 400

        restaurant_doc = db.collection("restaurants").document(rid).get()
        restaurant_name = restaurant_doc.to_dict().get("name", "Restaurant") if restaurant_doc.exists else "Restaurant"

        rows, total_amount = _date_range_paid_rows(rid, from_date, to_date)

        # ---------- Build the workbook ----------
        wb = Workbook()
        ws = wb.active
        ws.title = "Orders Report"
        ws.page_setup.orientation = "landscape"

        ws.merge_cells("A1:G1")
        ws["A1"] = restaurant_name
        ws["A1"].font = Font(bold=True, size=16)
        ws["A1"].alignment = Alignment(horizontal="center")

        ws.merge_cells("A2:G2")
        ws["A2"] = f"Orders Report — {from_date} to {to_date}"
        ws["A2"].font = Font(size=11, color="666666")
        ws["A2"].alignment = Alignment(horizontal="center")

        headers = ["Time", "Table", "Items", "Amount", "Waiter Name", "Waiter Employee ID", "Waiter Phone"]
        header_row = 4
        for col, h in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1A1A2E")
            cell.alignment = Alignment(horizontal="center")

        r = header_row + 1
        for row in rows:
            ws.cell(row=r, column=1, value=row["time"])
            ws.cell(row=r, column=2, value=row["table"])
            ws.cell(row=r, column=3, value=row["items"])
            ws.cell(row=r, column=4, value=row["amount"]).number_format = '"$"#,##0.00'
            ws.cell(row=r, column=5, value=row["waiter_name"])
            ws.cell(row=r, column=6, value=row["waiter_id"])
            ws.cell(row=r, column=7, value=row["waiter_phone"])
            r += 1

        # Total row
        ws.cell(row=r, column=3, value="TOTAL").font = Font(bold=True)
        total_cell = ws.cell(row=r, column=4, value=round(total_amount, 2))
        total_cell.font = Font(bold=True)
        total_cell.number_format = '"$"#,##0.00'

        widths = [17, 8, 34, 12, 18, 16, 16]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        buffer = BytesIO()
        wb.save(buffer)
        excel_bytes = buffer.getvalue()
        buffer.close()

        filename = f"{restaurant_name}_orders_{from_date}_to_{to_date}.xlsx".replace(" ", "_")
        return Response(
            excel_bytes,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        return f"Excel Report Error ❌ {str(e)}", 500


# =====================================
# 🧑‍🍳 WAITER / 💰 CASHIER — GENERIC LOGIN (no rid in URL)
# Used by the homepage portal cards. Looks up the employee_id across
# every restaurant's staff_accounts, then hands off to the same
# restaurant-scoped session + dashboard as the per-restaurant login
# routes below.
#
# NOTE: this deliberately does NOT use db.collection_group("staff_accounts")
# — a collection_group query needs a special "collection group" scoped
# index created in the Firebase console first, otherwise Firestore raises
# FAILED_PRECONDITION and every login silently looks like a wrong
# employee ID / PIN. Looping per-restaurant needs no extra index setup.
# =====================================
def _find_staff_across_restaurants(employee_id, role, pin):
    """Returns (rid, staff_dict) or (None, None)."""
    try:
        for rdoc in db.collection("restaurants").stream():
            for doc in rdoc.reference.collection("staff_accounts") \
                    .where("employee_id", "==", employee_id).stream():
                s = doc.to_dict()
                if s.get("role") == role and s.get("pin") == pin:
                    s["id"] = doc.id
                    return rdoc.id, s
    except Exception as e:
        print("Staff lookup error:", e)
    return None, None


@app.route("/waiter_login", methods=["GET", "POST"])
def waiter_login_generic():
    error = None
    if request.method == "POST":
        employee_id = request.form.get("employee_id", "").strip().upper()
        pin         = request.form.get("pin", "").strip()
        rid, found  = _find_staff_across_restaurants(employee_id, "waiter", pin)

        if not found:
            error = "Wrong employee ID or PIN ❌"
        elif not found.get("active", True):
            error = "This account has been disabled — contact your admin ❌"
        else:
            active, restaurant = restaurant_is_active(rid)
            if not active:
                return render_template("staff_suspended.html", restaurant_name=restaurant.get("name", "Restaurant"), portal="Waiter")
            session["staff_ok"]   = True
            session["staff_role"] = "waiter"
            session["staff_rid"]  = rid
            session["staff_id"]   = found["id"]
            session["staff_name"] = found.get("name", "")
            session["staff_employee_id"] = employee_id
            return redirect(f"/waiter_dashboard/{rid}")

    return render_template("waiter_login.html", rid=None, restaurant_name="Sahal Server POS", error=error)


@app.route("/cashier_login", methods=["GET", "POST"])
def cashier_login_generic():
    error = None
    if request.method == "POST":
        employee_id = request.form.get("employee_id", "").strip().upper()
        pin         = request.form.get("pin", "").strip()
        rid, found  = _find_staff_across_restaurants(employee_id, "cashier", pin)

        if not found:
            error = "Wrong employee ID or PIN ❌"
        elif not found.get("active", True):
            error = "This account has been disabled — contact your admin ❌"
        else:
            active, restaurant = restaurant_is_active(rid)
            if not active:
                return render_template("staff_suspended.html", restaurant_name=restaurant.get("name", "Restaurant"), portal="Cashier")
            session["staff_ok"]   = True
            session["staff_role"] = "cashier"
            session["staff_rid"]  = rid
            session["staff_id"]   = found["id"]
            session["staff_name"] = found.get("name", "")
            session["staff_employee_id"] = employee_id
            return redirect(f"/cashier_dashboard/{rid}")

    return render_template("cashier_login.html", rid=None, restaurant_name="Sahal Server POS", error=error)


# =====================================
# 🧑‍🍳 WAITER — LOGIN / DASHBOARD / LOGOUT
# =====================================
@app.route("/waiter_login/<rid>", methods=["GET", "POST"])
def waiter_login(rid):
    active, restaurant = restaurant_is_active(rid)
    if not restaurant:
        return "Restaurant not found ❌"
    restaurant_name = restaurant.get("name", "Restaurant")

    if not active:
        return render_template("staff_suspended.html", restaurant_name=restaurant_name, portal="Waiter")

    error = None
    if request.method == "POST":
        employee_id = request.form.get("employee_id", "").strip().upper()
        pin         = request.form.get("pin", "").strip()
        found = None
        for doc in db.collection("restaurants").document(rid).collection("staff_accounts") \
                     .where("employee_id", "==", employee_id).stream():
            found = doc.to_dict()
            found["id"] = doc.id
            break

        if not found or found.get("role") != "waiter" or found.get("pin") != pin:
            error = "Wrong employee ID or PIN ❌"
        elif not found.get("active", True):
            error = "This account has been disabled — contact your admin ❌"
        else:
            session["staff_ok"]   = True
            session["staff_role"] = "waiter"
            session["staff_rid"]  = rid
            session["staff_id"]   = found["id"]
            session["staff_name"] = found.get("name", "")
            session["staff_employee_id"] = employee_id
            next_url = request.form.get("next") or request.args.get("next")
            if next_url and next_url.startswith(f"/waiter_dashboard/{rid}"):
                return redirect(next_url)
            return redirect(f"/waiter_dashboard/{rid}")

    return render_template("waiter_login.html", rid=rid, restaurant_name=restaurant_name, error=error, next=request.args.get("next", ""))


# =====================================
# 📱 TABLE QR DEEP-LINK — jumps a logged-in waiter straight to
# "New Order" with the table pre-filled (no typing, no menu digging).
# If not logged in yet, sends them to login first and continues here
# automatically once they authenticate.
# =====================================
@app.route("/waiter_order/<rid>/<table>")
def waiter_order_deeplink(rid, table):
    target = f"/waiter_dashboard/{rid}?table={quote(str(table))}&view=neworder"
    if not session.get("staff_ok") or session.get("staff_role") != "waiter" or session.get("staff_rid") != rid:
        # target itself contains ? and & — it MUST be percent-encoded before
        # being embedded as the value of another query string param, or the
        # outer parser splits it apart and table/view get silently dropped
        # (this was the exact bug: worked when already logged in — no
        # redirect through login needed — but broke on a fresh mobile
        # session that had to log in first).
        return redirect(f"/waiter_login/{rid}?next={quote(target, safe='')}")
    return redirect(target)


@app.route("/waiter_dashboard/<rid>")
def waiter_dashboard(rid):
    if not session.get("staff_ok") or session.get("staff_role") != "waiter" or session.get("staff_rid") != rid:
        return redirect(f"/waiter_login/{rid}")

    active, restaurant = restaurant_is_active(rid)
    if not restaurant:
        return "Restaurant not found ❌"
    if not active:
        session.pop("staff_ok", None)
        return render_template("staff_suspended.html", restaurant_name=restaurant.get("name", "Restaurant"), portal="Waiter")

    restaurant_ref = db.collection("restaurants").document(rid)
    employee_id = session.get("staff_employee_id", "")

    menu = []
    for doc in restaurant_ref.collection("menu").stream():
        m = doc.to_dict()
        m["id"] = doc.id
        menu.append(m)

    today = datetime.now().strftime("%Y-%m-%d")
    my_orders = []
    pending_count = 0
    items_sold = 0
    total_sales = 0.0
    for doc in restaurant_ref.collection("orders") \
            .order_by("created_at", direction=firestore.Query.DESCENDING).limit(150).stream():
        o = doc.to_dict()
        if o.get("employee_id") != employee_id:
            continue
        created = o.get("created_at")
        created_date = created.strftime("%Y-%m-%d") if hasattr(created, "strftime") else str(created)[:10]
        o["id"] = doc.id
        my_orders.append(o)
        if created_date == today:
            status = str(o.get("status", "")).lower()
            if status != "paid":
                pending_count += 1
            else:
                total_sales += float(o.get("price", 0))
                items_sold += sum(int(i.get("qty", 1)) for i in (o.get("cart") or []))

    orders_today = sum(1 for o in my_orders
                        if (o.get("created_at").strftime("%Y-%m-%d") if hasattr(o.get("created_at"), "strftime") else str(o.get("created_at"))[:10]) == today)

    # ---- My Customers: aggregate by customer_phone across this waiter's orders ----
    customer_agg = {}
    for o in my_orders:
        phone = (o.get("customer_phone") or "").strip()
        if not phone:
            continue
        if phone not in customer_agg:
            customer_agg[phone] = {"phone": phone, "visits": 0, "total_spent": 0.0, "last_order": ""}
        customer_agg[phone]["visits"] += 1
        if str(o.get("status", "")).lower() == "paid":
            customer_agg[phone]["total_spent"] += float(o.get("price", 0))
        created = o.get("created_at")
        created_str = created.strftime("%Y-%m-%d %H:%M") if hasattr(created, "strftime") else str(created)
        if created_str > customer_agg[phone]["last_order"]:
            customer_agg[phone]["last_order"] = created_str
    my_customers = sorted(customer_agg.values(), key=lambda x: x["visits"], reverse=True)
    for c in my_customers:
        c["total_spent"] = round(c["total_spent"], 2)

    # ---- Orders still waiting to be printed (not yet marked printed, not paid) ----
    waiting_orders = [o for o in my_orders
                       if not o.get("receipt_printed") and str(o.get("status", "")).lower() != "paid"]

    # ---- This waiter's payments today (for My Transactions + donut) ----
    today_payments = list(restaurant_ref.collection("payments")
                           .where("waiter_employee_id", "==", employee_id)
                           .where("date", "==", today).stream())

    method_totals = {"cash": 0.0, "evc": 0.0, "edahab": 0.0, "card": 0.0, "other": 0.0}
    my_transactions = []
    for doc in today_payments:
        p = doc.to_dict()
        p["id"] = doc.id
        method = str(p.get("method", "other")).lower()
        if method not in method_totals:
            method = "other"
        method_totals[method] += float(p.get("amount", 0))
        my_transactions.append(p)
    my_transactions.sort(key=lambda x: x.get("created_at", ""), reverse=True)

    donut_colors = [("cash", "#1f9e57"), ("evc", "#2f6fed"), ("edahab", "#7b3fb5"),
                     ("card", "#e0651f"), ("other", "#cccccc")]
    if total_sales > 0:
        stops = []
        cursor = 0.0
        for key, color in donut_colors:
            pct = (method_totals[key] / total_sales) * 100
            start = round(cursor, 2)
            end = round(cursor + pct, 2)
            stops.append(f"{color} {start}% {end}%")
            cursor += pct
        donut_gradient = "conic-gradient(" + ", ".join(stops) + ")"
    else:
        donut_gradient = "conic-gradient(#e8ecf4 0% 100%)"

    resp = make_response(render_template(
        "waiter_dashboard.html",
        rid=rid,
        restaurant_name=restaurant.get("name", "Restaurant"),
        staff_name=session.get("staff_name", ""),
        employee_id=employee_id,
        menu=menu,
        my_orders=my_orders[:30],
        orders_today=orders_today,
        pending_count=pending_count,
        items_sold=items_sold,
        total_sales=round(total_sales, 2),
        my_customers=my_customers,
        waiting_orders=waiting_orders,
        my_transactions=my_transactions[:30],
        method_totals={k: round(v, 2) for k, v in method_totals.items()},
        donut_gradient=donut_gradient,
        prefill_table=request.args.get("table", ""),
        open_view=request.args.get("view", ""),
        categories=_get_or_seed_categories(rid)
    ))
    # Categories (and other data here) can change on the cashier's side
    # between page loads — force the browser to always fetch a fresh
    # copy instead of quietly serving a stale one from its cache.
    resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    return resp

@app.route("/mark_receipt_printed/<rid>/<order_id>", methods=["POST"])
def mark_receipt_printed(rid, order_id):
    if not session.get("staff_ok") or session.get("staff_role") != "waiter" or session.get("staff_rid") != rid:
        return jsonify({"success": False, "error": "Unauthorized"}), 401
    try:
        db.collection("restaurants").document(rid).collection("orders").document(order_id) \
          .update({"receipt_printed": True})
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


# The kitchen's own auto-print tracking — separate from mark_receipt_printed
# above, which is the WAITER's customer-copy tracking. No staff-role check
# here: the kitchen screen isn't behind a waiter/cashier login, only the
# shared kitchen password (or nothing, depending on setup), so this just
# needs a valid order under this restaurant.
@app.route("/mark_kitchen_printed/<rid>/<order_id>", methods=["POST"])
def mark_kitchen_printed(rid, order_id):
    try:
        order_ref = db.collection("restaurants").document(rid).collection("orders").document(order_id)
        if not order_ref.get().exists:
            return jsonify({"success": False, "error": "Order not found"})
        order_ref.update({"kitchen_printed": True})
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


@app.route("/waiter_logout")
def waiter_logout():
    for k in ("staff_ok", "staff_role", "staff_rid", "staff_id", "staff_name", "staff_employee_id"):
        session.pop(k, None)
    return redirect("/")


# =====================================
# 💰 CASHIER — LOGIN / DASHBOARD / LOGOUT
# =====================================
@app.route("/cashier_login/<rid>", methods=["GET", "POST"])
def cashier_login(rid):
    active, restaurant = restaurant_is_active(rid)
    if not restaurant:
        return "Restaurant not found ❌"
    restaurant_name = restaurant.get("name", "Restaurant")

    if not active:
        return render_template("staff_suspended.html", restaurant_name=restaurant_name, portal="Cashier")

    error = None
    if request.method == "POST":
        employee_id = request.form.get("employee_id", "").strip().upper()
        pin         = request.form.get("pin", "").strip()
        found = None
        for doc in db.collection("restaurants").document(rid).collection("staff_accounts") \
                     .where("employee_id", "==", employee_id).stream():
            found = doc.to_dict()
            found["id"] = doc.id
            break

        if not found or found.get("role") != "cashier" or found.get("pin") != pin:
            error = "Wrong employee ID or PIN ❌"
        elif not found.get("active", True):
            error = "This account has been disabled — contact your admin ❌"
        else:
            session["staff_ok"]   = True
            session["staff_role"] = "cashier"
            session["staff_rid"]  = rid
            session["staff_id"]   = found["id"]
            session["staff_name"] = found.get("name", "")
            session["staff_employee_id"] = employee_id
            return redirect(f"/cashier_dashboard/{rid}")

    return render_template("cashier_login.html", rid=rid, restaurant_name=restaurant_name, error=error)


@app.route("/cashier_dashboard/<rid>")
def cashier_dashboard(rid):
    if not session.get("staff_ok") or session.get("staff_role") != "cashier" or session.get("staff_rid") != rid:
        return redirect(f"/cashier_login/{rid}")

    active, restaurant = restaurant_is_active(rid)
    if not restaurant:
        return "Restaurant not found ❌"
    if not active:
        session.pop("staff_ok", None)
        return render_template("staff_suspended.html", restaurant_name=restaurant.get("name", "Restaurant"), portal="Cashier")

    restaurant_ref = db.collection("restaurants").document(rid)
    employee_id = session.get("staff_employee_id", "")

    shift_id, shift = get_active_cashier_shift(rid, employee_id)

    if not shift:
        # No open shift — must open one before doing anything else.
        return render_template(
            "cashier_dashboard.html",
            rid=rid,
            restaurant_name=restaurant.get("name", "Restaurant"),
            staff_name=session.get("staff_name", ""),
            employee_id=employee_id,
            has_shift=False
        )

    today = datetime.now().strftime("%Y-%m-%d")
    report_date = request.args.get("report_date", today)
    try:
        rd_obj = datetime.strptime(report_date, "%Y-%m-%d")
    except Exception:
        rd_obj = datetime.now()
        report_date = today
    report_day_start = rd_obj.replace(hour=0, minute=0, second=0, microsecond=0)
    report_day_end = report_day_start + timedelta(days=1)

    menu_items = []
    for mdoc in restaurant_ref.collection("menu").stream():
        m = mdoc.to_dict()
        m["id"] = mdoc.id
        menu_items.append(m)

    categories = _get_or_seed_categories(rid)

    pending_orders = []
    todays_paid_orders = []
    for doc in restaurant_ref.collection("orders") \
            .order_by("created_at", direction=firestore.Query.DESCENDING).limit(200).stream():
        o = doc.to_dict()
        o["id"] = doc.id
        status = str(o.get("status", "")).lower()
        created = o.get("created_at")
        created_date = created.strftime("%Y-%m-%d") if hasattr(created, "strftime") else str(created)[:10]
        if status != "paid":
            pending_orders.append(o)
        elif created_date == today:
            todays_paid_orders.append(o)

    # ---- "Orders Awaiting Payment" broken down by waiter (for the
    # Dhammaan / Waiter toggle) — orders with no employee_id are
    # grouped under "Aan La Qeexin" (Unassigned). Sorted so whichever
    # waiter has the most pending value always sits at the top. ----
    pending_by_waiter = {}
    pending_total_amount = 0.0
    for o in pending_orders:
        wid  = o.get("employee_id") or "__unassigned__"
        name = o.get("employee_name") or "Aan La Qeexin"
        if wid not in pending_by_waiter:
            pending_by_waiter[wid] = {"employee_id": wid, "name": name, "orders": [], "count": 0, "total": 0.0}
        pending_by_waiter[wid]["orders"].append(o)
        pending_by_waiter[wid]["count"] += 1
        amount = float(o.get("price", 0))
        pending_by_waiter[wid]["total"] += amount
        pending_total_amount += amount

    for w in pending_by_waiter.values():
        w["total"] = round(w["total"], 2)
        w["percentage"] = round((w["total"] / pending_total_amount) * 100, 1) if pending_total_amount > 0 else 0.0

    pending_waiters = sorted(pending_by_waiter.values(), key=lambda x: x["total"], reverse=True)

    # ---- Restaurant-wide TODAY stats (all cashiers/shifts) ----
    today_payments = list(restaurant_ref.collection("payments").where("date", "==", today).stream())
    payment_summary = {"cash": 0.0, "evc": 0.0, "edahab": 0.0, "card": 0.0, "other": 0.0}
    today_sales = 0.0
    for doc in today_payments:
        p = doc.to_dict()
        amount = float(p.get("amount", 0))
        method = str(p.get("method", "other")).lower()
        if method not in payment_summary:
            method = "other"
        payment_summary[method] += amount
        today_sales += amount

    today_cash = payment_summary["cash"]
    today_mobile = payment_summary["evc"] + payment_summary["edahab"] + payment_summary["card"] + payment_summary["other"]

    # ---- Donut chart gradient (computed server-side so the template
    # never has to embed Jinja math/conditionals inside a CSS value —
    # linters choke on that and it's fragile to maintain) ----
    donut_colors = [("cash", "#1f9e57"), ("evc", "#2f6fed"), ("edahab", "#7b3fb5"),
                     ("card", "#e0651f"), ("other", "#cccccc")]
    if today_sales > 0:
        stops = []
        cursor = 0.0
        for key, color in donut_colors:
            pct = (payment_summary[key] / today_sales) * 100
            start = round(cursor, 2)
            end = round(cursor + pct, 2)
            stops.append(f"{color} {start}% {end}%")
            cursor += pct
        donut_gradient = "conic-gradient(" + ", ".join(stops) + ")"
    else:
        donut_gradient = "conic-gradient(#e8ecf4 0% 100%)"

    # ---- Top Selling Items (from today's paid orders' cart) ----
    item_agg = {}
    for o in todays_paid_orders:
        for it in (o.get("cart") or []):
            name = it.get("name", "Item")
            qty  = int(it.get("qty", 1))
            price = float(it.get("price", 0))
            if name not in item_agg:
                item_agg[name] = {"name": name, "qty": 0, "sales": 0.0}
            item_agg[name]["qty"] += qty
            item_agg[name]["sales"] += price * qty
    top_items = sorted(item_agg.values(), key=lambda x: x["sales"], reverse=True)[:5]

    # ---- Waiters Performance for the searched report_date (defaults to
    # today) — includes every registered waiter, even ones with zero
    # orders that day, with % share of that day's sales and an "active
    # hours" span (first order -> last order that day, since there's
    # no separate waiter clock-in/out system — this is a real, derived
    # figure, not a fabricated one). ----
    report_orders = []
    for doc in restaurant_ref.collection("orders") \
            .where("created_at", ">=", report_day_start) \
            .where("created_at", "<", report_day_end).stream():
        o = doc.to_dict()
        report_orders.append(o)

    waiter_agg = {}
    for doc in restaurant_ref.collection("staff_accounts").where("role", "==", "waiter").stream():
        s = doc.to_dict()
        wid = s.get("employee_id")
        if wid:
            waiter_agg[wid] = {"employee_id": wid, "name": s.get("name", wid),
                                "orders": 0, "sales": 0.0, "percentage": 0.0,
                                "first_order": None, "last_order": None, "active_hours": "—"}

    report_day_sales_total = 0.0
    for o in report_orders:
        wid = o.get("employee_id")
        if not wid:
            continue
        if wid not in waiter_agg:
            waiter_agg[wid] = {"employee_id": wid, "name": o.get("employee_name", wid),
                                "orders": 0, "sales": 0.0, "percentage": 0.0,
                                "first_order": None, "last_order": None, "active_hours": "—"}
        created = o.get("created_at")
        if hasattr(created, "timestamp"):
            if waiter_agg[wid]["first_order"] is None or created < waiter_agg[wid]["first_order"]:
                waiter_agg[wid]["first_order"] = created
            if waiter_agg[wid]["last_order"] is None or created > waiter_agg[wid]["last_order"]:
                waiter_agg[wid]["last_order"] = created
        if str(o.get("status", "")).lower() == "paid":
            waiter_agg[wid]["orders"] += 1
            amount = float(o.get("price", 0))
            waiter_agg[wid]["sales"] += amount
            report_day_sales_total += amount

    for w in waiter_agg.values():
        if report_day_sales_total > 0:
            w["percentage"] = round((w["sales"] / report_day_sales_total) * 100, 1)
        w["sales"] = round(w["sales"], 2)
        if w["first_order"] and w["last_order"]:
            span = w["last_order"] - w["first_order"]
            total_minutes = int(span.total_seconds() // 60)
            h, m = divmod(total_minutes, 60)
            w["active_hours"] = f"{h}h {m}m" if total_minutes > 0 else "< 1m"

    waiters_performance = sorted(waiter_agg.values(), key=lambda x: x["sales"], reverse=True)

    # ---- THIS cashier's shift-scoped totals ----
    shift_payments = list(restaurant_ref.collection("payments")
                           .where("shift_id", "==", shift_id).stream())

    shift_method_totals = {"cash": 0.0, "evc": 0.0, "edahab": 0.0, "card": 0.0, "other": 0.0}
    shift_transactions = []
    shift_collected = 0.0
    for doc in shift_payments:
        p = doc.to_dict()
        p["id"] = doc.id
        amount = float(p.get("amount", 0))
        method = str(p.get("method", "other")).lower()
        if method not in shift_method_totals:
            method = "other"
        shift_method_totals[method] += amount
        shift_collected += amount
        shift_transactions.append(p)

    shift_transactions.sort(key=lambda x: x.get("created_at", ""), reverse=True)

    opening_cash = float(shift.get("opening_cash", 0))
    expected_cash = opening_cash + shift_method_totals["cash"]

    return render_template(
        "cashier_dashboard.html",
        rid=rid,
        restaurant_name=restaurant.get("name", "Restaurant"),
        staff_name=session.get("staff_name", ""),
        employee_id=employee_id,
        has_shift=True,
        shift=shift,
        shift_id=shift_id,
        opening_cash=round(opening_cash, 2),
        expected_cash=round(expected_cash, 2),
        pending_orders=pending_orders[:30],
        pending_waiters=pending_waiters,
        pending_count=len(pending_orders),
        # restaurant-wide today
        today_sales=round(today_sales, 2),
        today_orders_count=len(todays_paid_orders) + len(pending_orders),
        today_cash=round(today_cash, 2),
        today_mobile=round(today_mobile, 2),
        payment_summary={k: round(v, 2) for k, v in payment_summary.items()},
        donut_gradient=donut_gradient,
        top_items=top_items,
        waiters_performance=waiters_performance,
        report_date=report_date,
        # this cashier's shift
        shift_collected=round(shift_collected, 2),
        shift_orders_count=len(shift_transactions),
        shift_method_totals={k: round(v, 2) for k, v in shift_method_totals.items()},
        shift_transactions=shift_transactions[:30],
        menu_items=menu_items,
        categories=categories
    )


@app.route("/cashier_shift/<rid>/open", methods=["POST"])
def cashier_shift_open(rid):
    if not session.get("staff_ok") or session.get("staff_role") != "cashier" or session.get("staff_rid") != rid:
        return jsonify({"success": False, "error": "Unauthorized"}), 401
    try:
        employee_id = session.get("staff_employee_id", "")
        existing_id, existing = get_active_cashier_shift(rid, employee_id)
        if existing:
            return jsonify({"success": False, "error": "You already have an open shift"})

        data = request.get_json() or {}
        opening_cash = float(data.get("opening_cash", 0))

        date_str = datetime.now().strftime("%Y%m%d")
        shift_code = f"SHIFT-{date_str}-{employee_id.replace('-', '')}"

        shift_ref = db.collection("restaurants").document(rid).collection("cashier_shifts").document()
        shift_ref.set({
            "shift_code": shift_code,
            "cashier_employee_id": employee_id,
            "cashier_name": session.get("staff_name", ""),
            "opening_cash": opening_cash,
            "opened_at": datetime.now().isoformat(),
            "status": "open"
        })
        return jsonify({"success": True, "shift_id": shift_ref.id})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


@app.route("/cashier_shift/<rid>/close", methods=["POST"])
def cashier_shift_close(rid):
    if not session.get("staff_ok") or session.get("staff_role") != "cashier" or session.get("staff_rid") != rid:
        return jsonify({"success": False, "error": "Unauthorized"}), 401
    try:
        employee_id = session.get("staff_employee_id", "")
        shift_id, shift = get_active_cashier_shift(rid, employee_id)
        if not shift:
            return jsonify({"success": False, "error": "No open shift"})

        data = request.get_json() or {}
        actual_cash = float(data.get("actual_cash", 0))

        restaurant_ref = db.collection("restaurants").document(rid)
        payments = list(restaurant_ref.collection("payments").where("shift_id", "==", shift_id).stream())
        cash_sales = sum(float(p.to_dict().get("amount", 0)) for p in payments
                          if str(p.to_dict().get("method", "")).lower() == "cash")
        opening_cash = float(shift.get("opening_cash", 0))
        expected_cash = opening_cash + cash_sales
        difference = round(actual_cash - expected_cash, 2)

        restaurant_ref.collection("cashier_shifts").document(shift_id).update({
            "status": "closed",
            "closed_at": datetime.now().isoformat(),
            "cash_sales": round(cash_sales, 2),
            "expected_cash": round(expected_cash, 2),
            "actual_cash": round(actual_cash, 2),
            "difference": difference,
            "transactions_count": len(payments)
        })
        return jsonify({"success": True, "expected_cash": round(expected_cash, 2),
                         "actual_cash": round(actual_cash, 2), "difference": difference})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


@app.route("/cashier_pay/<rid>/<order_id>", methods=["POST"])
def cashier_pay(rid, order_id):
    if not session.get("staff_ok") or session.get("staff_role") != "cashier" or session.get("staff_rid") != rid:
        return jsonify({"success": False, "error": "Unauthorized"}), 401
    try:
        employee_id = session.get("staff_employee_id", "")
        shift_id, shift = get_active_cashier_shift(rid, employee_id)
        if not shift:
            return jsonify({"success": False, "error": "Open a shift before taking payments"})

        data = request.get_json() or {}
        method = str(data.get("method", "cash")).lower()
        if method not in ("cash", "evc", "edahab", "card", "other"):
            method = "other"

        restaurant_ref = db.collection("restaurants").document(rid)
        order_ref = restaurant_ref.collection("orders").document(order_id)
        order_doc = order_ref.get()
        if not order_doc.exists:
            return jsonify({"success": False, "error": "Order not found"})

        order_data_dict = order_doc.to_dict()
        amount = float(order_data_dict.get("price", 0))
        payment_id = get_next_payment_id(rid)
        now = datetime.now()

        restaurant_ref.collection("payments").document().set({
            "payment_id":     payment_id,
            "order_id":       order_id,
            "restaurant_id":  rid,
            "table":          order_data_dict.get("table", ""),
            "cashier_id":     employee_id,
            "cashier_name":   session.get("staff_name", ""),
            "waiter_employee_id": order_data_dict.get("employee_id", ""),
            "waiter_name":    order_data_dict.get("employee_name", ""),
            "shift_id":       shift_id,
            "amount":         amount,
            "method":         method,
            "status":         "paid",
            "date":           now.strftime("%Y-%m-%d"),
            "time":           now.strftime("%H:%M:%S"),
            "created_at":     now.isoformat()
        })

        order_ref.update({
            "status":          "paid",
            "payment_method":  method,
            "payment_id":      payment_id,
            "paid_by":         employee_id,
            "cashier_id":      employee_id,
            "updated_at":      datetime.utcnow()
        })

        return jsonify({"success": True, "payment_id": payment_id})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


@app.route("/cashier_logout")
def cashier_logout():
    for k in ("staff_ok", "staff_role", "staff_rid", "staff_id", "staff_name", "staff_employee_id"):
        session.pop(k, None)
    return redirect("/")


# =====================================
# 👥 STAFF LIST
# =====================================
@app.route("/staff_list/<rid>")
def staff_list(rid):
    try:
        docs = db.collection("restaurants").document(rid).collection("staff").stream()
        staff = []
        for doc in docs:
            item = doc.to_dict()
            item["id"] = doc.id
            staff.append(item)
        return render_template("staff_list.html", staff=staff)

    except Exception as e:
        return f"Staff list error ❌ {str(e)}"


# =====================================
# 📰 SEND NEWS
# =====================================
@app.route("/send_news/<rid>", methods=["POST"])
def send_news(rid):
    try:
        news_data = {
            "title": request.form["title"],
            "message": request.form["message"],
            "created_at": datetime.now()
        }
        db.collection("restaurants").document(rid).collection("staff_news").add(news_data)
        return redirect("/dashboard/" + rid)

    except Exception as e:
        return f"Send news error ❌ {str(e)}"


# =====================================
# 📰 STAFF NEWS
# =====================================
@app.route("/staff_news/<rid>")
def staff_news(rid):
    try:
        docs = db.collection("restaurants").document(rid).collection("staff_news").stream()
        news = []
        for doc in docs:
            item = doc.to_dict()
            item["id"] = doc.id
            news.append(item)
        return render_template("staff_news.html", news=news)

    except Exception as e:
        return f"Staff news error ❌ {str(e)}"


# =====================================
# 📊 STATS
# =====================================
@app.route("/stats/<rid>")
def stats(rid):
    try:
        conn = sqlite3.connect("database.db")
        c = conn.cursor()

        c.execute("SELECT COUNT(*) FROM orders WHERE restaurant_id=?", (rid,))
        orders = c.fetchone()[0]

        c.execute("SELECT AVG(CAST(price AS FLOAT)) FROM menu WHERE restaurant_id=?", (rid,))
        row = c.fetchone()
        avg_price = row[0] if row and row[0] else 0

        conn.close()

        revenue = orders * avg_price
        profit = round(revenue * 0.7, 2)

        return jsonify({
            "orders": orders,
            "revenue": round(revenue, 2),
            "profit": profit
        })

    except Exception as e:
        return jsonify({"error": str(e)})


# =====================================
# 🔔 GET CALLS
# =====================================
@app.route("/get_calls/<rid>")
def get_calls(rid):
    try:
        conn = sqlite3.connect("database.db")
        c = conn.cursor()
        c.execute("SELECT COUNT(*) FROM waiter_calls WHERE restaurant_id=?", (rid,))
        count = c.fetchone()[0]
        conn.close()
        return jsonify({"count": count})

    except Exception as e:
        return jsonify({"error": str(e)})


# =====================================
# 🏷️ MENU CATEGORIES — cashier-managed, not hardcoded. Seeded once
# with the 4 defaults (plus the 2 already in use by existing menu
# items), then the cashier can add or delete freely from there.
# =====================================
DEFAULT_MENU_CATEGORIES = [
    {"key": "quraac", "label": "🌅 Quraac"},
    {"key": "qado", "label": "☀️ Qado"},
    {"key": "casariyo", "label": "🌤️ Casariyo"},
    {"key": "casho", "label": "🌙 Casho"},
    {"key": "cold_drink", "label": "🥤 Cold Drink"},
    {"key": "hot_drink", "label": "☕ Hot Drink"},
]


def _get_or_seed_categories(rid):
    cats_ref = db.collection("restaurants").document(rid).collection("categories")
    docs = list(cats_ref.stream())
    if not docs:
        for c in DEFAULT_MENU_CATEGORIES:
            cats_ref.add(c)
        docs = list(cats_ref.stream())
    categories = []
    for d in docs:
        c = d.to_dict()
        c["id"] = d.id
        categories.append(c)
    categories.sort(key=lambda x: x.get("label", ""))
    return categories


@app.route("/menu_categories/<rid>")
def menu_categories_list(rid):
    if not session.get("staff_ok") or session.get("staff_role") != "cashier" or session.get("staff_rid") != rid:
        return jsonify({"success": False, "error": "Unauthorized"}), 401
    try:
        return jsonify({"success": True, "categories": _get_or_seed_categories(rid)})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


@app.route("/menu_categories/<rid>/add", methods=["POST"])
def menu_categories_add(rid):
    if not session.get("staff_ok") or session.get("staff_role") != "cashier" or session.get("staff_rid") != rid:
        return jsonify({"success": False, "error": "Unauthorized"}), 401
    try:
        data = request.get_json() or {}
        label = data.get("label", "").strip()
        if not label:
            return jsonify({"success": False, "error": "Category name is required"})

        key = re.sub(r'[^a-z0-9]+', '_', label.lower()).strip('_')
        if not key:
            return jsonify({"success": False, "error": "Category name must include letters or numbers"})

        cats_ref = db.collection("restaurants").document(rid).collection("categories")
        existing = list(cats_ref.where("key", "==", key).limit(1).stream())
        if existing:
            return jsonify({"success": False, "error": f"'{label}' already exists"})

        new_cat = {"key": key, "label": label}
        doc_ref = cats_ref.add(new_cat)
        new_cat["id"] = doc_ref[1].id
        return jsonify({"success": True, "category": new_cat})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


@app.route("/menu_categories/<rid>/delete/<category_id>", methods=["DELETE"])
def menu_categories_delete(rid, category_id):
    if not session.get("staff_ok") or session.get("staff_role") != "cashier" or session.get("staff_rid") != rid:
        return jsonify({"success": False, "error": "Unauthorized"}), 401
    try:
        db.collection("restaurants").document(rid).collection("categories").document(category_id).delete()
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


# =====================================
# ✅ ADD MENU
# =====================================
@app.route("/add_menu/<rid>", methods=["POST"])
def add_menu(rid):
    if not session.get("staff_ok") or session.get("staff_role") != "cashier" or session.get("staff_rid") != rid:
        return jsonify({"success": False, "error": "Unauthorized — only the Cashier can add menu items"}), 401
    try:
        name = request.form.get("name", "").strip()
        price = request.form.get("price", "").strip()
        category = request.form.get("category", "").strip()
        image_file = request.files.get("image")

        if not name or not price:
            return jsonify({"success": False, "error": "Food name and price are required"})

        image_url = ""
        if image_file and image_file.filename:
            image_url = upload_to_firebase_storage(image_file, folder=f"menu/{rid}")

        menu_data = {
            "name": name,
            "price": price,
            "category": category,
            "image": image_url,
            "created_at": datetime.now()
        }

        doc_ref = db.collection("restaurants").document(rid).collection("menu").add(menu_data)
        menu_data["id"] = doc_ref[1].id
        return jsonify({"success": True, "item": menu_data})

    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


@app.route("/edit_menu_item/<rid>/<item_id>", methods=["POST"])
def edit_menu_item(rid, item_id):
    if not session.get("staff_ok") or session.get("staff_role") != "cashier" or session.get("staff_rid") != rid:
        return jsonify({"success": False, "error": "Unauthorized — only the Cashier can edit menu items"}), 401
    try:
        name = request.form.get("name", "").strip()
        price = request.form.get("price", "").strip()
        category = request.form.get("category", "").strip()
        image_file = request.files.get("image")

        if not name or not price:
            return jsonify({"success": False, "error": "Food name and price are required"})

        item_ref = db.collection("restaurants").document(rid).collection("menu").document(item_id)
        if not item_ref.get().exists:
            return jsonify({"success": False, "error": "Menu item not found"})

        update_fields = {"name": name, "price": price, "category": category}
        if image_file and image_file.filename:
            update_fields["image"] = upload_to_firebase_storage(image_file, folder=f"menu/{rid}")

        item_ref.update(update_fields)
        updated = item_ref.get().to_dict()
        updated["id"] = item_id
        return jsonify({"success": True, "item": updated})

    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


@app.route("/delete_menu_item/<rid>/<item_id>", methods=["DELETE"])
def delete_menu_item(rid, item_id):
    if not session.get("staff_ok") or session.get("staff_role") != "cashier" or session.get("staff_rid") != rid:
        return jsonify({"success": False, "error": "Unauthorized — only the Cashier can delete menu items"}), 401
    try:
        db.collection("restaurants").document(rid).collection("menu").document(item_id).delete()
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


# =====================================
# 📢 ADD AD
# =====================================
@app.route("/add_ad/<rid>", methods=["POST"])
def add_ad(rid):
    try:
        restaurant_ref = db.collection("restaurants").document(rid)
        title = request.form.get("title", "").strip()

        image_file = request.files.get("image")
        audio_file = request.files.get("audio")

        image_url = upload_to_firebase_storage(image_file, folder=f"ads/{rid}") if image_file else ""
        audio_url = upload_to_firebase_storage(audio_file, folder=f"ads/{rid}") if audio_file else ""

        restaurant_ref.collection("ads").add({
            "title": title,
            "image": image_url,
            "audio": audio_url,
            "created_at": datetime.utcnow()
        })

        return redirect(f"/dashboard/{rid}")

    except Exception as e:
        print("Add Ad Error:", e)
        return f"Add Ad Error ❌ {str(e)}"


# =====================================
# 📱 GENERATE QR
# =====================================
@app.route("/generate_qr/<rid>", methods=["POST"])
def generate_qr(rid):
    try:
        table = request.form.get("table", "").strip()

        if not table.isdigit():
            return "<p>Table number must be number only ❌</p>"

        restaurant_ref = db.collection("restaurants").document(rid)
        restaurant_doc = restaurant_ref.get()

        if not restaurant_doc.exists:
            return "Restaurant not found ❌"

        restaurant = restaurant_doc.to_dict()
        restaurant_name = restaurant.get("name", "Restaurant")

        filename = f"qr_{rid}_{table}.png"
        qr_folder = os.path.join("static", "qr")
        os.makedirs(qr_folder, exist_ok=True)
        file_path = os.path.join(qr_folder, filename)

        url = f"https://sahalserver.com/waiter_order/{rid}/{table}"

        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_H,
            box_size=10,
            border=4
        )
        qr.add_data(url)
        qr.make(fit=True)

        img = qr.make_image(fill_color="black", back_color="white")
        img.save(file_path)

        return render_template(
            "qr.html",
            rid=rid,
            img=filename,
            url=url,
            table=table,
            restaurant=restaurant_name,
            payment_methods=restaurant.get("payment_methods", [])
        )

    except Exception as e:
        print("QR Error:", e)
        return f"QR Error ❌ {str(e)}"


# =====================================
# 🍽 CLEAN MENU ROUTE
# =====================================
@app.route("/<restaurant_slug>/table-<table_no>")
def clean_table_menu(restaurant_slug, table_no):
    try:
        rid = request.args.get("rid")

        if not rid:
            return "Restaurant ID missing ❌"

        restaurant_ref = db.collection("restaurants").document(rid)
        restaurant_doc = restaurant_ref.get()

        if not restaurant_doc.exists:
            return "Restaurant not found ❌"

        restaurant = restaurant_doc.to_dict()

        menu = []
        menu_docs = restaurant_ref.collection("menu").stream()
        for doc in menu_docs:
            item = doc.to_dict()
            item["id"] = doc.id
            menu.append(item)

        return render_template(
            "customer_menu.html",
            menu=menu,
            table=table_no,
            rid=rid,
            restaurant=restaurant.get("name", "Restaurant")
        )

    except Exception as e:
        print("Menu Error:", e)
        return f"Menu Error ❌ {str(e)}"

# =====================================
# 📦 CREATE ORDER - FINAL FIX
# =====================================
# =====================================
# 🔢 SEQUENTIAL RECEIPT REF (per restaurant)
# Marnaba kuma laabto 0 — 1, 2, 3, ... weligiis kor u socda.
# =====================================
def get_next_receipt_ref(rid):
    counter_ref = db.collection("restaurants").document(rid) \
                    .collection("meta").document("receipt_counter")
    counter_ref.set({"count": firestore.Increment(1)}, merge=True)
    snap = counter_ref.get()
    return snap.to_dict().get("count", 1)


@app.route("/order/<rid>", methods=["POST"])
def create_order(rid):
    try:
        data = request.get_json()
        if not data:
            return jsonify({"error": "No data"}), 400

        table = str(data.get("table", "")).strip()
        cart  = data.get("cart", [])
        customer_phone = str(data.get("customer_phone", "")).strip()

        if not table or not cart:
            return jsonify({"error": "Invalid order"}), 400

        waiter_employee_id = ""
        if session.get("staff_ok") and session.get("staff_role") == "waiter" and session.get("staff_rid") == rid:
            waiter_employee_id = session.get("staff_employee_id", "")

        restaurant_ref = db.collection("restaurants").document(rid)

        # If this same waiter has an OPEN (not-yet-paid, not kitchen-cleared)
        # order still on this table, add the new items to it instead of
        # creating a second separate ticket — matches "same table, same
        # waiter, still open -> add to it" behaviour.
        existing_order_id = None
        existing_cart = []
        if waiter_employee_id:
            for doc in restaurant_ref.collection("orders") \
                    .where("table", "==", table) \
                    .where("employee_id", "==", waiter_employee_id) \
                    .where("status", "in", ["pending", "preparing", "ready"]) \
                    .limit(1).stream():
                existing_order_id = doc.id
                existing_cart = doc.to_dict().get("cart", [])
                break

        if existing_order_id:
            merged_cart = existing_cart + cart
            items_text  = ", ".join([f"{i.get('qty')}x {i.get('name')}" for i in merged_cart])
            total_price = sum(float(i.get("price", 0)) * int(i.get("qty", 1)) for i in merged_cart)

            # The items sent in THIS request only — the delta the kitchen
            # actually needs to prepare, since everything else in
            # merged_cart was already sent and (maybe) already made.
            new_items_text = ", ".join([f"{i.get('qty')}x {i.get('name')}" for i in cart])

            order_ref = restaurant_ref.collection("orders").document(existing_order_id)
            update_fields = {
                "items":            items_text,
                "cart":             merged_cart,
                "price":            total_price,
                "status":           "pending",          # back to kitchen's attention
                "receipt_printed":  False,               # updated — needs reprint
                "kitchen_printed":  False,               # new items — kitchen needs to re-print too
                "last_added_cart":  cart,
                "last_added_items": new_items_text,
                "last_added_at":    datetime.utcnow(),
                "updated_at":       datetime.utcnow()
            }
            if customer_phone:
                update_fields["customer_phone"] = customer_phone
            order_ref.update(update_fields)
            order_id = existing_order_id
        else:
            items_text  = ", ".join([f"{i.get('qty')}x {i.get('name')}" for i in cart])
            total_price = sum(float(i.get("price", 0)) * int(i.get("qty", 1)) for i in cart)

            order_data = {
                "items":      items_text,
                "cart":       cart,
                "table":      table,
                "price":      total_price,
                "status":     "pending",
                "created_at": datetime.utcnow(),
                "kitchen_cleared": False,
                "receipt_printed": False,
                "kitchen_printed": False,
                "receipt_ref": get_next_receipt_ref(rid)
            }
            if customer_phone:
                order_data["customer_phone"] = customer_phone
            if waiter_employee_id:
                order_data["employee_id"]   = waiter_employee_id
                order_data["employee_name"] = session.get("staff_name", "")

            order_ref = restaurant_ref.collection("orders").document()
            order_id  = order_ref.id
            order_ref.set(order_data)

        return jsonify({
            "success":     True,
            "message":     "Order sent ✅",
            "order_id":    order_id,
            "merged":      bool(existing_order_id),
            "receipt_url": f"/receipt/{rid}/{order_id}"
        })

    except Exception as e:
        print("ORDER ERROR:", e)
        return jsonify({"error": str(e)})

# =====================================
# 🔄 UPDATE STATUS
# =====================================
@app.route("/update_status/<rid>/<order_id>/<status>")
def update_status(rid, order_id, status):
    try:
        order_ref = db.collection("restaurants") \
            .document(rid) \
            .collection("orders") \
            .document(order_id)

        order_doc = order_ref.get()

        if not order_doc.exists:
            return jsonify({"success": False, "message": "Order not found ❌"})

        order_ref.update({
            "status": status,
            "updated_at": datetime.utcnow()
        })

        updated_data = order_ref.get().to_dict()

        return jsonify({
            "success": True,
            "message": f"Status updated to {updated_data.get('status')} ✅",
            "status": updated_data.get("status"),
            "order_id": order_id,
            "table": updated_data.get("table"),
            "items": updated_data.get("items")
        })

    except Exception as e:
        print("Update Status Error:", e)
        return jsonify({"success": False, "message": f"Update failed ❌ {str(e)}"})


# =====================================
# 🔔 GET ORDERS COUNT
# =====================================
@app.route("/get_orders_count/<int:rid>")
def get_orders_count(rid):
    try:
        conn = sqlite3.connect("database.db")
        c = conn.cursor()
        c.execute("SELECT COUNT(*) FROM orders WHERE restaurant_id=?", (rid,))
        count = c.fetchone()[0]
        conn.close()
        return jsonify({"count": count})

    except Exception as e:
        return jsonify({"error": str(e)})


# =====================================
# 🔔 CALL WAITER
# =====================================
@app.route("/call_waiter/<rid>", methods=["POST"])
def call_waiter(rid):
    try:
        table = request.form.get("table")
        restaurant_ref = db.collection("restaurants").document(rid)
        restaurant_ref.collection("waiter_calls").add({
            "table": table,
            "created_at": firestore.SERVER_TIMESTAMP
        })
        return "success"

    except Exception as e:
        return str(e)


# =====================================
# 📋 ORDER STATUS
# =====================================
@app.route("/order_status/<rid>")
def order_status(rid):
    try:
        table = request.args.get("table")

        # Two equality filters + order_by on a third field needs a
        # composite index Firestore doesn't have by default — sort in
        # Python instead (same fix as the cashier-shift 400 error).
        docs = list(db.collection("orders") \
            .where("restaurant_id", "==", rid) \
            .where("table_no", "==", table) \
            .stream())

        if docs:
            latest = max(docs, key=lambda d: d.to_dict().get("created_at") or "")
            return latest.to_dict().get("status", "pending")

        return "waiting"

    except Exception as e:
        return str(e)


@app.route("/kitchen/<rid>", methods=["GET", "POST"])
def kitchen(rid):
    try:
        restaurant_ref = db.collection("restaurants").document(rid)
        restaurant_doc = restaurant_ref.get()

        if not restaurant_doc.exists:
            return "Restaurant not found ❌"

        restaurant = restaurant_doc.to_dict()
        real_pass = restaurant.get("kitchen_password", "7890")

        if not restaurant.get("active", True):
            session.pop("kitchen_" + str(rid), None)
            return render_template("staff_suspended.html", restaurant_name=restaurant.get("name", "Restaurant"), portal="Kitchen")

        if request.method == "POST":
            user_pass = request.form.get("password", "").strip()
            if user_pass != str(real_pass).strip():
                return render_template("kitchen_login.html", rid=rid, error="Wrong password ❌")
            session["kitchen_" + str(rid)] = True

        if not session.get("kitchen_" + str(rid)):
            return render_template("kitchen_login.html", rid=rid)

        # ✅ Menu image lookup: { food_name (lowercase): image_url }
        # Sawirrada menu-ga ayaa laga soo helayaa magaca cuntada ee order-ka.
        menu_images = {}
        for mdoc in restaurant_ref.collection("menu").stream():
            md = mdoc.to_dict() or {}
            nm = (md.get("name") or "").strip().lower()
            if nm:
                menu_images[nm] = md.get("image", "")

        def build_display_items(order, cart_override=None):
            """Ka dhig order-ka (ama cart gaar ah) liis nadiif ah: [{name, qty, image}].
            Cart ayaa la door bidayaa (waa nadiif), haddii uusan jirin items
            text-ka ('2x Burger') ayaa la kala qaadayaa."""
            display = []
            cart = cart_override if cart_override is not None else order.get("cart")
            if cart and isinstance(cart, list):
                for c in cart:
                    nm = (c.get("name") or "").strip()
                    if not nm:
                        continue
                    display.append({
                        "name": nm,
                        "qty": c.get("qty", 1),
                        "image": menu_images.get(nm.lower(), "")
                    })
            elif cart_override is None:
                raw = order.get("items", "") or ""
                for part in raw.split(","):
                    part = part.strip()
                    if not part:
                        continue
                    qty, name = 1, part
                    if "x " in part:
                        head, tail = part.split("x ", 1)
                        if head.strip().isdigit():
                            qty = int(head.strip())
                            name = tail.strip()
                    display.append({
                        "name": name,
                        "qty": qty,
                        "image": menu_images.get(name.lower(), "")
                    })
            return display

        order_docs = restaurant_ref.collection("orders") \
            .order_by("created_at", direction=firestore.Query.DESCENDING) \
            .stream()

        orders = []
        for doc in order_docs:
            order = doc.to_dict()
            order["id"] = doc.id

            if order.get("kitchen_cleared") == True:
                continue

            created_at = order.get("created_at")
            if created_at:
                try:
                    order["created_at"] = created_at.astimezone(
                        ZoneInfo("Africa/Mogadishu")
                    ).strftime("%Y-%m-%d %I:%M:%S %p")
                except:
                    order["created_at"] = str(created_at)
            else:
                order["created_at"] = "N/A"

            # ✅ Sawirrada iyo liiska cuntada
            order["display_items"] = build_display_items(order)
            order["main_image"] = next(
                (d["image"] for d in order["display_items"] if d["image"]), ""
            )

            # Newly-added items on this order since it was first sent
            # (present only when a waiter edited an already-open order) —
            # lets the kitchen see just the delta instead of re-reading
            # the whole ticket as if it were brand new.
            last_added_cart = order.get("last_added_cart")
            if last_added_cart:
                order["new_display_items"] = build_display_items(order, cart_override=last_added_cart)
                last_added_at = order.get("last_added_at")
                order["last_added_marker"] = last_added_at.isoformat() if hasattr(last_added_at, "isoformat") else str(last_added_at)
                # Old items always come first in the merged cart (new
                # items are appended on top of it when a waiter edits an
                # order) — so everything before that boundary was already
                # sent/confirmed to the kitchen on an earlier update, and
                # gets struck through in the "All Items" list below.
                old_count = max(0, len(order["display_items"]) - len(order["new_display_items"]))
                for idx, it in enumerate(order["display_items"]):
                    it["already_confirmed"] = idx < old_count
            else:
                order["new_display_items"] = []
                order["last_added_marker"] = ""
                for it in order["display_items"]:
                    it["already_confirmed"] = False

            orders.append(order)

        # ✅ Tirooyinka stat-cards
        stats = {
            "total": len(orders),
            "preparing": sum(1 for o in orders if o.get("status") == "preparing"),
            "ready": sum(1 for o in orders if o.get("status") == "ready"),
        }

        return render_template(
            "kitchen.html",
            orders=orders,
            rid=rid,
            stats=stats,
            restaurant_name=restaurant.get("name", "Restaurant")
        )

    except Exception as e:
        print("Kitchen Error:", e)
        return f"Kitchen error ❌ {str(e)}"


# =====================================
# 🤖 AI CHAT
# =====================================
@app.route("/ai_chat/<int:rid>", methods=["POST"])
def ai_chat(rid):
    try:
        msg = request.form.get("message")
        table = request.form.get("table")
        reply = "Mahadsanid 🙏 fariintaada waa la gudbiyey"

        conn = sqlite3.connect("database.db")
        c = conn.cursor()
        c.execute("""
            INSERT INTO ai_messages (restaurant_id, table_no, message, time)
            VALUES (?,?,?,?)
        """, (rid, table, msg, datetime.now().strftime("%Y-%m-%d %H:%M")))
        conn.commit()
        conn.close()

        return jsonify({"reply": reply})

    except Exception as e:
        return jsonify({"error": str(e)})


# =====================================
# 📊 TODAY STATS
# =====================================
@app.route("/today_stats/<rid>")
def today_stats(rid):
    try:
        today = datetime.now().strftime("%Y-%m-%d")
        yesterday = (datetime.now() - timedelta(days=1)).strftime("%Y-%m-%d")

        conn = sqlite3.connect("database.db")
        c = conn.cursor()

        c.execute("SELECT COUNT(*) FROM orders WHERE restaurant_id=? AND time LIKE ?", (rid, today + "%"))
        today_orders = c.fetchone()[0]

        c.execute("SELECT COUNT(*) FROM orders WHERE restaurant_id=? AND time LIKE ?", (rid, yesterday + "%"))
        yesterday_orders = c.fetchone()[0]

        c.execute("SELECT AVG(CAST(price AS FLOAT)) FROM menu WHERE restaurant_id=?", (rid,))
        row = c.fetchone()
        avg_price = row[0] if row and row[0] else 0

        conn.close()

        today_revenue = round(today_orders * avg_price, 2)
        yesterday_revenue = round(yesterday_orders * avg_price, 2)
        today_profit = round(today_revenue * 0.7, 2)
        yesterday_profit = round(yesterday_revenue * 0.7, 2)
        diff_profit = round(today_profit - yesterday_profit, 2)

        return jsonify({
            "today_orders": today_orders,
            "today_revenue": today_revenue,
            "today_profit": today_profit,
            "yesterday_orders": yesterday_orders,
            "yesterday_revenue": yesterday_revenue,
            "yesterday_profit": yesterday_profit,
            "diff_profit": diff_profit
        })

    except Exception as e:
        return jsonify({"error": str(e)})


# =====================================
# 📊 ANALYTICS PAGE
# =====================================
@app.route("/analytics/<rid>")
def analytics(rid):
    try:
        return render_template("stats.html", rid=rid)
    except Exception as e:
        return f"Analytics error ❌ {e}"

# =====================================
# 📅 ORDERS BY DATE
# =====================================
@app.route("/orders_by_date/<rid>")
def orders_by_date(rid):
    try:
        date = request.args.get("date")

        docs = db.collection("orders") \
            .where("restaurant_id", "==", rid) \
            .stream()

        result = []

        for doc in docs:
            item = doc.to_dict()
            created = item.get("created_at")

            if created:
                try:
                    created_str = created.strftime("%Y-%m-%d")
                    if created_str == date:
                        result.append({
                            "table": item.get("table_no"),
                            "food": item.get("food"),
                            "time": created.strftime("%H:%M")
                        })
                except:
                    pass

        return jsonify({"orders": result, "total": len(result)})

    except Exception as e:
        return jsonify({"error": str(e)})


# =====================================
# 📈 COMPARE TODAY VS YESTERDAY
# =====================================
@app.route("/compare/<rid>")
def compare(rid):
    try:
        today = datetime.now().strftime("%Y-%m-%d")
        yesterday = (datetime.now() - timedelta(days=1)).strftime("%Y-%m-%d")

        conn = sqlite3.connect("database.db")
        c = conn.cursor()

        c.execute("""
            SELECT COUNT(*) FROM orders
            WHERE restaurant_id=? AND date(time)=?
        """, (str(rid), today))
        today_orders = c.fetchone()[0] or 0

        c.execute("""
            SELECT COUNT(*) FROM orders
            WHERE restaurant_id=? AND date(time)=?
        """, (str(rid), yesterday))
        yesterday_orders = c.fetchone()[0] or 0

        c.execute("""
            SELECT AVG(CAST(price AS FLOAT))
            FROM menu WHERE restaurant_id=?
        """, (str(rid),))
        row = c.fetchone()
        avg_price = row[0] if row and row[0] else 0

        conn.close()

        today_total = round(today_orders * avg_price, 2)
        yesterday_total = round(yesterday_orders * avg_price, 2)
        diff = round(today_total - yesterday_total, 2)

        if diff > 0:
            status = "PROFIT 📈"
        elif diff < 0:
            status = "LOSS 📉"
        else:
            status = "EVEN ⚖️"

        return jsonify({
            "today_orders": today_orders,
            "yesterday_orders": yesterday_orders,
            "today": today_total,
            "yesterday": yesterday_total,
            "difference": diff,
            "status": status
        })

    except Exception as e:
        return jsonify({"error": str(e)})


# =====================================
# 🧹 CLEAR ORDERS
# =====================================
@app.route("/clear_orders/<rid>")
def clear_orders(rid):
    try:
        orders_ref = db.collection("restaurants").document(rid).collection("orders")
        docs = orders_ref.stream()

        for doc in docs:
            orders_ref.document(doc.id).update({
                "cleared_from_kitchen": True
            })

        return "OK"

    except Exception as e:
        return f"Error ❌ {str(e)}"


@app.route("/clear_calls/<rid>")
def clear_calls(rid):
    conn = sqlite3.connect("database.db")
    c = conn.cursor()
    c.execute("DELETE FROM waiter_calls WHERE restaurant_id=?", (rid,))
    conn.commit()
    conn.close()
    return "ok"

# =========================
# 🔔 CHECK NEW ORDER
# =========================
last_order_map = {}

@app.route("/check_new_order/<rid>")
def check_new_order(rid):
    try:
        conn = sqlite3.connect("database.db")
        c = conn.cursor()

        c.execute("""
            SELECT id, table_no
            FROM orders
            WHERE restaurant_id=?
            ORDER BY id DESC
            LIMIT 1
        """, (rid,))

        row = c.fetchone()
        conn.close()

        if not row:
            return jsonify({"new_order": False})

        order_id, table = row

        if rid not in last_order_map:
            last_order_map[rid] = order_id
            return jsonify({"new_order": False})

        if order_id != last_order_map[rid]:
            last_order_map[rid] = order_id
            return jsonify({
                "new_order": True,
                "table": table
            })

        return jsonify({"new_order": False})

    except Exception as e:
        return jsonify({"error": str(e)})

@app.route("/receipt/<rid>/<order_id>")
def receipt(rid, order_id):
    try:
        order_ref = db.collection("restaurants").document(rid)\
                      .collection("orders").document(order_id)
        order_doc = order_ref.get()

        if not order_doc.exists:
            return "<h2 style='text-align:center;margin-top:100px;font-family:Arial'>❌ Receipt not found</h2>", 404

        order = order_doc.to_dict()
        rest_doc = db.collection("restaurants").document(rid).get()
        rest = rest_doc.to_dict() if rest_doc.exists else {}

        cart = order.get("cart", [])
        subtotal = float(order.get("price", 0))
        vat = round(subtotal * 0.05, 2)
        total = round(subtotal + vat, 2)

        items = []
        for i in cart:
            qty   = int(i.get("qty", 1))
            price = float(i.get("price", 0))
            items.append({
                "food":  i.get("name", "Item"),
                "qty":   qty,
                "price": price,
                "total": round(qty * price, 2)
            })

        # ✅ Somalia time (UTC+3)
        created_raw = order.get("created_at")
        try:
            created_at = created_raw.astimezone(ZoneInfo("Africa/Mogadishu"))
        except:
            created_at = created_raw

        # Sequential integer ref — orders made before this field existed
        # get one assigned now, once, and it's saved so it never changes again.
        receipt_ref = order.get("receipt_ref")
        if not receipt_ref:
            receipt_ref = get_next_receipt_ref(rid)
            order_ref.update({"receipt_ref": receipt_ref})

        return render_template(
            "receipt.html",
            rid             = rid,
            order_id        = order_id,
            restaurant_name = rest.get("name", "Restaurant"),
            phone           = rest.get("phone", ""),
            payment         = rest.get("payment", ""),
            table           = order.get("table", ""),
            customer_phone  = order.get("customer_phone", ""),
            ref             = receipt_ref,
            items           = items,
            subtotal        = subtotal,
            vat             = vat,
            total           = total,
            created_at      = created_at
        )

    except Exception as e:
        print("Receipt Error:", e)
        return f"Receipt Error ❌ {str(e)}"

@app.route("/dashboard_receipts/<rid>")
def dashboard_receipts(rid):
    try:
        if not session.get("restaurant_login"):
            return redirect("/login")

        order_docs = db.collection("restaurants").document(rid)\
                       .collection("orders")\
                       .order_by("created_at", direction=firestore.Query.DESCENDING)\
                       .stream()

        orders = []
        count = 1
        for doc in order_docs:
            o = doc.to_dict()
            if o.get("kitchen_cleared"):
                continue

            # ✅ FIX: items text u beddel
            items_raw = o.get("items", "")
            if isinstance(items_raw, list):
                items_text = ", ".join(items_raw)
            elif isinstance(items_raw, str):
                items_text = items_raw
            else:
                items_text = ""

            orders.append({
                "order_num":  count,
                "order_id":   doc.id,
                "table":      o.get("table", "?"),
                "items_text": items_text,   # ✅ string
                "price":      o.get("price", 0),
                "status":     o.get("status", "pending"),
                "created_at": o.get("created_at")
            })
            count += 1

        rest_doc = db.collection("restaurants").document(rid).get()
        rest = rest_doc.to_dict() if rest_doc.exists else {}

        return render_template(
            "dashboard_receipts.html",
            orders=orders,
            rid=rid,
            restaurant=rest.get("name", "Restaurant")
        )

    except Exception as e:
        return f"Receipt List Error ❌ {str(e)}"

@app.route("/receipt_view/<rid>/<table>")
def receipt_view(rid, table):
    try:
        order_docs = db.collection("restaurants").document(rid)\
                       .collection("orders")\
                       .order_by("created_at", direction=firestore.Query.DESCENDING)\
                       .stream()

        for doc in order_docs:
            o = doc.to_dict()
            if not o.get("kitchen_cleared"):
                order_id = doc.id
                cart = o.get("cart", [])
                subtotal = float(o.get("price", 0))
                vat = round(subtotal * 0.05, 2)
                total = round(subtotal + vat, 2)

                items = []
                for i in cart:
                    qty = int(i.get("qty", 1))
                    price = float(i.get("price", 0))
                    items.append({
                        "food": i.get("name", "Item"),
                        "qty": qty,
                        "price": price,
                        "total": round(qty * price, 2)
                    })

                rest_doc = db.collection("restaurants").document(rid).get()
                rest = rest_doc.to_dict() if rest_doc.exists else {}

                receipt_ref = o.get("receipt_ref")
                if not receipt_ref:
                    receipt_ref = get_next_receipt_ref(rid)
                    doc.reference.update({"receipt_ref": receipt_ref})

                return render_template(
                    "receipt.html",
                    rid=rid,
                    order_id=order_id,
                    restaurant_name=rest.get("name", "Restaurant"),
                    phone=rest.get("phone", ""),
                    payment=rest.get("payment", ""),
                    table=o.get("table", table),
                    ref=receipt_ref,
                    items=items,
                    subtotal=subtotal,
                    vat=vat,
                    total=total,
                    created_at=o.get("created_at")
                )

        return "<h2 style='font-family:monospace;text-align:center;margin-top:100px'>📭 No orders found</h2>"

    except Exception as e:
        return f"Error ❌ {str(e)}"


# ==========================================
# PHARMACY SALE RECEIPT (printable, same style as restaurant receipt.html)
# ==========================================
@app.route("/pharmacy_receipt/<pid>/<int:sale_ref>")
def pharmacy_receipt(pid, sale_ref):
    try:
        conn = sqlite3.connect(DB_PATH)
        c    = conn.cursor()
        init_pharmacy_sql(conn, c)
        c.execute("""SELECT medicine_name, quantity_sold, selling_price, sale_date
                     FROM pharmacy_sales
                     WHERE pharmacy_id=? AND sale_ref=?
                     ORDER BY id ASC""", (pid, sale_ref))
        rows = c.fetchall()
        conn.close()

        if not rows:
            return "<h2 style='text-align:center;margin-top:100px;font-family:Arial'>❌ Receipt not found</h2>", 404

        items = []
        subtotal = 0.0
        for name, qty, price, _ in rows:
            qty   = int(qty)
            price = float(price)
            line_total = round(qty * price, 2)
            subtotal  += line_total
            items.append({"food": name, "qty": qty, "price": price, "total": line_total})

        vat   = round(subtotal * 0.05, 2)
        total = round(subtotal + vat, 2)

        sale_date_raw = rows[0][3]
        try:
            created_at = datetime.strptime(sale_date_raw, "%Y-%m-%d %H:%M:%S")
        except Exception:
            created_at = None

        pharmacy_name = pid
        phone = ""
        payment_account = ""
        ph_doc = db.collection("pharmacies").where("username", "==", pid).limit(1).stream()
        for d in ph_doc:
            ph = d.to_dict()
            pharmacy_name   = ph.get("pharmacy_name", pid)
            phone           = ph.get("phone", "")
            payment_account = ph.get("payment_account", "")
            break
        else:
            pu_doc = db.collection("pharmacy_users").document(pid).get()
            if pu_doc.exists:
                pu = pu_doc.to_dict()
                pharmacy_name   = pu.get("pharmacy_name", pid)
                phone           = pu.get("phone", "")
                payment_account = pu.get("payment_account", "")

        return render_template(
            "pharmacy_receipt.html",
            pharmacy_name = pharmacy_name,
            phone         = phone,
            payment       = payment_account,
            table         = session.get("pharmacy_name", pid),
            ref           = sale_ref,
            items         = items,
            subtotal      = subtotal,
            vat           = vat,
            total         = total,
            created_at    = created_at
        )
    except Exception as e:
        print("Pharmacy Receipt Error:", e)
        return f"Receipt Error ❌ {str(e)}"


# ==========================================
# 📢 SYSTEM INFO — ADD
# ==========================================
@app.route("/add_info", methods=["POST"])
def add_info():
    try:
        title = request.form.get("title", "").strip()
        content = request.form.get("content", "").strip()

        image_file = request.files.get("image")
        video_file = request.files.get("video")

        image_name = ""
        video_name = ""

        if image_file and image_file.filename:
            image_name = secure_filename(image_file.filename)
            image_file.save(os.path.join("static/uploads", image_name))

        if video_file and video_file.filename:
            video_name = secure_filename(video_file.filename)
            video_file.save(os.path.join("static/uploads", video_name))

        # ==========================================
        # 🔥 SAVE TO FIREBASE
        # ==========================================
        db.collection("system_info").add({

            "title": title,
            "content": content,
            "image": image_name,
            "video": video_name,

            # 🔥 EXTRA INFO
            "date": datetime.now().strftime("%Y-%m-%d"),
            "timestamp": datetime.utcnow(),

            # 🔥 ORDER SYSTEM
            "position": int(time.time())

        })

        return jsonify({
            "success": True,
            "message": "Information saved successfully"
        })

    except Exception as e:

        return jsonify({
            "success": False,
            "error": str(e)
        })

# ==========================================
# 📢 PUBLIC INFO PAGE (READ ONLY)
# ==========================================
@app.route("/info")
def show_info():

    try:

        return render_template("info_public.html")

    except Exception as e:

        print("INFO PAGE ERROR:", e)

        return f"Info Page Error ❌ {str(e)}"


# ==========================================
# 📢 ADMIN INFO PAGE
# ==========================================
@app.route("/admin_info")
def admin_info():

    try:

        return render_template("info.html")

    except Exception as e:

        print("ADMIN INFO ERROR:", e)

        return f"Admin Info Error ❌ {str(e)}"


# ==========================================
# 📢 GET ALL INFO JSON
# ==========================================
@app.route("/get_all_info")
def get_all_info():

    try:

        docs = db.collection("system_info").stream()

        all_info = []

        for doc in docs:

            data = doc.to_dict()

            all_info.append({

                "id": doc.id,
                "title": data.get("title", ""),
                "content": data.get("content", ""),
                "image": data.get("image", ""),
                "video": data.get("video", ""),
                "date": str(data.get("date", "")),
                "position": data.get("position", 0)

            })

        # ✅ SORT BY POSITION
        all_info.sort(
            key=lambda x: x.get("position", 0)
        )

        # ✅ RETURN JSON
        return jsonify({
            "success": True,
            "data": all_info
        })

    except Exception as e:

        print("GET INFO ERROR:", e)

        return jsonify({
            "success": False,
            "error": str(e)
        })


# ==========================================
# 🗑 DELETE INFO (ADMIN ONLY)
# ==========================================
@app.route("/delete_info/<doc_id>", methods=["DELETE"])
def delete_info(doc_id):

    try:

        db.collection("system_info") \
            .document(doc_id) \
            .delete()

        return jsonify({
            "success": True
        })

    except Exception as e:

        return jsonify({
            "success": False,
            "error": str(e)
        })


# ==========================================
# 🔥 UPDATE POSITIONS (ADMIN ONLY)
# ==========================================
@app.route("/update_info_positions", methods=["POST"])
def update_info_positions():

    try:

        data = request.get_json()

        positions = data.get("positions", [])

        for item in positions:

            db.collection("system_info") \
                .document(item["id"]) \
                .update({

                    "position": item["position"]

                })

        return jsonify({
            "success": True
        })

    except Exception as e:

        return jsonify({
            "success": False,
            "error": str(e)
        })


# ==========================================
# ✏ EDIT INFO
# ==========================================
@app.route("/edit_info/<doc_id>")
def edit_info(doc_id):

    try:

        doc = db.collection("system_info") \
            .document(doc_id) \
            .get()

        if not doc.exists:

            return "Info not found"

        data = doc.to_dict()

        data["id"] = doc.id

        return render_template(
            "edit_info.html",
            info=data
        )

    except Exception as e:

        return str(e)


# ==========================================
# 💾 UPDATE INFO
# ==========================================
@app.route("/update_info/<doc_id>", methods=["POST"])
def update_info(doc_id):

    try:

        title = request.form.get("title")
        content = request.form.get("content")

        db.collection("system_info") \
            .document(doc_id) \
            .update({

                "title": title,
                "content": content

            })

        return redirect("/admin_info")

    except Exception as e:

        return str(e)

@app.route("/dashboard_login", methods=["POST"])
def dashboard_login():
    try:
        email    = request.form.get("email", "").strip().lower()
        password = request.form.get("password", "").strip()

        docs = dhibic_db.collection("dashboard_users") \
                        .where("email", "==", email).limit(1).get()

        if len(docs) == 0:
            return jsonify({"success": False, "error": "Email not found"})

        user_data   = docs[0].to_dict()
        db_password = str(user_data.get("password", "")).strip()

        if db_password != password:
            return jsonify({"success": False, "error": "Wrong password"})

        session["dashboard_user"] = email
        return jsonify({"success": True, "redirect": "/view-orders"})

    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500


# ==============================
# VIEW ORDERS PAGE
# ==============================
@app.route("/view-orders")
def view_orders():
    if "dashboard_user" not in session:
        return redirect("/")

    def fmt_ts(ts):
        try:
            return ts.strftime("%Y-%m-%d %H:%M") if ts else ""
        except:
            return str(ts) if ts else ""

    try:
        # ── data_orders ──
        do_docs = dhibic_db.collection("data_orders") \
            .order_by("createdAt", direction=firestore.Query.DESCENDING) \
            .limit(200).get()

        data_orders = []
        for doc in do_docs:
            d = doc.to_dict()
            data_orders.append({
                "docId":         doc.id,
                "referenceId":   d.get("referenceId", ""),
                "senderPhone":   d.get("senderPhone", ""),
                "receiverPhone": d.get("receiverPhone", ""),
                "packageName":   d.get("packageName", ""),
                "packageData":   d.get("packageData", ""),
                "description":   d.get("description", ""),
                "amount":        d.get("amount", "0"),
                "status":        d.get("status", "PENDING"),
                "createdAt":     fmt_ts(d.get("createdAt")),
            })

        # ── orders ──
        or_docs = dhibic_db.collection("orders") \
            .order_by("createdAt", direction=firestore.Query.DESCENDING) \
            .limit(200).get()

        orders = []
        for doc in or_docs:
            d = doc.to_dict()
            orders.append({
                "docId":        doc.id,
                "orderId":      d.get("orderId", ""),
                "customerId":   d.get("customerId", ""),
                "address":      d.get("address", ""),
                "merchantId":   d.get("merchantId", ""),
                "merchantName": d.get("merchantName", ""),
                "merchantPhone":d.get("merchantPhone", ""),
                "deliveryType": d.get("deliveryType", ""),
                "price":        d.get("price", d.get("amount", "0")),
                "status":       d.get("status", "PENDING"),
                "createdAt":    fmt_ts(d.get("createdAt")),
            })

        # ── exchange_orders ──
        ex_docs = dhibic_db.collection("exchange_orders") \
            .order_by("createdAt", direction=firestore.Query.DESCENDING) \
            .limit(200).get()

        exchange_orders = []
        for doc in ex_docs:
            d = doc.to_dict()
            exchange_orders.append({
                "docId":         doc.id,
                "senderNumber":  d.get("senderNumber", ""),
                "receiverNumber":d.get("receiverNumber", ""),
                "fromCompany":   d.get("fromCompany", ""),
                "toCompany":     d.get("toCompany", ""),
                "amount":        d.get("amount", "0"),
                "finalAmount":   d.get("finalAmount", "0"),
                "customerEmail": d.get("customerEmail", ""),
                "status":        d.get("status", "PENDING"),
                "createdAt":     fmt_ts(d.get("createdAt")),
                "approvedAt":    fmt_ts(d.get("approvedAt")),
            })

        return render_template(
            "view_orders.html",
            data_orders=data_orders,
            orders=orders,
            exchange_orders=exchange_orders,
        )

    except Exception as e:
        import traceback; traceback.print_exc()
        return f"Error: {str(e)}", 500


# ==============================
# APPROVE ORDER  (3 collection)
# ==============================
@app.route("/approve-order/<collection>/<doc_id>", methods=["POST"])
def approve_order(collection, doc_id):
    if "dashboard_user" not in session:
        return jsonify({"success": False, "error": "Not logged in"})

    allowed = {"data_orders", "orders", "exchange_orders"}
    if collection not in allowed:
        return jsonify({"success": False, "error": "Invalid collection"})

    try:
        status_val = "approved" if collection == "exchange_orders" else "APPROVED"
        dhibic_db.collection(collection).document(doc_id).update({
            "status": status_val
        })
        return jsonify({"success": True})

    except Exception as e:
        return jsonify({"success": False, "error": str(e)})
    

# ── KITCHEN joins its room on page load ──
@socketio.on("kitchen_join")
def kitchen_join(data):
    rid  = data.get("rid", "")
    room = f"kitchen_{rid}"
    join_room(room)
    print(f"✅ Kitchen joined room: {room}")
    emit("kitchen_ready", {"room": room, "rid": rid})

@app.route("/call_waiter_webrtc/<rid>", methods=["POST"])
def call_waiter_webrtc(rid):
    """Kaydi wicitaanka (SDP-ga la socda) Firestore si kitchen-ku polling-ka ugu arko
    haddii socket-ku ku fashilmo — kani waa backup, socket-ku waa mid degdeg ah."""
    try:
        data = request.get_json(silent=True) or {}
        table = (data.get("table") or request.form.get("table", "") or "").strip()
        sdp = data.get("sdp")
        message = data.get("message", "")

        if not table:
            return jsonify({"success": False, "error": "Missing table"})

        doc = {
            "table": table,
            "status": "ringing",
            "created_at": firestore.SERVER_TIMESTAMP
        }
        if sdp:
            doc["sdp"] = sdp
        if message:
            doc["message"] = message

        db.collection("restaurants").document(rid).collection("active_calls").document(table).set(doc)
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


@app.route("/get_active_calls/<rid>")
def get_active_calls(rid):
    try:
        docs = db.collection("restaurants").document(rid).collection("active_calls").where("status", "==", "ringing").stream()
        calls = []
        for d in docs:
            v = d.to_dict()
            calls.append({
                "table": v.get("table"),
                "sdp": v.get("sdp"),
                "message": v.get("message", "")
            })
        return jsonify({"calls": calls})
    except Exception as e:
        return jsonify({"calls": [], "error": str(e)})

# =========================
# 🪪 CREATE ID CARD
# =========================
@app.route("/create_id_card", methods=["GET", "POST"])
def create_id_card():
    if request.method == "POST" and not session.get("idcard_ok"):
        try:
            doc_ref = db.collection("idCardLogin").document("idcard")
            doc = doc_ref.get()
            real_pass = doc.to_dict().get("password", "") if doc.exists else ""
            entered = request.form.get("password", "").strip()

            if entered != real_pass:
                return render_template("create_id_card.html", error="Wrong password ❌")

            session["idcard_ok"] = True
            return redirect("/create_id_card")

        except Exception as e:
            print("IDCARD LOGIN ERROR:", e)
            return render_template("create_id_card.html", error=f"System Error ❌ {str(e)}")

    if not session.get("idcard_ok"):
        return render_template("create_id_card.html")

    return render_template("create_id_card.html", logged_in=True)


@app.route("/logout_id_card")
def logout_id_card():
    session.pop("idcard_ok", None)
    return redirect("/create_id_card")


# ── Save a newly generated ID card into history ──
@app.route("/save_id_card", methods=["POST"])
def save_id_card():
    if not session.get("idcard_ok"):
        return jsonify({"success": False, "error": "Not authorized"}), 403
    try:
        data = request.get_json(force=True) or {}
        id_no   = (data.get("id_no") or "").strip()
        name    = (data.get("name") or "").strip()
        title   = (data.get("title") or "").strip()
        issue   = (data.get("issue") or "").strip()
        expiry  = (data.get("expiry") or "").strip()
        photo   = data.get("photo") or ""

        if not id_no or not name:
            return jsonify({"success": False, "error": "ID No iyo Name waa waajib"}), 400

        card_data = {
            "id_no": id_no,
            "id_no_lower": id_no.lower(),
            "name": name,
            "name_lower": name.lower(),
            "title": title,
            "issue": issue,
            "expiry": expiry,
            "photo": photo,
            "printed": False,
            "created_at": firestore.SERVER_TIMESTAMP,
        }

        db.collection("id_cards").document(id_no).set(card_data, merge=True)
        return jsonify({"success": True, "id": id_no})
    except Exception as e:
        print("SAVE ID CARD ERROR:", e)
        return jsonify({"success": False, "error": str(e)}), 500


# ── List / search saved ID cards ──
@app.route("/list_id_cards", methods=["GET"])
def list_id_cards():
    if not session.get("idcard_ok"):
        return jsonify({"success": False, "error": "Not authorized"}), 403
    try:
        q = (request.args.get("q") or "").strip().lower()
        docs = db.collection("id_cards").order_by(
            "created_at", direction=firestore.Query.DESCENDING
        ).stream()

        cards = []
        for d in docs:
            item = d.to_dict()
            item["doc_id"] = d.id
            ca = item.get("created_at")
            item["created_at"] = ca.isoformat() if hasattr(ca, "isoformat") else None

            if q:
                if q not in (item.get("id_no_lower") or "") and q not in (item.get("name_lower") or ""):
                    continue

            cards.append(item)

        return jsonify({"success": True, "cards": cards})
    except Exception as e:
        print("LIST ID CARDS ERROR:", e)
        return jsonify({"success": False, "error": str(e)}), 500


# ── Mark an ID card as already printed ──
@app.route("/mark_id_card_printed/<doc_id>", methods=["POST"])
def mark_id_card_printed(doc_id):
    if not session.get("idcard_ok"):
        return jsonify({"success": False, "error": "Not authorized"}), 403
    try:
        db.collection("id_cards").document(doc_id).update({"printed": True})
        return jsonify({"success": True})
    except Exception as e:
        print("MARK PRINTED ERROR:", e)
        return jsonify({"success": False, "error": str(e)}), 500


# ── Delete a single ID card ──
@app.route("/delete_id_card/<doc_id>", methods=["POST"])
def delete_id_card(doc_id):
    if not session.get("idcard_ok"):
        return jsonify({"success": False, "error": "Not authorized"}), 403
    try:
        db.collection("id_cards").document(doc_id).delete()
        return jsonify({"success": True})
    except Exception as e:
        print("DELETE ID CARD ERROR:", e)
        return jsonify({"success": False, "error": str(e)}), 500


# ── Delete ALL ID cards ──
@app.route("/delete_all_id_cards", methods=["POST"])
def delete_all_id_cards():
    if not session.get("idcard_ok"):
        return jsonify({"success": False, "error": "Not authorized"}), 403
    try:
        docs = db.collection("id_cards").stream()
        for doc in docs:
            doc.reference.delete()
        return jsonify({"success": True})
    except Exception as e:
        print("DELETE ALL ID CARDS ERROR:", e)
        return jsonify({"success": False, "error": str(e)}), 500


# =========================
# ⚙️ ID CARD SETTINGS (change login password)
# =========================
@app.route("/id_card_settings", methods=["GET"])
def id_card_settings():
    if not session.get("idcard_ok"):
        return render_template("id_card_settings.html")
    return render_template("id_card_settings.html", logged_in=True)


@app.route("/change_id_card_password", methods=["POST"])
def change_id_card_password():
    if not session.get("idcard_ok"):
        return jsonify({"success": False, "error": "Not authorized"}), 403
    try:
        data = request.get_json(force=True) or {}
        current_password = (data.get("current_password") or "").strip()
        new_password     = (data.get("new_password") or "").strip()

        if not current_password or not new_password:
            return jsonify({"success": False, "error": "Fadlan buuxi dhammaan fields-ka"}), 400

        doc_ref = db.collection("idCardLogin").document("idcard")
        doc = doc_ref.get()
        real_pass = doc.to_dict().get("password", "") if doc.exists else ""

        if current_password != real_pass:
            return jsonify({"success": False, "error": "Password-ka hadda jira waa khalad"}), 400

        doc_ref.set({"password": new_password}, merge=True)
        return jsonify({"success": True})
    except Exception as e:
        print("CHANGE ID CARD PASSWORD ERROR:", e)
        return jsonify({"success": False, "error": str(e)}), 500


# =========================
# 🗂️ ALL ID CARDS (list/search/print/delete table page)
# =========================
@app.route("/all_id_cards", methods=["GET"])
def all_id_cards_page():
    if not session.get("idcard_ok"):
        return render_template("all_id_cards.html")
    return render_template("all_id_cards.html", logged_in=True)
    
    
# ── CUSTOMER joins its room ──
@socketio.on("customer_join")
def customer_join(data):
    rid   = data.get("rid", "")
    table = data.get("table", "")
    room  = f"customer_{rid}_{table}"
    join_room(room)
    emit("customer_ready", {"room": room})

# ── Customer → Kitchen: offer ──
@socketio.on("webrtc_offer")
def webrtc_offer(data):
    rid   = data.get("rid", "")
    table = data.get("table", "")
    data["table"] = table
    emit("webrtc_offer", data, to=f"kitchen_{rid}")

# ── Customer → Kitchen: ICE ──
@socketio.on("webrtc_ice_customer")
def webrtc_ice_customer(data):
    rid = data.get("rid", "")
    emit("webrtc_ice_customer", data, to=f"kitchen_{rid}")

# ── Customer → Kitchen: end ──
@socketio.on("webrtc_end_customer")
def webrtc_end_customer(data):
    rid   = data.get("rid", "")
    table = data.get("table", "")
    if rid and table:
        try:
            db.collection("restaurants").document(rid).collection("active_calls").document(table).delete()
        except Exception:
            pass
    emit("webrtc_end", data, to=f"kitchen_{rid}")

# ── Kitchen → Customer: answer ──
@socketio.on("webrtc_answer")
def webrtc_answer(data):
    rid   = data.get("rid", "")
    table = data.get("table", "")
    if rid and table:
        try:
            db.collection("restaurants").document(rid).collection("active_calls").document(table).delete()
        except Exception:
            pass
    emit("webrtc_answer", data, to=f"customer_{rid}_{table}")

# ── Kitchen → Customer: ICE ──
@socketio.on("webrtc_ice_kitchen")
def webrtc_ice_kitchen(data):
    rid   = data.get("rid", "")
    table = data.get("table", "")
    emit("webrtc_ice_kitchen", data, to=f"customer_{rid}_{table}")

# ── Kitchen → Customer: end ──
@socketio.on("webrtc_end_kitchen")
def webrtc_end_kitchen(data):
    rid   = data.get("rid", "")
    table = data.get("table", "")
    if rid and table:
        try:
            db.collection("restaurants").document(rid).collection("active_calls").document(table).delete()
        except Exception:
            pass
    emit("webrtc_end", data, to=f"customer_{rid}_{table}")

# ==========================================
# 💊 PHARMACY ROUTES — UPDATED VERSION
# ==========================================
# Medicines  → Firestore: pharmacy_product/{username}/medicines
# Sales      → SQLite
# Debts      → SQLite
# Images     → static/uploads/ (filename stored in Firestore)
# ==========================================

PHARMACY_IMG_FOLDER = "static/uploads"
os.makedirs(PHARMACY_IMG_FOLDER, exist_ok=True)


# ==========================================
# GET PHARMACY FIRESTORE REF
# ==========================================
def get_pharmacy_ref():
    """
    Returns reference to: pharmacy_product/{username}/medicines
    The document ID (username) is always the pharmacy's own username.
    """
    username = session.get("pharmacy_user", "")
    if not username:
        username = session.get("pharmacy_id", "unknown")
    return (
        db.collection("pharmacy_product")
          .document(username)
          .collection("medicines")
    )


def get_username():
    """Returns the current pharmacy's username (used as document ID)."""
    return session.get("pharmacy_user", session.get("pharmacy_id", "unknown"))


def get_next_pharmacy_sale_ref(username):
    """Sequential per-pharmacy integer ref number for sale receipts (never resets)."""
    counter_ref = db.collection("pharmacy_product").document(username) \
                    .collection("meta").document("sale_counter")
    counter_ref.set({"count": firestore.Increment(1)}, merge=True)
    snap = counter_ref.get()
    return snap.to_dict().get("count", 1)


# ==========================================
# INIT PHARMACY SALES/DEBTS (SQLite only)
# ==========================================
def init_pharmacy_sql(conn, c):
    c.execute("""
        CREATE TABLE IF NOT EXISTS pharmacy_users (
            id       INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE,
            password TEXT,
            role     TEXT DEFAULT 'pharmacist'
        )
    """)
    c.execute("""
        CREATE TABLE IF NOT EXISTS pharmacy_sales (
            id             INTEGER PRIMARY KEY AUTOINCREMENT,
            pharmacy_id    TEXT,
            sale_ref       INTEGER,
            medicine_id    TEXT,
            medicine_name  TEXT,
            barcode        TEXT,
            quantity_sold  INTEGER,
            cost_price     REAL,
            selling_price  REAL,
            profit         REAL,
            sale_date      TEXT DEFAULT CURRENT_TIMESTAMP
        )
    """)
    # sale_ref added after original table creation — backfill for older DBs
    try:
        c.execute("ALTER TABLE pharmacy_sales ADD COLUMN sale_ref INTEGER")
    except Exception:
        pass
    c.execute("""
        CREATE TABLE IF NOT EXISTS pharmacy_debts (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            pharmacy_id TEXT,
            name        TEXT NOT NULL,
            phone       TEXT,
            type        TEXT DEFAULT 'cash',
            amount      REAL DEFAULT 0,
            description TEXT,
            date        TEXT,
            status      TEXT DEFAULT 'unpaid',
            created_at  TEXT DEFAULT CURRENT_TIMESTAMP
        )
    """)
    conn.commit()


# ==========================================
# PHARMACY LOGIN
# ==========================================
@app.route("/pharmacy_login", methods=["GET", "POST"])
def pharmacy_login():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()
        try:
            today = datetime.now().strftime("%Y-%m-%d")
            for doc in db.collection("pharmacies").where("username", "==", username).stream():
                ph = doc.to_dict()
                if ph.get("password") == password:
                    expiry = ph.get("expiry_date", "")
                    session["pharmacy_ok"]     = True
                    session["pharmacy_user"]   = username
                    session["pharmacy_name"]   = ph.get("pharmacy_name", username)
                    session["pharmacy_id"]     = doc.id
                    session["pharmacy_expiry"] = expiry
                    if expiry and expiry < today:
                        session["pharmacy_suspended"] = True
                    elif not ph.get("active", True):
                        session["pharmacy_suspended"] = True
                    else:
                        session.pop("pharmacy_suspended", None)
                    return redirect("/pharmacy")
        except Exception as e:
            print("Pharmacy login pharmacies error:", e)
        try:
            pu_doc = db.collection("pharmacy_users").document(username).get()
            if pu_doc.exists:
                pu = pu_doc.to_dict()
                if pu.get("password") == password:
                    session["pharmacy_ok"]   = True
                    session["pharmacy_user"] = username
                    session["pharmacy_name"] = pu.get("pharmacy_name", username)
                    session["pharmacy_id"]   = pu.get("pharmacy_id", username)
                    return redirect("/pharmacy")
        except Exception as e:
            print("Pharmacy login pharmacy_users error:", e)
        try:
            conn = sqlite3.connect(DB_PATH)
            c    = conn.cursor()
            c.execute("""CREATE TABLE IF NOT EXISTS pharmacy_users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT UNIQUE, password TEXT, role TEXT DEFAULT 'pharmacist')""")
            c.execute("SELECT * FROM pharmacy_users WHERE username=? AND password=?", (username, password))
            user = c.fetchone()
            conn.close()
            if user:
                session["pharmacy_ok"]   = True
                session["pharmacy_user"] = username
                session["pharmacy_name"] = username
                session["pharmacy_id"]   = username
                return redirect("/pharmacy")
        except Exception as e:
            print("Pharmacy login SQLite error:", e)
        return render_template("pharmacy_login.html", error="Wrong username or password")
    return render_template("pharmacy_login.html")


# ==========================================
# PHARMACY LOGOUT
# ==========================================
@app.route("/pharmacy/logout")
def pharmacy_logout():
    session.pop("pharmacy_ok",     None)
    session.pop("pharmacy_user",   None)
    session.pop("pharmacy_name",   None)
    session.pop("pharmacy_id",     None)
    session.pop("pharmacy_expiry", None)
    return redirect("/pharmacy_login")


# ==========================================
# PHARMACY DASHBOARD
# ==========================================
@app.route("/pharmacy")
def pharmacy():
    if not session.get("pharmacy_ok"):
        return redirect("/pharmacy_login")
    if session.get("pharmacy_suspended"):
        pharmacy_id = session.get("pharmacy_id", "")
        pharmacy_fee = 0.0
        if pharmacy_id:
            ph_doc = db.collection("pharmacies").document(pharmacy_id).get()
            if ph_doc.exists:
                pharmacy_fee = _monthly_fee_for("pharmacy", ph_doc.to_dict())
        return render_template(
            "renew.html",
            entity_type    = "pharmacy",
            entity_label   = "Pharmacy",
            entity_id      = pharmacy_id,
            business_name  = session.get("pharmacy_name", ""),
            monthly_fee    = pharmacy_fee,
            payment_number = ADMIN_PAYMENT_NUMBER,
            logout_url     = "/pharmacy/logout"
        )
    try:
        ref       = get_pharmacy_ref()
        medicines = []
        for doc in ref.stream():
            m = doc.to_dict()
            medicines.append((
                doc.id,
                m.get("name", ""),
                m.get("barcode", ""),
                m.get("cost_price", 0),
                m.get("selling_price", 0),
                m.get("stock_quantity", 0),
                m.get("expiry_date", ""),
                m.get("category", "General"),
                m.get("created_at", ""),
                m.get("image", ""),       # filename for static/uploads
                m.get("imageUrl", "")     # optional external URL
            ))
        today      = datetime.now().strftime("%Y-%m-%d")
        alert_date = (datetime.now() + timedelta(days=90)).strftime("%Y-%m-%d")
        pid        = get_username()
        conn = sqlite3.connect(DB_PATH)
        c    = conn.cursor()
        init_pharmacy_sql(conn, c)
        c.execute("""SELECT COUNT(*), SUM(quantity_sold), SUM(profit)
                     FROM pharmacy_sales WHERE pharmacy_id=? AND date(sale_date)=?""", (pid, today))
        row          = c.fetchone()
        today_sales  = row[0] or 0
        today_qty    = row[1] or 0
        today_profit = round(row[2] or 0, 2)
        c.execute("""SELECT medicine_name, SUM(quantity_sold) as total
                     FROM pharmacy_sales WHERE pharmacy_id=? AND date(sale_date)=?
                     GROUP BY medicine_name ORDER BY total DESC LIMIT 5""", (pid, today))
        top_selling = c.fetchall()
        conn.close()
        expired       = [m for m in medicines if m[6] and m[6] < today]
        expiry_alerts = [m for m in medicines if m[6] and today <= m[6] <= alert_date]
        low_stock     = [m for m in medicines if int(m[5]) <= 3]
        return render_template(
            "pharmacy.html",
            medicines=medicines, today_sales=today_sales, today_qty=today_qty,
            today_profit=today_profit, expiry_alerts=expiry_alerts, expired=expired,
            low_stock=low_stock, top_selling=top_selling, today=today,
            now_date=today, expiry_warn=alert_date,
            pid=pid,
            broadcasts=_get_active_broadcasts("pharmacy", pid)
        )
    except Exception as e:
        return f"Pharmacy Error: {str(e)}"


# ==========================================
# ADD MEDICINE → pharmacy_product/{username}/medicines
# ==========================================
@app.route("/pharmacy/add_medicine", methods=["POST"])
def add_medicine():
    if not session.get("pharmacy_ok"):
        return jsonify({"success": False, "error": "Not logged in"}), 401
    try:
        name          = request.form.get("name", "").strip()
        barcode       = request.form.get("barcode", "").strip()
        cost_price    = float(request.form.get("cost_price", 0))
        selling_price = float(request.form.get("selling_price", 0))
        stock_qty     = int(request.form.get("stock_quantity", 0))
        expiry_date   = request.form.get("expiry_date", "").strip()
        category      = request.form.get("category", "General").strip()
        if not name:
            return jsonify({"success": False, "error": "Medicine name required"})

        image_filename = ""
        image_file     = request.files.get("image")
        if image_file and image_file.filename:
            ext       = os.path.splitext(image_file.filename)[1].lower()
            safe_name = secure_filename(f"{name.replace(' ','_')}_{int(datetime.now().timestamp())}{ext}")
            image_file.save(os.path.join(PHARMACY_IMG_FOLDER, safe_name))
            image_filename = safe_name

        ref = get_pharmacy_ref()
        ref.add({
            "name":           name,
            "barcode":        barcode or "",
            "cost_price":     cost_price,
            "selling_price":  selling_price,
            "stock_quantity": stock_qty,
            "expiry_date":    expiry_date,
            "category":       category,
            "image":          image_filename,
            "imageUrl":       "",
            "created_at":     datetime.now().isoformat()
        })
        return jsonify({"success": True, "message": "Medicine added"})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


# ==========================================
# SEARCH MEDICINE → pharmacy_product/{username}/medicines
# ==========================================
@app.route("/pharmacy/search")
def search_medicine():
    if not session.get("pharmacy_ok"):
        return jsonify({"error": "Not logged in"}), 401
    query = request.args.get("q", "").strip()
    try:
        ref     = get_pharmacy_ref()
        results = []
        q_low   = query.lower()
        for doc in ref.stream():
            m    = doc.to_dict()
            name = m.get("name", "")
            bc   = m.get("barcode", "")
            if not query or bc == query or q_low in name.lower():
                results.append({
                    "medicine_id":    doc.id,
                    "name":           name,
                    "barcode":        bc,
                    "cost_price":     m.get("cost_price", 0),
                    "selling_price":  m.get("selling_price", 0),
                    "stock_quantity": m.get("stock_quantity", 0),
                    "expiry_date":    m.get("expiry_date", ""),
                    "category":       m.get("category", "General"),
                    "image":          m.get("image", ""),
                    "imageUrl":       m.get("imageUrl", "")
                })
        return jsonify(results)
    except Exception as e:
        return jsonify({"error": str(e)})


# ==========================================
# SELL MEDICINE → SQLite sales, update Firestore stock
# ==========================================
@app.route("/pharmacy/sell", methods=["POST"])
def sell_medicine():
    if not session.get("pharmacy_ok"):
        return jsonify({"success": False, "error": "Not logged in"}), 401
    try:
        data = request.get_json()
        cart = data.get("cart", [])
        if not cart:
            return jsonify({"success": False, "error": "Cart is empty"})
        ref          = get_pharmacy_ref()
        pid          = get_username()
        total_profit = 0
        sale_ref     = get_next_pharmacy_sale_ref(pid)
        sale_time    = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        conn         = sqlite3.connect(DB_PATH)
        c            = conn.cursor()
        init_pharmacy_sql(conn, c)
        for item in cart:
            med_id     = item.get("medicine_id")
            qty        = int(item.get("quantity", 1))
            paid_price = float(item.get("price", 0))
            med_doc    = ref.document(str(med_id)).get()
            if not med_doc.exists:
                conn.close()
                return jsonify({"success": False, "error": f"Medicine not found: {med_id}"})
            m     = med_doc.to_dict()
            name  = m.get("name", "")
            bc    = m.get("barcode", "")
            cost  = float(m.get("cost_price", 0))
            stock = int(m.get("stock_quantity", 0))
            if qty > stock:
                conn.close()
                return jsonify({"success": False, "error": f"Not enough stock for {name}. Available: {stock}"})
            profit       = (paid_price - cost) * qty
            total_profit += profit
            ref.document(str(med_id)).update({"stock_quantity": stock - qty})
            c.execute("""INSERT INTO pharmacy_sales
                (pharmacy_id, sale_ref, medicine_id, medicine_name, barcode,
                 quantity_sold, cost_price, selling_price, profit, sale_date)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (pid, sale_ref, med_id, name, bc, qty, cost, paid_price,
                 profit, sale_time))
        conn.commit()
        conn.close()
        return jsonify({"success": True, "message": "Sale recorded",
                        "total_profit": round(total_profit, 2),
                        "items_sold": len(cart),
                        "sale_ref": sale_ref,
                        "receipt_url": f"/pharmacy_receipt/{pid}/{sale_ref}"})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


# ==========================================
# EDIT MEDICINE → Firestore
# ==========================================
@app.route("/pharmacy/edit/<med_id>", methods=["PUT"])
def edit_medicine(med_id):
    if not session.get("pharmacy_ok"):
        return jsonify({"success": False, "error": "Not logged in"}), 401
    try:
        data = request.get_json()
        ref  = get_pharmacy_ref()
        ref.document(med_id).update({
            "name":           data.get("name"),
            "barcode":        data.get("barcode", ""),
            "cost_price":     float(data.get("cost_price", 0)),
            "selling_price":  float(data.get("selling_price", 0)),
            "stock_quantity": int(data.get("stock_quantity", 0)),
            "expiry_date":    data.get("expiry_date", ""),
            "category":       data.get("category", "General")
        })
        return jsonify({"success": True, "message": "Updated"})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


# ==========================================
# DELETE MEDICINE → Firestore
# ==========================================
@app.route("/pharmacy/delete/<med_id>", methods=["DELETE"])
def delete_medicine(med_id):
    if not session.get("pharmacy_ok"):
        return jsonify({"success": False, "error": "Not logged in"}), 401
    try:
        ref = get_pharmacy_ref()
        ref.document(med_id).delete()
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


# ==========================================
# ALERTS → Firestore
# ==========================================
@app.route("/pharmacy/alerts")
def pharmacy_alerts():
    if not session.get("pharmacy_ok"):
        return jsonify({"error": "Not logged in"}), 401
    try:
        today      = datetime.now().strftime("%Y-%m-%d")
        alert_date = (datetime.now() + timedelta(days=90)).strftime("%Y-%m-%d")
        ref        = get_pharmacy_ref()
        expiring = []; expired = []; low_st = []
        for doc in ref.stream():
            m     = doc.to_dict()
            exp   = m.get("expiry_date", "")
            stock = int(m.get("stock_quantity", 0))
            entry = {"medicine_id": doc.id, "name": m.get("name",""), "stock": stock, "expiry": exp}
            if exp:
                if exp < today: expired.append(entry)
                elif exp <= alert_date: expiring.append(entry)
            if stock <= 3: low_st.append(entry)
        return jsonify({"expiring_soon": expiring, "expired": expired, "low_stock": low_st,
                        "total_alerts": len(expiring) + len(expired) + len(low_st)})
    except Exception as e:
        return jsonify({"error": str(e)})


# ==========================================
# REPORT → SQLite
# ==========================================
@app.route("/pharmacy/report")
def pharmacy_report():
    if not session.get("pharmacy_ok"):
        return jsonify({"error": "Not logged in"}), 401
    try:
        pid       = get_username()
        date_from = request.args.get("from", datetime.now().strftime("%Y-%m-%d"))
        date_to   = request.args.get("to",   datetime.now().strftime("%Y-%m-%d"))
        conn      = sqlite3.connect(DB_PATH)
        c         = conn.cursor()
        init_pharmacy_sql(conn, c)
        c.execute("""SELECT COUNT(*), SUM(quantity_sold),
                     SUM(selling_price * quantity_sold),
                     SUM(cost_price    * quantity_sold),
                     SUM(profit)
                     FROM pharmacy_sales
                     WHERE pharmacy_id=? AND date(sale_date) BETWEEN ? AND ?""",
                  (pid, date_from, date_to))
        row = c.fetchone()
        c.execute("""SELECT medicine_name,
                     SUM(quantity_sold) as qty,
                     SUM(profit)        as profit
                     FROM pharmacy_sales
                     WHERE pharmacy_id=? AND date(sale_date) BETWEEN ? AND ?
                     GROUP BY medicine_name ORDER BY qty DESC LIMIT 5""",
                  (pid, date_from, date_to))
        top_medicines = [{"name": r[0], "qty": r[1], "profit": round(r[2], 2)}
                         for r in c.fetchall()]
        c.execute("""SELECT date(sale_date),
                     COUNT(*),
                     SUM(quantity_sold),
                     SUM(profit)
                     FROM pharmacy_sales
                     WHERE pharmacy_id=? AND date(sale_date) BETWEEN ? AND ?
                     GROUP BY date(sale_date) ORDER BY date(sale_date) DESC""",
                  (pid, date_from, date_to))
        daily = [{"date": r[0], "transactions": r[1], "qty": r[2], "profit": round(r[3], 2)}
                 for r in c.fetchall()]
        conn.close()
        return jsonify({
            "from": date_from, "to": date_to,
            "total_transactions": row[0] or 0,
            "total_qty":          row[1] or 0,
            "total_revenue":      round(row[2] or 0, 2),
            "total_cost":         round(row[3] or 0, 2),
            "net_profit":         round(row[4] or 0, 2),
            "top_medicines":      top_medicines,
            "daily":              daily
        })
    except Exception as e:
        return jsonify({"error": str(e)})


# ==========================================
# ADD DEBT → SQLite
# ==========================================
@app.route("/pharmacy/add_debt", methods=["POST"])
def add_debt():
    if not session.get("pharmacy_ok"):
        return jsonify({"success": False, "error": "Not logged in"}), 401
    try:
        pid  = get_username()
        data = request.get_json()
        name = data.get("name", "").strip()
        if not name:
            return jsonify({"success": False, "error": "Name required"})
        conn = sqlite3.connect(DB_PATH)
        c    = conn.cursor()
        init_pharmacy_sql(conn, c)
        c.execute("""INSERT INTO pharmacy_debts
            (pharmacy_id, name, phone, type, amount, description, date, status)
            VALUES (?, ?, ?, ?, ?, ?, ?, 'unpaid')""",
            (pid, name, data.get("phone", ""), data.get("type", "cash"),
             float(data.get("amount", 0)), data.get("description", ""),
             data.get("date", datetime.now().strftime("%Y-%m-%d"))))
        conn.commit()
        conn.close()
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


# ==========================================
# GET DEBTS → SQLite
# ==========================================
@app.route("/pharmacy/debts")
def get_debts():
    if not session.get("pharmacy_ok"):
        return jsonify({"error": "Not logged in"}), 401
    try:
        pid  = get_username()
        conn = sqlite3.connect(DB_PATH)
        c    = conn.cursor()
        init_pharmacy_sql(conn, c)
        c.execute("SELECT * FROM pharmacy_debts WHERE pharmacy_id=? ORDER BY created_at DESC", (pid,))
        rows = c.fetchall()
        c.execute("SELECT COUNT(*) FROM pharmacy_debts WHERE pharmacy_id=?", (pid,))
        total = c.fetchone()[0]
        c.execute("SELECT COUNT(*) FROM pharmacy_debts WHERE pharmacy_id=? AND type='product' AND status!='paid'", (pid,))
        product_count = c.fetchone()[0]
        c.execute("SELECT SUM(amount) FROM pharmacy_debts WHERE pharmacy_id=? AND type='cash' AND status!='paid'", (pid,))
        cash_total = c.fetchone()[0] or 0
        c.execute("SELECT COUNT(*) FROM pharmacy_debts WHERE pharmacy_id=? AND status='paid'", (pid,))
        paid_count = c.fetchone()[0]
        conn.close()
        debts = [{"id": r[0], "name": r[2], "phone": r[3], "type": r[4],
                  "amount": r[5], "description": r[6], "date": r[7],
                  "status": r[8], "created_at": r[9]} for r in rows]
        return jsonify({"debts": debts, "total": total, "product_count": product_count,
                        "cash_total": round(cash_total, 2), "paid_count": paid_count})
    except Exception as e:
        return jsonify({"error": str(e)})


# ==========================================
# MARK DEBT PAID
# ==========================================
@app.route("/pharmacy/debt_paid/<int:debt_id>", methods=["POST"])
def mark_debt_paid(debt_id):
    if not session.get("pharmacy_ok"):
        return jsonify({"success": False, "error": "Not logged in"}), 401
    try:
        conn = sqlite3.connect(DB_PATH)
        c    = conn.cursor()
        c.execute("UPDATE pharmacy_debts SET status='paid' WHERE id=?", (debt_id,))
        conn.commit()
        conn.close()
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


# ==========================================
# DELETE DEBT
# ==========================================
@app.route("/pharmacy/delete_debt/<int:debt_id>", methods=["DELETE"])
def delete_debt(debt_id):
    if not session.get("pharmacy_ok"):
        return jsonify({"success": False, "error": "Not logged in"}), 401
    try:
        conn = sqlite3.connect(DB_PATH)
        c    = conn.cursor()
        c.execute("DELETE FROM pharmacy_debts WHERE id=?", (debt_id,))
        conn.commit()
        conn.close()
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


# ==========================================
# REGISTER PHARMACY (ADMIN)
# ==========================================
@app.route("/admin/register_pharmacy", methods=["POST"])
def admin_register_pharmacy():
    try:
        if not session.get("admin_ok"):
            return jsonify({"success": False, "error": "Unauthorized"}), 401
        data            = request.get_json()
        pharmacy_name   = data.get("pharmacy_name", "").strip()
        phone           = data.get("phone", "").strip()
        monthly_fee     = float(data.get("monthly_fee", 0))
        months          = int(data.get("months", 3))
        username        = data.get("username", "").strip()
        password        = data.get("password", "").strip()
        payment_account = data.get("payment_account", "").strip()
        if not pharmacy_name or not phone or not username or not password:
            return jsonify({"success": False, "error": "Fill all required fields"})
        if payment_account and not payment_account.isdigit():
            return jsonify({"success": False, "error": "Payment account must contain digits only"})
        if payment_account and len(payment_account) > 4:
            return jsonify({"success": False, "error": "Payment account must be 1-4 digits"})
        expiry_date = (datetime.now() + timedelta(days=months * 30)).strftime("%Y-%m-%d")
        total_fee   = round(monthly_fee * months, 2)
        doc_ref = db.collection("pharmacies").add({
            "pharmacy_name": pharmacy_name, "phone": phone,
            "username": username, "password": password,
            "monthly_fee": monthly_fee, "months": months, "total_fee": total_fee,
            "payment_account": payment_account,
            "created_at": datetime.now().isoformat(), "expiry_date": expiry_date, "active": True
        })
        db.collection("pharmacy_users").document(username).set({
            "username": username, "password": password,
            "pharmacy_name": pharmacy_name, "phone": phone,
            "payment_account": payment_account,
            "pharmacy_id": doc_ref[1].id, "created_at": datetime.now().isoformat()
        })
        # Create the pharmacy_product document for this user
        db.collection("pharmacy_product").document(username).set({
            "pharmacy_name": pharmacy_name,
            "username":      username,
            "created_at":    datetime.now().isoformat()
        })
        return jsonify({"success": True, "message": "Pharmacy registered",
                        "expiry_date": expiry_date, "total_fee": total_fee})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


# ==========================================
# RENEW PHARMACY (ADMIN)
# ==========================================
@app.route("/admin/renew_pharmacy/<string:pid>", methods=["POST"])
def admin_renew_pharmacy(pid):
    try:
        if not session.get("admin_ok"):
            return jsonify({"success": False, "error": "Unauthorized"}), 401
        data   = request.get_json()
        months = int(data.get("months", 3))
        ph_ref = db.collection("pharmacies").document(pid)
        ph_doc = ph_ref.get()
        if not ph_doc.exists:
            return jsonify({"success": False, "error": "Pharmacy not found"})
        ph         = ph_doc.to_dict()
        old_expiry = ph.get("expiry_date", datetime.now().strftime("%Y-%m-%d"))
        try:
            base = max(datetime.strptime(old_expiry, "%Y-%m-%d"), datetime.now())
        except:
            base = datetime.now()
        new_expiry = (base + timedelta(days=months * 30)).strftime("%Y-%m-%d")
        ph_ref.update({"expiry_date": new_expiry, "active": True,
                       "last_renewed": datetime.now().isoformat(), "renewed_months": months})
        return jsonify({"success": True, "expiry_date": new_expiry})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


# ==========================================
# DELETE PHARMACY (ADMIN)
# ==========================================
@app.route("/admin/delete_pharmacy/<string:pid>", methods=["DELETE"])
def admin_delete_pharmacy(pid):
    try:
        if not session.get("admin_ok"):
            return jsonify({"success": False, "error": "Unauthorized"}), 401
        ph_doc = db.collection("pharmacies").document(pid).get()
        if not ph_doc.exists:
            return jsonify({"success": False, "error": "Not found"})
        ph       = ph_doc.to_dict()
        username = ph.get("username", "")
        db.collection("pharmacies").document(pid).delete()
        if username:
            try:
                db.collection("pharmacy_users").document(username).delete()
            except:
                pass
            try:
                db.collection("pharmacy_product").document(username).delete()
            except:
                pass
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


# ==========================================
# ✅ ACTIVATE / ❌ DISABLE PHARMACY (ADMIN)
# ==========================================
@app.route("/admin/activate_pharmacy/<string:pid>", methods=["POST"])
def admin_activate_pharmacy(pid):
    if not session.get("admin_ok"):
        return jsonify({"success": False, "error": "Unauthorized"}), 401
    try:
        ph_ref = db.collection("pharmacies").document(pid)
        if not ph_ref.get().exists:
            return jsonify({"success": False, "error": "Not found"})
        ph_ref.update({"active": True, "activated_at": datetime.now().isoformat()})
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


@app.route("/admin/disable_pharmacy/<string:pid>", methods=["POST"])
def admin_disable_pharmacy(pid):
    if not session.get("admin_ok"):
        return jsonify({"success": False, "error": "Unauthorized"}), 401
    try:
        ph_ref = db.collection("pharmacies").document(pid)
        if not ph_ref.get().exists:
            return jsonify({"success": False, "error": "Not found"})
        ph_ref.update({"active": False, "disabled_at": datetime.now().isoformat()})
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


# ==========================================
# CREATE PHARMACY USER (ADMIN)
# ==========================================
@app.route("/admin/create_pharmacy_user", methods=["POST"])
def admin_create_pharmacy_user():
    try:
        if not session.get("admin_ok"):
            return jsonify({"success": False, "error": "Unauthorized ❌"}), 401

        data     = request.get_json()
        username = data.get("username", "").strip()
        password = data.get("password", "").strip()

        if not username or not password:
            return jsonify({"success": False, "error": "Fill all fields ❌"})

        conn = sqlite3.connect(DB_PATH)
        c    = conn.cursor()
        c.execute("""
            CREATE TABLE IF NOT EXISTS pharmacy_users (
                id        INTEGER PRIMARY KEY AUTOINCREMENT,
                username  TEXT UNIQUE,
                password  TEXT,
                role      TEXT DEFAULT 'pharmacist'
            )
        """)
        conn.commit()
        c.execute("SELECT id FROM pharmacy_users WHERE username=?", (username,))
        if c.fetchone():
            conn.close()
            return jsonify({"success": False, "error": f"Username '{username}' already exists ❌"})
        c.execute(
            "INSERT INTO pharmacy_users (username, password) VALUES (?, ?)",
            (username, password)
        )
        conn.commit()
        conn.close()

        db.collection("pharmacy_users").document(username).set({
            "username":   username,
            "password":   password,
            "created_at": datetime.now().isoformat()
        })

        # Auto-create the pharmacy_product document so collection is ready
        db.collection("pharmacy_product").document(username).set({
            "username":   username,
            "created_at": datetime.now().isoformat()
        }, merge=True)

        return jsonify({"success": True, "message": f"User '{username}' created ✅"})

    except Exception as e:
        return jsonify({"success": False, "error": str(e)})

# =========================
# 💳 PAYMENT QR GENERATOR (A4)
# =========================
@app.route("/payment_qr", methods=["GET", "POST"])
def payment_qr():
    if not session.get("admin_ok"):
        return redirect("/admin")

    img_file = ""
    ussd_code = ""

    if request.method == "POST":
        ussd_code = request.form.get("ussd_code", "").strip()

        if not ussd_code:
            return render_template("qr_payment.html", img="", ussd="", error="Fadlan geli USSD code ❌")

        # tel: link — si scan-ku toos ugu wanqali karo dial-ka
        # # waxaa loo beddelaa %23 si telefoonku u aqbalo
        tel_link = "tel:" + ussd_code.replace("#", "%23")

        filename = f"payqr_{int(time.time())}.png"
        qr_folder = os.path.join("static", "qr")
        os.makedirs(qr_folder, exist_ok=True)
        file_path = os.path.join(qr_folder, filename)

        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_H,
            box_size=12,
            border=2
        )
        qr.add_data(tel_link)
        qr.make(fit=True)

        img = qr.make_image(fill_color="black", back_color="white")
        img.save(file_path)

        img_file = filename

    return render_template("qr_payment.html", img=img_file, ussd=ussd_code)


# ==========================================
# 🕌 IMAM MEDIA — a role for Imam University's media team. Password
# lives in Firestore ("imam-media" collection, "password" field on
# any doc there — matches how it was already set up). Once logged in,
# they get an upload area for images/videos/PDFs, all logged as a
# transaction-style history.
# ==========================================
@app.route("/imam_media_login", methods=["GET", "POST"])
def imam_media_login():
    if request.method == "POST":
        password = request.form.get("password", "").strip()
        try:
            for doc in db.collection("imam-media").stream():
                stored_password = (doc.to_dict() or {}).get("password", "")
                if stored_password and password == str(stored_password):
                    session["imam_media_ok"] = True
                    return redirect("/imam_media_dashboard")
        except Exception as e:
            print("Imam media login error:", e)
        return render_template("imam_media_login.html", error="Wrong password")
    return render_template("imam_media_login.html")


@app.route("/imam_media_logout")
def imam_media_logout():
    session.pop("imam_media_ok", None)
    return redirect("/imam_media_login")


@app.route("/imam_media_dashboard")
def imam_media_dashboard():
    if not session.get("imam_media_ok"):
        return redirect("/imam_media_login")
    try:
        uploads = []
        for doc in db.collection("imam_media_uploads").order_by(
                "uploaded_at", direction=firestore.Query.DESCENDING).stream():
            u = doc.to_dict()
            u["id"] = doc.id
            uploads.append(u)
        return render_template("imam_media_dashboard.html", uploads=uploads)
    except Exception as e:
        return f"Imam Media Dashboard Error ❌ {str(e)}"


@app.route("/imam_media_dashboard/report")
def imam_media_report():
    if not session.get("imam_media_ok"):
        return redirect("/imam_media_login")
    try:
        images, videos, documents = [], [], []
        for doc in db.collection("imam_media_uploads").order_by(
                "uploaded_at", direction=firestore.Query.DESCENDING).stream():
            u = doc.to_dict()
            u["id"] = doc.id
            if u.get("file_type") == "image":
                images.append(u)
            elif u.get("file_type") == "video":
                videos.append(u)
            elif u.get("file_type") in ("pdf", "document"):
                documents.append(u)
        return render_template(
            "imam_media_report.html",
            images=images, videos=videos, documents=documents,
            total=len(images) + len(videos) + len(documents)
        )
    except Exception as e:
        return f"Imam Media Report Error ❌ {str(e)}"


@app.route("/imam_media_dashboard/upload", methods=["POST"])
def imam_media_upload():
    if not session.get("imam_media_ok"):
        return jsonify({"success": False, "error": "Unauthorized"}), 401
    try:
        files  = request.files.getlist("file")
        note   = request.form.get("note", "").strip()
        period = request.form.get("period", "daily").strip().lower()
        if period not in ("daily", "weekly", "monthly", "yearly"):
            period = "daily"
        files = [f for f in files if f and f.filename]
        if not files:
            return jsonify({"success": False, "error": "Choose at least one file first"})

        uploaded = []
        skipped = []
        for file in files:
            ext = os.path.splitext(file.filename)[1].lower()
            if ext in (".jpg", ".jpeg", ".png", ".gif", ".webp"):
                file_type = "image"
            elif ext in (".mp4", ".mov", ".avi", ".webm", ".mkv"):
                file_type = "video"
            elif ext in (".pdf", ".ppt", ".pptx", ".doc", ".docx"):
                file_type = "document"
            else:
                skipped.append(file.filename)
                continue

            url = upload_to_firebase_storage(file, folder="imam_media")
            if not url:
                skipped.append(file.filename)
                continue

            record = {
                "filename": secure_filename(file.filename),
                "url": url,
                "file_type": file_type,
                "note": note,
                "period": period,
                "uploaded_at": datetime.now().isoformat()
            }
            doc_ref = db.collection("imam_media_uploads").add(record)
            record["id"] = doc_ref[1].id
            uploaded.append(record)

        if not uploaded:
            return jsonify({"success": False, "error": "None of the selected files could be uploaded (images, videos, PDFs, Word, or PowerPoint files only)"})

        return jsonify({"success": True, "uploaded": uploaded, "skipped": skipped, "count": len(uploaded)})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


@app.route("/imam_media_dashboard/delete/<upload_id>", methods=["DELETE"])
def imam_media_delete(upload_id):
    if not session.get("imam_media_ok"):
        return jsonify({"success": False, "error": "Unauthorized"}), 401
    try:
        doc_ref = db.collection("imam_media_uploads").document(upload_id)
        doc = doc_ref.get()
        if doc.exists:
            url = doc.to_dict().get("url", "")
            delete_from_firebase_storage(url)
            doc_ref.delete()
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 8080))

    init_db()

    socketio.run(
        app,
        host="0.0.0.0",
        port=port
    )